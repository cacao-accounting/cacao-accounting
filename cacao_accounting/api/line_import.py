# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2025 - 2026 William José MORENO Reyes

"""Logic for importing detail lines from external sources."""

from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
import os
import tempfile
import unicodedata
import zipfile
from typing import Any, cast

from flask import Blueprint, jsonify, request
from flask.typing import ResponseReturnValue
from flask_login import current_user, login_required
from sqlalchemy import or_
from openpyxl import load_workbook

from cacao_accounting.api.line_import_registry import LineImportSchemaRegistry
from cacao_accounting.auth.permisos import Permisos
from cacao_accounting.database import (
    ARAPOpenItem,
    Accounts,
    Book,
    CompanyParty,
    CostCenter,
    Currency,
    Entity,
    Item,
    Party,
    PaymentEntry,
    PurchaseInvoice,
    Project,
    SalesInvoice,
    UOM,
    Warehouse,
    database,
)
from cacao_accounting.database.helpers import obtener_id_modulo_por_nombre
from cacao_accounting.document_flow.status import _
from cacao_accounting.document_flow.registry import normalize_doctype

line_import_bp = Blueprint("line_import", __name__)

DOCTYPES_MODULES = {
    "purchase_request": "purchases",
    "purchase_quotation": "purchases",
    "supplier_quotation": "purchases",
    "purchase_order": "purchases",
    "purchase_receipt": "inventory",
    "purchase_invoice": "purchases",
    "sales_request": "sales",
    "sales_quotation": "sales",
    "sales_order": "sales",
    "delivery_note": "inventory",
    "sales_invoice": "sales",
    "journal_entry": "accounting",
    "payment_reconciliation": "cash",
    "bank_transaction": "cash",
    "stock_entry": "inventory",
}

PAYMENT_RECONCILIATION_MODELS = {
    "sales_invoice": SalesInvoice,
    "purchase_invoice": PurchaseInvoice,
    "sales_credit_note": SalesInvoice,
    "sales_debit_note": SalesInvoice,
    "purchase_credit_note": PurchaseInvoice,
    "purchase_debit_note": PurchaseInvoice,
}


@dataclass(frozen=True)
class LineValidationPayload:
    """Normalized payload for line import validation."""

    doctype: str | None
    context: dict[str, Any]
    rows: list[dict[str, Any]]


def _is_decimal(value: Any) -> bool:
    """Check if a value can be converted to Decimal."""
    if value is None or str(value).strip() == "":
        return False
    try:
        return Decimal(str(value)).is_finite()
    except (InvalidOperation, ValueError):
        return False


def _is_date(value: Any) -> bool:
    """Check if a value can be converted to date."""
    if value is None or str(value).strip() == "":
        return False
    try:
        date.fromisoformat(str(value))
        return True
    except ValueError:
        return False


@line_import_bp.route("/api/line-import/schema")
@login_required
def get_schema() -> ResponseReturnValue:
    """Return the import schema for a given doctype."""
    doctype = request.args.get("doctype")
    if not doctype:
        return jsonify({"error": _("Doctype no especificado")}), 400
    schema = LineImportSchemaRegistry.get_schema(doctype)
    if not schema:
        return jsonify({"error": _("Doctype no soportado")}), 400
    return jsonify(schema)


@line_import_bp.route("/api/line-import/open-items")
@login_required
def get_open_items() -> ResponseReturnValue:
    """Return open AP/AR documents for the journal reference selector."""
    from cacao_accounting.contabilidad.arap_allocation import list_cached_open_items, list_open_items

    company = (request.args.get("company") or "").strip()
    party_type = (request.args.get("party_type") or "").strip().lower()
    party = (request.args.get("party") or "").strip()
    reference_type = (request.args.get("reference_type") or "").strip()
    reference_type = {
        "invoice": "sales_invoice" if party_type == "customer" else "purchase_invoice",
        "debit_note": "sales_debit_note" if party_type == "customer" else "purchase_debit_note",
        "credit_note": "sales_credit_note" if party_type == "customer" else "purchase_credit_note",
        "payment": "payment_entry",
        "payment_entry": "payment_entry",
        "journal_entry": "journal_entry",
    }.get(reference_type, reference_type)
    query_text = (request.args.get("q") or "").strip().casefold()
    if not company or party_type not in {"customer", "supplier"} or not party:
        return jsonify({"results": []})
    items = list_cached_open_items(company=company, party_type=party_type, party_id=party)
    if not items:
        items = list_open_items(company=company, party_type=party_type, party_id=party)
    results = []
    for item in items:
        if reference_type and item.document_type != reference_type:
            continue
        document_id = str(item.document_id)
        label = str(item.document_no or document_id)
        if query_text and query_text not in label.casefold() and query_text not in document_id.casefold():
            continue
        selector_id = str(item.open_item_id or document_id)
        results.append(
            {
                "id": selector_id,
                "value": selector_id,
                "document_id": document_id,
                "label": label,
                "display_name": f"{label} ({item.outstanding} {item.currency})",
                "document_type": item.document_type,
                "currency": item.currency,
                "outstanding": str(item.outstanding),
                "economic_line_id": item.economic_line_id,
                "line_number": item.line_number,
            }
        )
    return jsonify({"results": results[:100]})


def _normalize_header(value: Any) -> str:
    """Normalize a spreadsheet header consistently across locales."""
    text = str(value or "").strip().casefold()
    while text.endswith("*"):
        text = text[:-1].rstrip()
    return " ".join(
        unicodedata.normalize("NFD", text)
        .encode("ascii", "ignore")
        .decode("ascii")
        .replace("_", " ")
        .replace("-", " ")
        .split()
    )


def _xlsx_archive_is_safe(path: str) -> tuple[bool, str | None]:
    """Reject compressed XLSX payloads that can exhaust parser resources."""
    try:
        with zipfile.ZipFile(path) as archive:
            entries = archive.infolist()
            if len(entries) > 500:
                return False, _("El archivo contiene demasiados elementos comprimidos.")
            uncompressed = sum(item.file_size for item in entries)
            compressed = max(sum(item.compress_size for item in entries), 1)
            if uncompressed > 64 * 1024 * 1024 or uncompressed / compressed > 100:
                return False, _("El archivo comprimido excede los límites de seguridad.")
            forbidden = ("vbaProject", "externalLink", "embeddings/", "connections.xml")
            if any(any(token in item.filename for token in forbidden) for item in entries):
                return False, _("El archivo contiene macros, enlaces o contenido embebido no permitido.")
    except (OSError, zipfile.BadZipFile):
        return False, _("El archivo XLSX está dañado o no es válido.")
    return True, None


def _xlsx_rows(path: str, schema: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Read an XLSX sheet and return canonical rows plus structural errors."""
    errors: list[dict[str, Any]] = []
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    if len(workbook.worksheets) > 10:
        return [], [{"row": None, "field": "file", "message": _("El archivo contiene demasiadas hojas.")}]
    named = [sheet for sheet in workbook.worksheets if _normalize_header(sheet.title) in {"lineas", "lines"}]
    if named:
        worksheet = named[0]
    elif len(workbook.worksheets) == 1:
        worksheet = workbook.worksheets[0]
    else:
        return [], [{"row": None, "field": "sheet", "message": _("Debe existir una única hoja de líneas importable.")}]
    rows = list(worksheet.iter_rows())
    if not rows:
        return [], [{"row": None, "field": "sheet", "message": _("La hoja no contiene encabezados.")}]
    header_cells = rows[0]
    normalized_headers = [_normalize_header(cell.value) for cell in header_cells]
    if len(normalized_headers) > 100:
        return [], [{"row": 1, "field": "header", "message": _("El archivo contiene demasiadas columnas.")}]
    column_by_header: dict[str, str] = {}
    for column in schema.get("columns", []):
        for candidate in [column.get("key"), column.get("label"), *(column.get("aliases") or [])]:
            column_by_header[_normalize_header(candidate)] = str(column["key"])
    mapped: list[str | None] = []
    seen: set[str] = set()
    for index, header in enumerate(normalized_headers):
        if not header:
            if any(cell.value not in (None, "") for cell in (row[index] for row in rows[1:] if index < len(row))):
                errors.append(
                    {
                        "row": 1,
                        "field": f"column_{index + 1}",
                        "message": _("La columna tiene datos pero no tiene encabezado."),
                    }
                )
            mapped.append(None)
            continue
        key = column_by_header.get(header)
        if key is None:
            errors.append({"row": 1, "field": header, "message": _("La columna no pertenece a la plantilla.")})
        elif key in seen:
            errors.append({"row": 1, "field": key, "message": _("El encabezado está duplicado.")})
        else:
            seen.add(key)
        mapped.append(key)
    required = {str(column["key"]) for column in schema.get("columns", []) if column.get("required")}
    missing = required - seen
    for key in sorted(missing):
        errors.append({"row": 1, "field": key, "message": _("Campo requerido faltante.")})
    canonical_rows: list[dict[str, Any]] = []
    for excel_row, row in enumerate(rows[1:], start=2):
        values = [cell.value for cell in row]
        if not any(value not in (None, "") for value in values):
            continue
        if len(canonical_rows) >= 500:
            errors.append({"row": excel_row, "field": "file", "message": _("Límite máximo de 500 líneas excedido.")})
            break
        item: dict[str, Any] = {"_excel_row": excel_row}
        for index, cell in enumerate(row):
            key = mapped[index] if index < len(mapped) else None
            if not key:
                continue
            value = cell.value
            if cell.data_type == "f" or (isinstance(value, str) and value.lstrip().startswith(("=", "+", "@"))):
                errors.append({"row": excel_row, "field": key, "message": _("Las fórmulas no están permitidas.")})
                continue
            if hasattr(value, "isoformat") and value.__class__.__name__ in {"date", "datetime"}:
                value = value.isoformat()[:10]
            item[key] = value
        canonical_rows.append(item)
    return canonical_rows, errors


@line_import_bp.route("/api/line-import/parse-xlsx", methods=["POST"])
@login_required
def parse_xlsx_lines() -> ResponseReturnValue:
    """Parse an XLSX on the server before client-side insertion."""
    doctype = request.form.get("doctype") or "journal_entry"
    schema = LineImportSchemaRegistry.get_schema(doctype)
    if not schema:
        return jsonify({"valid": False, "errors": [{"field": "doctype", "message": _("Doctype no soportado.")}]}), 400
    uploaded = request.files.get("file")
    if uploaded is None or not uploaded.filename:
        return jsonify({"valid": False, "errors": [{"field": "file", "message": _("Debe seleccionar un archivo XLSX.")}]}), 400
    if not str(uploaded.filename).casefold().endswith(".xlsx"):
        return jsonify({"valid": False, "errors": [{"field": "file", "message": _("Solo se admite formato XLSX.")}]}), 400
    with tempfile.NamedTemporaryFile(prefix="cacao-import-", suffix=".xlsx", delete=False) as temporary:
        uploaded.save(temporary)
        path = temporary.name
    try:
        safe, reason = _xlsx_archive_is_safe(path)
        if not safe:
            return jsonify({"valid": False, "errors": [{"field": "file", "message": reason}]}), 400
        rows, errors = _xlsx_rows(path, schema)
        if errors:
            return jsonify({"valid": False, "rows": [], "errors": errors})
        return jsonify({"valid": True, "rows": rows, "errors": []})
    except (OSError, ValueError, zipfile.BadZipFile) as exc:
        return jsonify({"valid": False, "errors": [{"field": "file", "message": str(exc)}]}), 400
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


@line_import_bp.route("/api/line-import/validate", methods=["POST"])
@login_required
def validate_lines() -> ResponseReturnValue:
    """Validate detail lines before importing them into a document."""
    payload = _get_validation_payload()
    schema, schema_error = _load_import_schema(payload.doctype)
    if schema_error:
        return schema_error
    if schema is None:
        return _invalid_payload_response("doctype", _("Doctype no soportado."), 400)

    company_id, company_error = _validate_company_context(payload.context)
    if company_error:
        return company_error
    if company_id is None:
        return _invalid_payload_response("company_id", _("Compañía no especificada en el contexto."), 400)

    permission_error = _validate_import_permission(str(payload.doctype))
    if permission_error:
        return permission_error
    module_name = DOCTYPES_MODULES.get(str(payload.doctype), "general")
    granular_permission = Permisos(modulo=obtener_id_modulo_por_nombre(module_name), usuario=current_user.id)
    if not granular_permission.administrador and not (
        granular_permission.crear and granular_permission.tiene_acceso_compania(company_id)
    ):
        return _error_response(_("No tiene acceso a la compañía seleccionada."), 403)

    rows_error = _validate_rows_limit(payload.rows)
    if rows_error:
        return rows_error

    errors: list[dict[str, Any]] = []
    validated_rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(payload.rows):
        row_errors, validated_row = _validate_import_row(
            row=row,
            row_no=row_index + 1,
            schema=schema,
            doctype=str(payload.doctype),
            company_id=company_id,
        )
        errors.extend(row_errors)
        validated_rows.append(validated_row)

    if str(payload.doctype) == "journal_entry" and not errors:
        _validate_complete_journal_import(payload.context, validated_rows, company_id, errors)

    return _validation_result_response(errors, validated_rows)


def _validate_complete_journal_import(
    context: dict[str, Any], imported_rows: list[dict[str, Any]], company_id: str, errors: list[dict[str, Any]]
) -> None:
    """Validate imported lines together with the lines already in the voucher."""
    from cacao_accounting.contabilidad.journal_service import (
        _normalize_line,
        _validate_ar_ap_lines,
        _validate_balanced_lines,
        _validate_line_books,
    )

    existing = context.get("existing_lines") or []
    if not isinstance(existing, list) or not all(isinstance(row, dict) for row in existing):
        errors.append(
            {"row": None, "field": "existing_lines", "message": _("Las líneas existentes no tienen un formato válido.")}
        )
        return
    combined_rows = [*existing, *imported_rows]
    try:
        normalized_rows = []
        for row in combined_rows:
            normalized = dict(row)
            if normalized.get("reference_document") and not normalized.get("reference_name"):
                normalized["reference_name"] = normalized["reference_document"]
            normalized_rows.append(normalized)
        lines = [_normalize_line(row, index + 1) for index, row in enumerate(normalized_rows)]
        lines = [line for line in lines if line.account or line.debit or line.credit]
        transaction_currency = str(context.get("transaction_currency") or "") or None
        header_rate = context.get("exchange_rate")
        _validate_balanced_lines(
            company_id,
            lines,
            transaction_currency,
            Decimal(str(header_rate)) if header_rate not in (None, "") else None,
            context.get("books") or None,
        )
        _validate_line_books(company_id, context.get("books") or None, lines)
        _validate_ar_ap_lines(company_id, lines)
    except (ValueError, InvalidOperation) as exc:
        errors.append({"row": None, "field": "voucher", "message": str(exc)})


def _get_validation_payload() -> LineValidationPayload:
    """Read and normalize the validation payload."""
    payload = request.get_json(silent=True)
    if not isinstance(payload, dict):
        return LineValidationPayload(doctype=None, context={}, rows=[])
    context = payload.get("context")
    rows = payload.get("rows")
    return LineValidationPayload(
        doctype=cast(str | None, payload.get("doctype")) if isinstance(payload.get("doctype"), str) else None,
        context=cast(dict[str, Any], context) if isinstance(context, dict) else {},
        rows=cast(list[dict[str, Any]], rows) if isinstance(rows, list) and all(isinstance(row, dict) for row in rows) else [],
    )


def _load_import_schema(doctype: str | None) -> tuple[dict[str, Any] | None, ResponseReturnValue | None]:
    """Load the import schema or return the corresponding HTTP error."""
    if not doctype:
        return None, _error_response(_("Doctype no especificado"), 400)
    schema = LineImportSchemaRegistry.get_schema(doctype)
    if not schema:
        return None, _error_response(_("Doctype no soportado"), 400)
    return schema, None


def _validate_company_context(context: dict[str, Any]) -> tuple[str | None, ResponseReturnValue | None]:
    """Validate that the request context contains an existing company."""
    company_id = context.get("company_id")
    if not company_id:
        return None, _invalid_payload_response("company_id", _("Compañía no especificada en el contexto."), 400)
    company = database.session.query(Entity).filter(or_(Entity.id == company_id, Entity.code == company_id)).first()
    if not company:
        return None, _error_response(_("La compañía seleccionada no existe."), 400)
    return str(company.code), None


def _validate_import_permission(doctype: str) -> ResponseReturnValue | None:
    """Validate the current user's import permission for the document module."""
    module_name = DOCTYPES_MODULES.get(doctype, "general")
    permission = Permisos(modulo=obtener_id_modulo_por_nombre(module_name), usuario=current_user.id)
    if not permission.autorizado or not permission.importar:
        return _error_response(_("No tiene permisos para importar en este módulo."), 403)
    return None


def _validate_rows_limit(rows: list[dict[str, Any]]) -> ResponseReturnValue | None:
    """Validate import row count limits."""
    if not rows:
        return _invalid_payload_response("rows", _("Debe importar al menos una línea."))
    if len(rows) > 500:
        return _invalid_payload_response("rows", _("Límite máximo de 500 líneas excedido."))
    return None


def _validate_import_row(
    *,
    row: dict[str, Any],
    row_no: int,
    schema: dict[str, Any],
    doctype: str,
    company_id: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Validate and enrich one imported line."""
    errors: list[dict[str, Any]] = []
    validated_row = row.copy()
    _validate_schema_columns(row, row_no, schema, errors)
    _enrich_and_validate_master_data(row, validated_row, row_no, company_id, errors)
    if doctype == "journal_entry":
        _validate_journal_entry_row(row, row_no, errors)
    elif doctype == "payment_reconciliation":
        _validate_payment_reconciliation_row(row, row_no, company_id, errors)
    return errors, validated_row


def _validate_schema_columns(
    row: dict[str, Any],
    row_no: int,
    schema: dict[str, Any],
    errors: list[dict[str, Any]],
) -> None:
    """Validate required fields and primitive schema types for a row."""
    for column in schema["columns"]:
        key = column["key"]
        value = row.get(key)
        is_empty = value is None or str(value).strip() == ""
        if column["required"] and is_empty:
            errors.append({"row": row_no, "field": key, "message": _("Campo requerido faltante.")})
            continue
        if not is_empty:
            _validate_typed_value(key, value, str(column["type"]), row_no, errors)


def _validate_typed_value(
    key: str,
    value: Any,
    column_type: str,
    row_no: int,
    errors: list[dict[str, Any]],
) -> None:
    """Validate a non-empty value according to its schema type."""
    if column_type == "decimal":
        _validate_decimal_value(key, value, row_no, errors)
    elif column_type == "boolean":
        normalized = str(value).strip().lower()
        if normalized not in {"1", "0", "true", "false", "yes", "no", "si", "sí"}:
            errors.append({"row": row_no, "field": key, "message": _("Valor booleano inválido.")})
    elif column_type == "date" and not _is_date(value):
        errors.append({"row": row_no, "field": key, "message": _("Formato de fecha inválido (AAAA-MM-DD).")})


def _validate_decimal_value(
    key: str,
    value: Any,
    row_no: int,
    errors: list[dict[str, Any]],
) -> None:
    """Validate decimal format and field-specific numeric constraints."""
    if not _is_decimal(value):
        errors.append({"row": row_no, "field": key, "message": _("Valor decimal inválido.")})
        return
    _validate_decimal_constraints(key, Decimal(str(value)), row_no, errors)


def _validate_decimal_constraints(
    key: str,
    decimal_value: Decimal,
    row_no: int,
    errors: list[dict[str, Any]],
) -> None:
    """Validate business constraints for known decimal fields."""
    if key == "quantity" and decimal_value <= 0:
        errors.append({"row": row_no, "field": key, "message": _("La cantidad debe ser mayor que cero.")})
    if key == "rate" and decimal_value < 0:
        errors.append({"row": row_no, "field": key, "message": _("El precio/tasa no puede ser negativo.")})
    if key in {"debit", "credit"} and decimal_value < 0:
        errors.append({"row": row_no, "field": key, "message": _("El débito/crédito no puede ser negativo.")})
    if key == "exchange_rate" and decimal_value <= 0:
        errors.append({"row": row_no, "field": key, "message": _("El tipo de cambio debe ser mayor que cero.")})
    if key in {"reference_exchange_rate", "payment_exchange_rate"} and decimal_value <= 0:
        errors.append({"row": row_no, "field": key, "message": _("El tipo de cambio debe ser mayor que cero.")})
    if key in {"discount_amount", "gain_loss_amount"} and decimal_value < 0:
        errors.append({"row": row_no, "field": key, "message": _("El importe no puede ser negativo.")})


def _enrich_and_validate_master_data(
    row: dict[str, Any],
    validated_row: dict[str, Any],
    row_no: int,
    company_id: str,
    errors: list[dict[str, Any]],
) -> None:
    """Validate master-data references and enrich the imported row."""
    _validate_item_reference(row, validated_row, row_no, errors)
    _validate_uom_reference(row, row_no, errors)
    _validate_account_reference(row, validated_row, row_no, company_id, errors)
    _validate_cost_center_reference(row, row_no, company_id, errors)
    _validate_project_reference(row, row_no, company_id, errors)
    _validate_warehouse_reference(row, row_no, company_id, errors)
    if row.get("book"):
        _validate_book_reference(row, validated_row, row_no, company_id, errors)
    if row.get("currency"):
        _validate_currency_reference(row, validated_row, row_no, errors)
    _validate_party_reference(row, validated_row, row_no, company_id, errors)
    _validate_open_item_reference(row, validated_row, row_no, company_id, errors)


def _validate_item_reference(
    row: dict[str, Any],
    validated_row: dict[str, Any],
    row_no: int,
    errors: list[dict[str, Any]],
) -> None:
    """Validate and enrich an item reference."""
    if not row.get("item_code"):
        return
    item = database.session.query(Item).filter_by(code=row["item_code"]).first()
    if not item:
        errors.append({"row": row_no, "field": "item_code", "message": _("El artículo no existe.")})
        return
    validated_row["item_name"] = item.name
    validated_row["item_id"] = item.id


def _validate_uom_reference(row: dict[str, Any], row_no: int, errors: list[dict[str, Any]]) -> None:
    """Validate a UOM reference."""
    if row.get("uom") and not database.session.query(UOM).filter_by(code=row["uom"]).first():
        errors.append({"row": row_no, "field": "uom", "message": _("La unidad de medida no existe.")})


def _validate_account_reference(
    row: dict[str, Any],
    validated_row: dict[str, Any],
    row_no: int,
    company_id: str,
    errors: list[dict[str, Any]],
) -> None:
    """Validate an account reference scoped by company."""
    account_value = str(row.get("account") or "").strip()
    if not account_value:
        return
    account = database.session.get(Accounts, account_value)
    if account is None:
        account = database.session.query(Accounts).filter_by(code=account_value, entity=company_id).first()
    if account is None or account.entity != company_id or account.group or account.enabled is False or account.active is False:
        errors.append({"row": row_no, "field": "account", "message": _("La cuenta contable no existe o no está activa.")})
    else:
        # Return canonical ID/code values for downstream adapters.
        validated_row["account"] = str(account.code)


def _validate_cost_center_reference(
    row: dict[str, Any],
    row_no: int,
    company_id: str,
    errors: list[dict[str, Any]],
) -> None:
    """Validate a cost center reference scoped by company."""
    cost_center = row.get("cost_center")
    if cost_center and not database.session.query(CostCenter).filter_by(code=cost_center, entity=company_id).first():
        errors.append({"row": row_no, "field": "cost_center", "message": _("El centro de costo no existe.")})


def _validate_project_reference(
    row: dict[str, Any],
    row_no: int,
    company_id: str,
    errors: list[dict[str, Any]],
) -> None:
    """Validate a project reference scoped by company."""
    if row.get("project") and not database.session.query(Project).filter_by(code=row["project"], entity=company_id).first():
        errors.append({"row": row_no, "field": "project", "message": _("El proyecto no existe.")})


def _validate_warehouse_reference(
    row: dict[str, Any],
    row_no: int,
    company_id: str,
    errors: list[dict[str, Any]],
) -> None:
    """Validate a warehouse reference scoped by company."""
    warehouse = row.get("warehouse")
    if warehouse and not database.session.query(Warehouse).filter_by(code=warehouse, company=company_id).first():
        errors.append({"row": row_no, "field": "warehouse", "message": _("La bodega no existe.")})


def _normalize_party_type(value: Any) -> str | None:
    """Normalize localized customer/supplier labels."""
    normalized = str(value or "").strip().casefold()
    return {
        "customer": "customer",
        "cliente": "customer",
        "customers": "customer",
        "supplier": "supplier",
        "proveedor": "supplier",
        "suppliers": "supplier",
    }.get(normalized)


def _validate_book_reference(
    row: dict[str, Any], validated_row: dict[str, Any], row_no: int, company_id: str, errors: list[dict[str, Any]]
) -> None:
    """Validate an optional line-specific accounting book."""
    value = str(row.get("book") or "").strip()
    book = database.session.query(Book).filter(Book.entity == company_id, (Book.id == value) | (Book.code == value)).first()
    if book is None or book.status != "activo":
        errors.append({"row": row_no, "field": "book", "message": _("El libro contable no existe o no está activo.")})
    else:
        validated_row["book"] = str(book.code)


def _validate_currency_reference(
    row: dict[str, Any], validated_row: dict[str, Any], row_no: int, errors: list[dict[str, Any]]
) -> None:
    """Validate an explicitly supplied currency code."""
    value = str(row.get("currency") or "").strip()
    currency = database.session.query(Currency).filter_by(code=value).first()
    if currency is None or currency.active is False:
        errors.append({"row": row_no, "field": "currency", "message": _("La moneda no existe o está inactiva.")})
    else:
        validated_row["currency"] = str(currency.code)


def _validate_party_reference(
    row: dict[str, Any], validated_row: dict[str, Any], row_no: int, company_id: str, errors: list[dict[str, Any]]
) -> None:
    """Validate and canonicalize a customer/supplier reference."""
    raw_type = row.get("party_type")
    raw_party = str(row.get("party") or "").strip()
    if not raw_type and not raw_party:
        return
    party_type = _normalize_party_type(raw_type)
    if party_type is None:
        errors.append({"row": row_no, "field": "party_type", "message": _("El tipo de tercero debe ser Cliente o Proveedor.")})
        return
    if not raw_party:
        errors.append({"row": row_no, "field": "party", "message": _("Debe indicar el tercero seleccionado.")})
        return
    party = database.session.get(Party, raw_party) or database.session.query(Party).filter_by(code=raw_party).first()
    if party is None or party.is_active is False:
        errors.append({"row": row_no, "field": "party", "message": _("El tercero no existe o está inactivo.")})
        return
    if party_type == "customer" and not party.is_customer:
        errors.append({"row": row_no, "field": "party", "message": _("El tercero no está configurado como Cliente.")})
    if party_type == "supplier" and not party.is_supplier:
        errors.append({"row": row_no, "field": "party", "message": _("El tercero no está configurado como Proveedor.")})
    company_party = database.session.query(CompanyParty).filter_by(company=company_id, party_id=party.id).first()
    if company_party is None or company_party.is_active is False:
        errors.append(
            {"row": row_no, "field": "party", "message": _("El tercero no está activo en la compañía seleccionada.")}
        )
        return
    validated_row["party_type"] = party_type
    validated_row["party"] = str(party.id)


def _normalize_reference_type(value: Any) -> str | None:
    """Normalize user-facing reference labels to stable document types."""
    normalized = str(value or "").strip().casefold()
    return {
        "invoice": "invoice",
        "factura": "invoice",
        "sales invoice": "sales_invoice",
        "purchase invoice": "purchase_invoice",
        "nota de débito": "debit_note",
        "nota de debito": "debit_note",
        "debit note": "debit_note",
        "sales debit note": "sales_debit_note",
        "purchase debit note": "purchase_debit_note",
        "nota de crédito": "credit_note",
        "nota de credito": "credit_note",
        "credit note": "credit_note",
        "sales credit note": "sales_credit_note",
        "purchase credit note": "purchase_credit_note",
        "pago": "payment_entry",
        "payment": "payment_entry",
        "payment entry": "payment_entry",
        "entrada de pago": "payment_entry",
        "journal entry": "journal_entry",
        "comprobante contable": "journal_entry",
        "otro comprobante contable": "journal_entry",
        "other journal entry": "journal_entry",
    }.get(normalized, normalized or None)


def _validate_open_item_reference(
    row: dict[str, Any], validated_row: dict[str, Any], row_no: int, company_id: str, errors: list[dict[str, Any]]
) -> None:
    """Resolve an optional AP/AR reference without applying it during import."""
    reference_type = _normalize_reference_type(row.get("reference_type"))
    reference_document = str(row.get("reference_document") or row.get("reference_name") or "").strip()
    if not reference_type and not reference_document:
        return
    if not reference_type or not reference_document:
        errors.append(
            {"row": row_no, "field": "reference_type", "message": _("Tipo y documento de referencia deben indicarse juntos.")}
        )
        return
    party_type = validated_row.get("party_type") or _normalize_party_type(row.get("party_type"))
    if reference_type == "invoice":
        reference_type = "sales_invoice" if party_type == "customer" else "purchase_invoice"
    elif reference_type == "debit_note":
        reference_type = "sales_debit_note" if party_type == "customer" else "purchase_debit_note"
    elif reference_type == "credit_note":
        reference_type = "sales_credit_note" if party_type == "customer" else "purchase_credit_note"
    query = database.session.query(ARAPOpenItem).filter(
        ARAPOpenItem.company == company_id,
        ARAPOpenItem.unallocated_amount > 0,
    )
    query = query.filter(ARAPOpenItem.document_type == reference_type)
    query = query.filter(
        (ARAPOpenItem.id == reference_document)
        | (ARAPOpenItem.document_id == reference_document)
        | (ARAPOpenItem.document_no == reference_document)
    )
    party_id = validated_row.get("party") or row.get("party")
    if party_id:
        query = query.filter(ARAPOpenItem.party_id == party_id)
    if party_type:
        query = query.filter(ARAPOpenItem.party_type == party_type)
    line_hint = str(row.get("reference_line") or "").strip()
    if line_hint:
        query = query.filter(
            ARAPOpenItem.line_number == int(line_hint) if line_hint.isdigit() else ARAPOpenItem.economic_line_id == line_hint
        )
    matches = query.all()
    if not matches:
        from cacao_accounting.contabilidad.arap_allocation import list_open_items

        ledger_matches = [
            item
            for item in list_open_items(company=company_id, party_type=party_type, party_id=party_id)
            if item.document_type == reference_type
            and (item.document_id == reference_document or item.document_no == reference_document)
            and item.outstanding > 0
        ]
        if len(ledger_matches) == 1:
            item = ledger_matches[0]
            validated_row["reference_type"] = reference_type
            validated_row["reference_document"] = str(item.document_no or item.document_id)
            return
        if len(ledger_matches) > 1:
            errors.append(
                {
                    "row": row_no,
                    "field": "reference_line",
                    "message": _("La referencia es ambigua; indique la línea del documento."),
                }
            )
            return
    if len(matches) == 0:
        errors.append(
            {"row": row_no, "field": "reference_document", "message": _("El documento abierto no existe o no tiene saldo.")}
        )
    elif len(matches) > 1:
        errors.append(
            {
                "row": row_no,
                "field": "reference_line",
                "message": _("La referencia es ambigua; indique la línea del documento."),
            }
        )
    else:
        validated_row["reference_type"] = reference_type
        validated_row["reference_open_item_id"] = str(matches[0].id)
        validated_row["reference_document"] = str(matches[0].document_no or matches[0].document_id)


def _validate_journal_entry_row(row: dict[str, Any], row_no: int, errors: list[dict[str, Any]]) -> None:
    """Validate debit and credit rules for imported journal lines."""
    debit_amount = Decimal(str(row.get("debit") or 0)) if _is_decimal(row.get("debit")) else Decimal(0)
    credit_amount = Decimal(str(row.get("credit") or 0)) if _is_decimal(row.get("credit")) else Decimal(0)
    if debit_amount == credit_amount == 0:
        errors.append({"row": row_no, "field": "debit", "message": _("Debe especificar un monto en Débito o Crédito.")})
    if debit_amount != 0 and credit_amount != 0:
        errors.append(
            {
                "row": row_no,
                "field": "debit",
                "message": _("No puede especificar Débito y Crédito en la misma línea."),
            }
        )


def _append_payment_reconciliation_error(errors: list[dict[str, Any]], row_no: int, field: str, message: str) -> None:
    """Append a normalized payment reconciliation validation error."""
    errors.append({"row": row_no, "field": field, "message": message})


def _get_payment_reconciliation_payment(
    row: dict[str, Any], company_id: str, row_no: int, errors: list[dict[str, Any]]
) -> Any:
    """Load the payment and report when it is not valid for the company."""
    payment_id = str(row.get("payment_id") or "").strip()
    payment = database.session.get(PaymentEntry, payment_id)
    if payment is None or payment.company != company_id or payment.docstatus != 1:
        _append_payment_reconciliation_error(
            errors, row_no, "payment_id", _("El pago no existe para la compañía seleccionada.")
        )
    return payment


def _get_payment_reconciliation_reference(
    row: dict[str, Any], company_id: str, row_no: int, errors: list[dict[str, Any]]
) -> tuple[Any, str]:
    """Load the referenced document using its normalized document type."""
    reference_type = normalize_doctype(str(row.get("reference_type") or ""))
    model = PAYMENT_RECONCILIATION_MODELS.get(reference_type)
    if model is None:
        _append_payment_reconciliation_error(errors, row_no, "reference_type", _("Tipo de documento no conciliable."))
        return None, reference_type
    reference_id = str(row.get("reference_id") or "").strip()
    reference = database.session.get(model, reference_id)
    if reference is None or reference.docstatus != 1 or reference.company != company_id:
        _append_payment_reconciliation_error(
            errors, row_no, "reference_id", _("El documento de referencia no existe o no está aprobado.")
        )
        return None, reference_type
    return reference, reference_type


def _validate_payment_reconciliation_party(
    payment: Any, reference: Any, reference_type: str, row_no: int, errors: list[dict[str, Any]]
) -> None:
    """Ensure the payment and referenced document belong to the same party."""
    expected_party_type = "supplier" if reference_type.startswith("purchase_") else "customer"
    party_attribute = "supplier_id" if expected_party_type == "supplier" else "customer_id"
    expected_party_id = getattr(reference, party_attribute, None)
    if payment.party_type != expected_party_type or str(payment.party_id) != str(expected_party_id):
        _append_payment_reconciliation_error(
            errors, row_no, "reference_id", _("El documento de referencia no coincide con el tercero del pago.")
        )


def _payment_reconciliation_rate(
    row: dict[str, Any], payment: Any, reference: Any, row_no: int, errors: list[dict[str, Any]]
) -> Decimal:
    """Resolve the exchange rate required to compare payment and document currencies."""
    document_currency = getattr(reference, "transaction_currency", None)
    payment_currency = getattr(payment, "currency", None)
    if not document_currency or not payment_currency or document_currency == payment_currency:
        return Decimal("1")
    raw_rate = row.get("payment_exchange_rate")
    if not (_is_decimal(raw_rate) and Decimal(str(raw_rate)) > 0):
        _append_payment_reconciliation_error(
            errors, row_no, "payment_exchange_rate", _("Se requiere una tasa para monedas distintas.")
        )
        return Decimal("1")
    return Decimal(str(raw_rate))


def _validate_payment_reconciliation_amounts(
    row: dict[str, Any], payment: Any, reference: Any, rate: Decimal, row_no: int, errors: list[dict[str, Any]]
) -> None:
    """Validate allocated amount against document and payment outstanding balances."""
    from cacao_accounting.document_flow.payment import compute_outstanding_amount, compute_payment_unallocated_amount

    allocated = Decimal(str(row["allocated_amount"])) if _is_decimal(row.get("allocated_amount")) else Decimal("0")
    discount = Decimal(str(row["discount_amount"])) if _is_decimal(row.get("discount_amount")) else Decimal("0")
    gain_loss = Decimal(str(row["gain_loss_amount"])) if _is_decimal(row.get("gain_loss_amount")) else Decimal("0")
    if allocated <= 0:
        return
    if allocated > compute_outstanding_amount(reference) + Decimal("0.01"):
        _append_payment_reconciliation_error(
            errors, row_no, "allocated_amount", _("El monto aplicado excede el saldo pendiente del documento.")
        )
    consumed = max(allocated - discount - gain_loss, Decimal("0")) * rate
    if consumed > compute_payment_unallocated_amount(payment) + Decimal("0.01"):
        _append_payment_reconciliation_error(
            errors, row_no, "allocated_amount", _("El monto aplicado excede el saldo disponible del pago.")
        )


def _validate_payment_reconciliation_row(
    row: dict[str, Any], row_no: int, company_id: str, errors: list[dict[str, Any]]
) -> None:
    """Validate payment/document identity before reconciliation is applied."""
    payment = _get_payment_reconciliation_payment(row, company_id, row_no, errors)
    reference, reference_type = _get_payment_reconciliation_reference(row, company_id, row_no, errors)
    if reference is not None and payment is not None:
        _validate_payment_reconciliation_party(payment, reference, reference_type, row_no, errors)
        rate = _payment_reconciliation_rate(row, payment, reference, row_no, errors)
        _validate_payment_reconciliation_amounts(row, payment, reference, rate, row_no, errors)
    allocated = row.get("allocated_amount")
    if _is_decimal(allocated) and Decimal(str(allocated)) <= 0:
        _append_payment_reconciliation_error(
            errors, row_no, "allocated_amount", _("El monto aplicado debe ser mayor que cero.")
        )


def _error_response(message: str, status_code: int) -> ResponseReturnValue:
    """Build a simple API error response."""
    return jsonify({"error": message}), status_code


def _invalid_payload_response(field: str, message: str, status_code: int = 200) -> ResponseReturnValue:
    """Build a validation response for request-level payload errors."""
    return jsonify({"valid": False, "errors": [{"row": None, "field": field, "message": message}]}), status_code


def _validation_result_response(
    errors: list[dict[str, Any]],
    validated_rows: list[dict[str, Any]],
) -> ResponseReturnValue:
    """Build the final line validation result response."""
    return jsonify({"valid": len(errors) == 0, "rows": validated_rows if len(errors) == 0 else [], "errors": errors})
