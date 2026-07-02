# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2025 - 2026 William José Moreno Reyes

"""Reportes operativos de subledgers, aging, Kardex y reconciliaciones."""

from __future__ import annotations

import csv
from dataclasses import replace
from datetime import date
from decimal import Decimal
from decimal import DecimalException
from io import BytesIO, StringIO
from typing import Any, cast

from flask import Blueprint, flash, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from cacao_accounting.auth.permisos import Permisos
from cacao_accounting.database import Accounts, AccountingPeriod, Book, Entity, UserFormPreference, database
from cacao_accounting.database.helpers import obtener_id_modulo_por_nombre
from cacao_accounting.decorators import modulo_activo, verifica_acceso
from cacao_accounting.form_preferences import get_form_preference, reset_form_preference, save_form_preference
from cacao_accounting.reportes.services import (
    AgingFilters,
    BankingFilters,
    FinancialReportFilters,
    KardexFilters,
    OperationalReportFilters,
    SubledgerFilters,
    PaginatedReport,
    get_account_movement_detail,
    get_account_summary_report,
    get_aging_report,
    get_ar_ap_subledger,
    get_batch_report,
    get_bank_balance_summary,
    get_bank_movement_detail,
    get_balance_sheet_report,
    get_gross_margin,
    get_income_statement_report,
    get_inventory_existence,
    get_inventory_valuation,
    get_kardex,
    get_purchases_by_item,
    get_purchases_by_supplier,
    get_reconciliation_report,
    get_sales_by_customer,
    get_sales_by_item,
    get_serial_report,
    get_stock_balance,
    get_trial_balance_report,
)
from cacao_accounting.version import APPNAME

try:  # pragma: no cover - fallback defensivo para contextos sin Flask-Babel inicializado.
    from flask_babel import gettext as _babel_gettext
except ImportError:  # pragma: no cover

    def _(value: str) -> str:
        return value

else:

    def _(value: str) -> str:
        try:
            return _babel_gettext(value)
        except (KeyError, RuntimeError):
            return value


reportes = Blueprint("reportes", __name__, template_folder="templates")

REPORT_TABLE_HTML = "reportes/report_table.html"

_COLUMN_LABELS = {
    "posting_date": "Posting Date",
    "accounting_period": "Period",
    "document_no": "Voucher",
    "voucher_type": "Type",
    "account_code": "Account",
    "account_name": "Account Name",
    "account_type": "Account Type",
    "classification": "Section",
    "debit": "Debit",
    "credit": "Credit",
    "running_balance": "Final Balance",
    "currency": "Currency",
    "ledger": "Ledger",
    "company": "Company",
    "opening_balance": "Opening Balance",
    "ending_balance": "Final Balance",
    "cost_center": "Cost Center",
    "unit": "Unit",
    "project": "Project",
    "party_type": "Party Type",
    "party_id": "Party",
    "created_by": "User",
    "created": "Creation Date",
    "created_at": "Creation Date",
    "movement_count": "Movements",
    "first_movement": "First Movement",
    "last_movement": "Last Movement",
    "line_comment": "Reference",
    "reference_type": "Reference Type",
    "is_reversal": "Is Reversal",
    "reversal_of": "Reversal Of",
    "status": "Status",
    "voucher_status": "Status",
    "section": "Section",
    "amount": "Amount",
    "bank_account": "Bank Account",
    "party_name": "Party",
    "payment_type": "Payment Type",
    "incoming_amount": "Incoming Amount",
    "outgoing_amount": "Outgoing Amount",
    "receipts_amount": "Receipts",
    "payments_amount": "Payments",
    "account_no": "Account Number",
    "item_name": "Item Name",
    "balance_qty": "Balance Qty",
    "incoming_qty": "Incoming Qty",
    "outgoing_qty": "Outgoing Qty",
    "value_change": "Value Change",
    "original_amount": "Original Amount",
    "paid_amount": "Paid Amount",
    "outstanding_amount": "Outstanding Amount",
    "days": "Days",
    "bucket": "Bucket",
    "remarks": "Remarks",
}
_MONEY_COLUMNS = {
    "debit",
    "credit",
    "difference",
    "opening_balance",
    "ending_balance",
    "running_balance",
    "amount",
    "assets",
    "liabilities",
    "equity",
    "period_profit",
    "income",
    "cost",
    "expense",
    "gross_profit",
    "net_profit",
    "incoming_amount",
    "outgoing_amount",
    "receipts_amount",
    "payments_amount",
    "original_amount",
    "paid_amount",
    "outstanding_amount",
    "value_change",
    "stock_value",
    "remaining_stock_value",
}
_RIGHT_ALIGN_COLUMNS = _MONEY_COLUMNS | {"level", "incoming_qty", "outgoing_qty", "balance_qty", "actual_qty", "days"}
_ALWAYS_VISIBLE_COLUMNS = {
    "debit",
    "credit",
    "difference",
    "account_code",
    "account_name",
    "section",
    "amount",
    "opening_balance",
    "ending_balance",
}
_EMPTY_CELL_VALUE = "—"
_FINANCIAL_FILTER_FIELDS = (
    "company",
    "ledger",
    "accounting_period",
    "voucher_number",
    "account_code",
    "account_from",
    "account_to",
    "cost_center_code",
    "unit_code",
    "project_code",
    "party_type",
    "party_id",
    "voucher_type",
    "status",
    "include_running_balance",
    "page_size",
    "sort_by",
    "sort_dir",
    "group_by",
)


def _to_decimal_or_zero(value: object) -> Decimal:
    try:
        return value if isinstance(value, Decimal) else Decimal(str(value))
    except DecimalException:
        return Decimal("0")


def _format_number(value: object) -> str:
    try:
        amount = value if isinstance(value, Decimal) else Decimal(str(value))
    except DecimalException:
        return _EMPTY_CELL_VALUE
    formatted = f"{abs(amount):,.2f}"
    return f"({formatted})" if amount < 0 else formatted


def _column_label(column: str, ledger_currency: str | None) -> str:
    label = _(_COLUMN_LABELS.get(column, column.replace("_", " ").title()))
    if column in _MONEY_COLUMNS and ledger_currency:
        return f"{label} ({ledger_currency})"
    return label


def _format_cell(column: str, value: object, ledger_currency: str | None) -> str:
    if value is None or value == "":
        return _EMPTY_CELL_VALUE
    if column in _MONEY_COLUMNS:
        return _format_number(value)
    if column == "posting_date" and isinstance(value, date):
        return value.isoformat()
    if column == "voucher_status":
        return _("Cancelado") if str(value).lower() == "cancelled" else _("Contabilizado")
    if column == "section":
        section_labels = {
            "assets": _("ACTIVOS"),
            "liabilities": _("PASIVOS"),
            "equity": _("PATRIMONIO"),
            "income": _("INGRESOS"),
            "cost": _("COSTOS"),
            "expense": _("GASTOS"),
            "gross_profit": _("UTILIDAD BRUTA"),
            "net_profit": _("UTILIDAD NETA"),
        }
        return section_labels.get(str(value), str(value))
    return str(value)


def _build_context_summary(report, report_filters: FinancialReportFilters) -> dict[str, str]:
    ledger_label = report_filters.ledger or "—"
    if report_filters.ledger and report.ledger_currency:
        ledger_label = f"{report_filters.ledger} ({report.ledger_currency})"
    status_value = report_filters.status or "submitted"
    status_label = _("Cancelado") if status_value == "cancelled" else _("Contabilizado")
    return {
        "company": report_filters.company,
        "ledger": ledger_label,
        "period": report_filters.accounting_period or "—",
        "status": status_label,
        "records": str(report.total_rows),
    }


def _is_report_balanced(raw_totals: dict[str, Decimal]) -> bool:
    difference = raw_totals.get("difference")
    if difference is None:
        return False
    try:
        return Decimal(str(difference)) == Decimal("0")
    except DecimalException:
        return False


def _report_form_key(report_code: str) -> str:
    return f"reports.financial.{report_code}"


def _load_report_view_options(report_code: str) -> list[str]:
    preferences = (
        database.session.execute(
            database.select(UserFormPreference.view_key)
            .filter_by(user_id=str(current_user.id), form_key=_report_form_key(report_code))
            .order_by(UserFormPreference.view_key.asc())
        )
        .scalars()
        .all()
    )
    views = ["default"]
    views.extend([view for view in preferences if view != "default"])
    return views


def _extract_filter_payload() -> dict[str, str]:
    payload: dict[str, str] = {}
    for key in _FINANCIAL_FILTER_FIELDS:
        value = request.args.get(key)
        if value:
            payload[key] = value
    visible_columns = request.args.getlist("visible_columns")
    if visible_columns:
        payload["visible_columns"] = ",".join(visible_columns)
    return payload


def _restore_filters_from_view(filters: FinancialReportFilters, report_code: str, view_key: str) -> FinancialReportFilters:
    preference = get_form_preference(str(current_user.id), _report_form_key(report_code), view_key)
    payload = preference.get("filters", {})
    if not isinstance(payload, dict):
        return filters
    page_size = _safe_page_size(payload.get("page_size"), filters.page_size)
    return cast(
        FinancialReportFilters,
        replace(
            filters,
            company=str(payload.get("company") or filters.company),
            ledger=_str_or_none(payload.get("ledger")),
            accounting_period=_str_or_none(payload.get("accounting_period")),
            voucher_number=_str_or_none(payload.get("voucher_number")),
            account_code=_str_or_none(payload.get("account_code")),
            account_from=_str_or_none(payload.get("account_from")),
            account_to=_str_or_none(payload.get("account_to")),
            cost_center_code=_str_or_none(payload.get("cost_center_code")),
            unit_code=_str_or_none(payload.get("unit_code")),
            project_code=_str_or_none(payload.get("project_code")),
            party_type=_str_or_none(payload.get("party_type")),
            party_id=_str_or_none(payload.get("party_id")),
            voucher_type=_str_or_none(payload.get("voucher_type")),
            status=str(payload.get("status") or filters.status or "submitted"),
            include_running_balance=str(payload.get("include_running_balance") or "").lower() in {"1", "true", "yes", "on"},
            page_size=page_size,
            sort_by=str(payload.get("sort_by") or filters.sort_by),
            sort_dir=str(payload.get("sort_dir") or filters.sort_dir),
            page=1,
        ),
    )


def _safe_page_size(raw: object, default: int) -> int:
    try:
        if raw is None or isinstance(raw, str):
            value = default
        else:
            value = int(raw)  # type: ignore[call-overload]
        return max(value, 1)
    except (TypeError, ValueError):
        return max(default, 1)


def _str_or_none(value: object) -> str | None:
    return str(value or "") or None


def _handle_saved_view_action(report_code: str, filters: FinancialReportFilters) -> tuple[FinancialReportFilters, str]:
    view_key = (request.args.get("saved_view") or "default").strip() or "default"
    action = request.args.get("view_action")
    if action == "save" and view_key:
        payload = {
            "schema_version": 1,
            "filters": _extract_filter_payload(),
            "columns": [
                {
                    "field": column,
                    "label": column,
                    "visible": True,
                    "width": 1,
                    "required": False,
                }
                for column in request.args.getlist("visible_columns")
            ],
        }
        save_form_preference(
            user_id=str(current_user.id),
            form_key=_report_form_key(report_code),
            view_key=view_key,
            payload=payload,
        )
        flash(_("Vista guardada correctamente."), "success")
    elif action == "reset" and view_key != "default":
        reset_form_preference(str(current_user.id), _report_form_key(report_code), view_key)
        flash(_("Vista eliminada correctamente."), "warning")
        view_key = "default"
    elif action == "apply" and view_key != "default":
        filters = _restore_filters_from_view(filters, report_code, view_key)
    return filters, view_key


def _resolve_view_context(report_code: str, filters: FinancialReportFilters) -> tuple[FinancialReportFilters, str, list[str]]:
    resolved_filters, selected_view = _handle_saved_view_action(report_code, filters)
    return resolved_filters, selected_view, _load_report_view_options(report_code)


def _preferred_columns_from_view(report_code: str, view_key: str) -> list[str]:
    if view_key == "default":
        return []
    preference = get_form_preference(str(current_user.id), _report_form_key(report_code), view_key)
    columns = preference.get("columns", [])
    if not isinstance(columns, list):
        return []
    visible: list[str] = []
    for column in columns:
        if not isinstance(column, dict):
            continue
        if not bool(column.get("visible", True)):
            continue
        field = str(column.get("field") or "").strip()
        if field:
            visible.append(field)
    return visible


def _preferred_group_by_from_view(report_code: str, view_key: str) -> str:
    if view_key == "default":
        return ""
    preference = get_form_preference(str(current_user.id), _report_form_key(report_code), view_key)
    filters = preference.get("filters", {})
    if not isinstance(filters, dict):
        return ""
    return str(filters.get("group_by") or "")


def _resolve_company(company_code: str) -> str:
    requested_company = company_code or "cacao"
    company_exists = database.session.execute(
        database.select(Entity.code).where(Entity.code == requested_company)
    ).scalar_one_or_none()
    if company_exists is not None:
        return requested_company
    default_company = database.session.execute(
        database.select(Entity.code).order_by(Entity.default.desc(), Entity.code.asc())
    ).scalar_one_or_none()
    return default_company or "cacao"


def _default_ledger_for_company(company_code: str) -> str | None:
    permisos = Permisos(modulo=obtener_id_modulo_por_nombre("accounting"), usuario=current_user.id)
    allowed_codes = permisos.obtener_libros_autorizados(company=company_code, return_codes=True)
    return (
        database.session.execute(
            database.select(Book.code)
            .where(Book.entity == company_code, Book.code.in_(allowed_codes))
            .order_by(Book.default.desc(), Book.is_primary.desc(), Book.code.asc())
        )
        .scalars()
        .first()
    )


def _default_period_for_company(company_code: str, target_date: date | None = None) -> str | None:
    effective_date = target_date or date.today()
    period_name = database.session.execute(
        database.select(AccountingPeriod.name)
        .where(
            AccountingPeriod.entity == company_code,
            AccountingPeriod.enabled.is_(True),
            AccountingPeriod.start <= effective_date,
            AccountingPeriod.end >= effective_date,
        )
        .order_by(AccountingPeriod.start.desc())
    ).scalar_one_or_none()
    if period_name:
        return period_name
    return database.session.execute(
        database.select(AccountingPeriod.name)
        .where(AccountingPeriod.entity == company_code, AccountingPeriod.enabled.is_(True))
        .order_by(AccountingPeriod.start.desc())
    ).scalar_one_or_none()


def _build_drill_down_url(
    values: dict[str, object], company: str, ledger: str | None, period: str | None = None
) -> str | None:
    account_code = values.get("account_code")
    if account_code in (None, "", _EMPTY_CELL_VALUE):
        return None
    query: dict[str, Any] = {
        "company": company,
        "account_code": str(account_code),
    }
    if ledger:
        query["ledger"] = ledger
    if period:
        query["accounting_period"] = period
    return url_for("reportes.account_movement", **query)


def _build_voucher_url(values: dict[str, object]) -> str | None:
    voucher_type = str(values.get("voucher_type") or "").lower()
    voucher_id = str(values.get("document_no") or values.get("voucher_id") or "").strip()
    if voucher_type == "journal_entry" and voucher_id:
        return url_for("contabilidad.ver_comprobante", identifier=voucher_id)
    return None


def _build_hierarchical_financial_rows(
    report_code: str, source_rows: list[dict[str, object]], company: str
) -> list[dict[str, object]]:
    if report_code not in {"trial-balance", "income-statement", "balance-sheet"}:
        return source_rows
    sections_order, section_nodes, section_non_account_rows = _collect_section_nodes(source_rows)
    _enrich_section_nodes(section_nodes, company)
    return _flatten_hierarchical_rows(sections_order, section_nodes, section_non_account_rows)


def _ensure_parent_nodes_exist(
    nodes: dict[str, dict[str, object]],
    account_code: str,
    row_section: object,
) -> None:
    """Garantiza que todos los nodos padre existan en la coleccion."""
    code_parts = account_code.split(".")
    for index in range(1, len(code_parts)):
        parent_code = ".".join(code_parts[:index])
        parent_node = nodes.setdefault(
            parent_code,
            {
                "section": row_section,
                "account_code": parent_code,
                "account_name": None,
            },
        )
        if row_section and not parent_node.get("section"):
            parent_node["section"] = row_section


def _collect_section_nodes(
    source_rows: list[dict[str, object]],
) -> tuple[list[str], dict[str, dict[str, dict[str, object]]], dict[str, list[dict[str, object]]]]:
    sections_order: list[str] = []
    section_nodes: dict[str, dict[str, dict[str, object]]] = {}
    section_non_account_rows: dict[str, list[dict[str, object]]] = {}
    for row in source_rows:
        section = str(row.get("section") or "__all__")
        if section not in sections_order:
            sections_order.append(section)
        account_code = str(row.get("account_code") or "").strip()
        if not account_code:
            section_non_account_rows.setdefault(section, []).append(dict(row))
            continue
        nodes = section_nodes.setdefault(section, {})
        existing_node = nodes.get(account_code, {})
        node = {**existing_node, **dict(row)}
        node["account_code"] = account_code
        nodes[account_code] = node
        _ensure_parent_nodes_exist(nodes, account_code, row.get("section"))
    return sections_order, section_nodes, section_non_account_rows


def _build_children_map(codes: set[str]) -> dict[str, list[str]]:
    """Construye mapa de relaciones padre-hijo para códigos de cuenta."""
    children_map: dict[str, list[str]] = {}
    for code in codes:
        parent = ".".join(code.split(".")[:-1])
        if parent:
            children_map.setdefault(parent, []).append(code)
    return children_map


def _compute_numeric_fields(nodes: dict[str, dict[str, object]]) -> set[str]:
    """Calcula los campos numéricos presentes en los nodos."""
    return {
        field
        for row in nodes.values()
        for field, value in row.items()
        if field in _MONEY_COLUMNS and isinstance(value, (int, float, Decimal, str))
    }


def _get_account_names(account_codes: list[str], company: str) -> dict[str, str]:
    """Obtiene nombres de cuentas desde la base de datos."""
    return {
        account.code: account.name
        for account in database.session.execute(
            database.select(Accounts).where(Accounts.entity == company, Accounts.code.in_(account_codes))
        ).scalars()
    }


def _enrich_node_metadata(
    node: dict[str, object],
    node_code: str,
    account_name: str | None,
    children_map: dict[str, list[str]],
) -> None:
    """Enriquece un nodo individual con metadatos calculados."""
    node["account_name"] = node.get("account_name") or account_name or node_code
    node["level"] = node_code.count(".") + 1
    node["is_group"] = bool(children_map.get(node_code))


def _enrich_section_nodes(section_nodes: dict[str, dict[str, dict[str, object]]], company: str) -> None:
    for section, nodes in section_nodes.items():
        if not nodes:
            continue
        account_codes = list(nodes)
        account_names = _get_account_names(account_codes, company)
        numeric_fields = _compute_numeric_fields(nodes)
        children_map = _build_children_map(set(nodes.keys()))
        _propagate_child_amounts_to_parents(nodes, numeric_fields)
        for node_code, node in nodes.items():
            _enrich_node_metadata(node, node_code, account_names.get(node_code), children_map)


def _propagate_child_amounts_to_parents(nodes: dict[str, dict[str, object]], numeric_fields: set[str]) -> None:
    for code in sorted(nodes.keys(), key=lambda value: value.count("."), reverse=True):
        parent = ".".join(code.split(".")[:-1])
        if not parent or parent not in nodes:
            continue
        for field in numeric_fields:
            parent_amount = _to_decimal_or_zero(nodes[parent].get(field))
            child_amount = _to_decimal_or_zero(nodes[code].get(field))
            nodes[parent][field] = parent_amount + child_amount


def _find_root_codes(nodes: dict[str, dict[str, object]]) -> list[str]:
    """Encuentra los códigos raíz que no tienen padre en el mismo conjunto."""
    return sorted(
        [code for code in nodes if ".".join(code.split(".")[:-1]) not in nodes],
        key=str,
    )


def _flatten_section(
    section: str,
    nodes: dict[str, dict[str, object]],
    non_account_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    """Flattenea una seccion individual."""
    result = list(non_account_rows)
    if not nodes:
        return result
    ordered_children_map = _build_children_map(set(nodes.keys()))
    root_codes = _find_root_codes(nodes)
    result.extend(_flatten_nodes_by_root(nodes, ordered_children_map, root_codes))
    return result


def _flatten_hierarchical_rows(
    sections_order: list[str],
    section_nodes: dict[str, dict[str, dict[str, object]]],
    section_non_account_rows: dict[str, list[dict[str, object]]],
) -> list[dict[str, object]]:
    flattened_rows: list[dict[str, object]] = []
    for section in sections_order:
        nodes = section_nodes.get(section, {})
        non_account = section_non_account_rows.get(section, [])
        flattened_rows.extend(_flatten_section(section, nodes, non_account))
    return flattened_rows


def _flatten_nodes_by_root(
    nodes: dict[str, dict[str, object]],
    ordered_children_map: dict[str, list[str]],
    root_codes: list[str],
) -> list[dict[str, object]]:
    result: list[dict[str, object]] = []

    def append_node(code: str) -> None:
        result.append(dict(nodes[code]))
        for child_code in sorted(ordered_children_map.get(code, []), key=str):
            append_node(child_code)

    for root_code in root_codes:
        append_node(root_code)
    return result


def _resolve_row_level(row: dict[str, object], account_code: str) -> int:
    level_value = row.get("level")
    if isinstance(level_value, int):
        return level_value
    try:
        if level_value is not None:
            return int(str(level_value))
    except (TypeError, ValueError):
        pass
    return account_code.count(".") + 1 if account_code else 0


def _date_arg(name: str) -> date | None:
    value = request.args.get(name)
    return date.fromisoformat(value) if value else None


def _int_arg(name: str, default: int) -> int:
    value = request.args.get(name, default=str(default))
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _bool_arg(name: str) -> bool:
    return request.args.get(name, "").lower() in {"1", "true", "yes", "on"}


def _financial_filters() -> FinancialReportFilters:
    company_code = _resolve_company(request.args.get("company", "cacao"))
    show_cancellations = _bool_arg("show_cancellations")
    ledger = request.args.get("ledger") or _default_ledger_for_company(company_code)
    accounting_period = request.args.get("accounting_period") or _default_period_for_company(company_code)
    return FinancialReportFilters(
        company=company_code,
        ledger=ledger,
        accounting_period=accounting_period,
        voucher_number=request.args.get("voucher_number") or None,
        account_code=request.args.get("account_code") or None,
        account_from=request.args.get("account_from") or None,
        account_to=request.args.get("account_to") or None,
        cost_center_code=request.args.get("cost_center_code") or None,
        unit_code=request.args.get("unit_code") or None,
        project_code=request.args.get("project_code") or None,
        party_type=request.args.get("party_type") or None,
        party_id=request.args.get("party_id") or None,
        voucher_type=request.args.get("voucher_type") or None,
        status=(request.args.get("status") or "submitted") if not show_cancellations else None,
        include_running_balance=_bool_arg("include_running_balance"),
        page=max(_int_arg("page", 1), 1),
        page_size=max(_int_arg("page_size", 100), 1),
        sort_by=request.args.get("sort_by", "posting_date"),
        sort_dir=request.args.get("sort_dir", "asc"),
        export_all=False,
    )


def _should_run_financial_report() -> bool:
    """Evita cargar datos al abrir la vista sin aplicar filtros explícitos."""
    if request.args.get("apply_filters") in {"1", "true", "yes", "on"}:
        return True
    return "export" in request.args


def _empty_financial_report() -> PaginatedReport:
    return PaginatedReport(rows=[], totals={}, columns=[], total_rows=0, page=1, page_size=100, ledger_currency=None)


def _report_to_matrix(report) -> tuple[list[str], list[list[object]]]:
    rows = getattr(report, "rows", [])
    columns = getattr(report, "columns", None) or (list(rows[0].values.keys()) if rows else [])
    data_rows = [[row.values.get(column) for column in columns] for row in rows]
    return columns, data_rows


def _write_operational_report_xlsx(report, report_code: str, title: str, filter_payload: dict[str, object]) -> bytes:
    columns, rows = _report_to_matrix(report)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report_code[:31]
    sheet.append([title])
    sheet.append([_("Fecha de generación"), date.today().isoformat()])
    sheet.append([_("Usuario"), getattr(current_user, "user", "")])
    sheet.append([])
    if columns:
        sheet.append([_column_label(column, report.ledger_currency) for column in columns])
        sheet.freeze_panes = "A5"
    for row in rows:
        sheet.append([_format_cell(column, row[index], report.ledger_currency) for index, column in enumerate(columns)])
    if report.totals:
        sheet.append([])
    for total_name, total_value in report.totals.items():
        sheet.append(
            [
                _("TOTAL"),
                _column_label(total_name, report.ledger_currency),
                _format_cell(total_name, total_value, report.ledger_currency),
            ]
        )
    for column_cells in sheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        max_length = max((len(value) for value in values), default=10)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 60)
    for column in range(1, sheet.max_column + 1):
        sheet.cell(row=4, column=column).alignment = Alignment(horizontal="center")

    filters_sheet = workbook.create_sheet(_("Filtros"))
    filters_sheet.append([_("Filtro"), _("Valor")])
    for key, value in filter_payload.items():
        if value in (None, "", False):
            continue
        filters_sheet.append([_(key.replace("_", " ").title()), str(value)])
    filters_sheet.freeze_panes = "A2"
    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def _export_operational_report(report, report_code: str, title: str, filter_payload: dict[str, object]):
    export_format = request.args.get("export")
    if export_format not in {"csv", "xlsx"}:
        return None

    columns, rows = _report_to_matrix(report)
    if export_format == "csv":
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(columns)
        writer.writerows(rows)
        return send_file(
            BytesIO(buffer.getvalue().encode("utf-8")),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"{report_code}.csv",
        )

    xlsx_content = _write_operational_report_xlsx(report, report_code, title, filter_payload)
    return send_file(
        BytesIO(xlsx_content),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{report_code}.xlsx",
    )


def _write_financial_report_xlsx(report, report_code: str, title: str, report_filters: FinancialReportFilters) -> bytes:
    columns, rows = _report_to_matrix(report)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report_code[:31]
    sheet.append([title])
    sheet.append([_("Fecha de generación"), date.today().isoformat()])
    sheet.append([_("Usuario"), getattr(current_user, "user", "")])
    sheet.append([])
    if columns:
        localized_headers = [_column_label(column, report.ledger_currency) for column in columns]
        sheet.append(localized_headers)
        sheet.freeze_panes = "A5"
    for row in rows:
        sheet.append([_format_cell(column, row[index], report.ledger_currency) for index, column in enumerate(columns)])
    if report.totals:
        sheet.append([])
    for total_name, total_value in report.totals.items():
        sheet.append(
            [
                _("TOTAL"),
                _column_label(total_name, report.ledger_currency),
                _format_cell(total_name, total_value, report.ledger_currency),
            ]
        )
    for column_cells in sheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        max_length = max((len(value) for value in values), default=10)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 60)
    for column in range(1, sheet.max_column + 1):
        sheet.cell(row=4, column=column).alignment = Alignment(horizontal="center")

    filters_sheet = workbook.create_sheet(_("Filtros"))
    filters_sheet.append([_("Filtro"), _("Valor")])
    for key in _FINANCIAL_FILTER_FIELDS:
        value = getattr(report_filters, key, None)
        if value in (None, "", False):
            continue
        filters_sheet.append([_(key.replace("_", " ").title()), str(value)])
    filters_sheet.freeze_panes = "A2"
    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def _export_financial_report(report, report_code: str, title: str, report_filters: FinancialReportFilters):
    export_format = request.args.get("export")
    if export_format not in {"csv", "xlsx"}:
        return None

    columns, rows = _report_to_matrix(report)
    if export_format == "csv":
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(columns)
        writer.writerows(rows)
        return send_file(
            BytesIO(buffer.getvalue().encode("utf-8")),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"{report_code}.csv",
        )

    xlsx_content = _write_financial_report_xlsx(report, report_code, title, report_filters)
    return send_file(
        BytesIO(xlsx_content),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{report_code}.xlsx",
    )


def _compute_display_columns(report, report_code: str, saved_view: str) -> list[str]:
    selected_columns = request.args.getlist("visible_columns")
    if not selected_columns:
        selected_columns = _preferred_columns_from_view(report_code, saved_view)
    columns = report.columns or []
    if selected_columns:
        columns = [column for column in columns if column in selected_columns]
    if report_code == "trial-balance":
        columns = [column for column in columns if column != "level"]
    display_columns = [
        column
        for column in columns
        if any((row.values.get(column) not in (None, "", "—") for row in report.rows)) or column in _ALWAYS_VISIBLE_COLUMNS
    ]
    return display_columns if display_columns else columns


def _compute_all_columns(report, report_code: str) -> list[str]:
    extra_columns = ["reference_type", "is_reversal", "reversal_of"]
    all_columns = list(dict.fromkeys([*(report.columns or []), *extra_columns]))
    if report_code == "trial-balance":
        all_columns = [column for column in all_columns if column != "level"]
    return all_columns


def _build_row_metadata(
    source_rows: list[dict[str, object]],
    report_filters: FinancialReportFilters,
) -> list[dict[str, object]]:
    child_counts: dict[str, int] = {}
    for row in source_rows:
        account_code = str(row.get("account_code") or "")
        if not account_code:
            continue
        parent_code = ".".join(account_code.split(".")[:-1])
        if parent_code:
            child_counts[parent_code] = child_counts.get(parent_code, 0) + 1
    row_metadata = []
    for row in source_rows:
        account_code = str(row.get("account_code") or "")
        parent_code = ".".join(account_code.split(".")[:-1]) if account_code else ""
        row_metadata.append(
            {
                "code": account_code,
                "parent": parent_code,
                "has_children": bool(child_counts.get(account_code)),
                "level": _resolve_row_level(row, account_code),
                "drilldown_url": _build_drill_down_url(
                    row, report_filters.company, report_filters.ledger, report_filters.accounting_period
                ),
                "voucher_url": _build_voucher_url(row),
                "is_group": bool(row.get("is_group")),
            }
        )
    return row_metadata


def _build_display_rows(
    source_rows: list[dict[str, object]],
    row_metadata: list[dict[str, object]],
    display_columns: list[str],
    ledger_currency: str | None,
) -> list[dict[str, object]]:
    display_rows: list[dict[str, object]] = []
    for index, row in enumerate(source_rows):
        formatted_row: dict[str, object] = {
            column: _format_cell(column, row.get(column), ledger_currency) for column in display_columns
        }
        formatted_row["__meta"] = row_metadata[index]
        display_rows.append(formatted_row)
    return display_rows


def _apply_grouping(
    display_rows: list[dict[str, object]],
    source_rows: list[dict[str, object]],
    report,
    report_code: str,
    saved_view: str,
) -> list[dict[str, object]]:
    group_by = request.args.get("group_by") or _preferred_group_by_from_view(report_code, saved_view)
    if not (report_code == "account-movement" and group_by and group_by in (report.columns or [])):
        return display_rows
    return _group_rows_by_field(display_rows, source_rows, group_by, report.ledger_currency)


def _group_rows_by_field(
    display_rows: list[dict[str, object]],
    source_rows: list[dict[str, object]],
    group_by: str,
    ledger_currency: str | None,
) -> list[dict[str, object]]:
    grouped_rows: list[dict[str, object]] = []
    current_group: str | None = None
    group_debit = Decimal("0")
    group_credit = Decimal("0")
    for index, row in enumerate(display_rows):
        raw_group_value = source_rows[index].get(group_by)
        group_value = _format_cell(group_by, raw_group_value, ledger_currency)
        if group_value != current_group:
            if current_group is not None:
                grouped_rows.append(_build_group_subtotal_row(group_debit, group_credit, ledger_currency))
            group_title = _(group_by.replace("_", " ").title())
            grouped_rows.append({"__row_type": "group", "__group_title": f"{group_title}: {group_value}"})
            current_group = group_value
            group_debit = Decimal("0")
            group_credit = Decimal("0")
        group_debit = _add_to_decimal_or_zero(group_debit, row.get("debit"))
        group_credit = _add_to_decimal_or_zero(group_credit, row.get("credit"))
        grouped_rows.append(row)
    if current_group is not None:
        grouped_rows.append(_build_group_subtotal_row(group_debit, group_credit, ledger_currency))
    return grouped_rows


def _build_group_subtotal_row(debit: Decimal, credit: Decimal, ledger_currency: str | None) -> dict[str, object]:
    return {
        "__row_type": "group_subtotal",
        "__group_title": _("Subtotal"),
        "debit": _format_cell("debit", debit, ledger_currency),
        "credit": _format_cell("credit", credit, ledger_currency),
    }


def _add_to_decimal_or_zero(current: Decimal, value: object) -> Decimal:
    try:
        return current + Decimal(str(value or "0").replace(",", "").replace("(", "-").replace(")", ""))
    except DecimalException:
        return current


def _render_financial_report(
    report_code: str,
    report_title: str,
    report,
    report_filters: FinancialReportFilters,
    saved_view: str,
    saved_views: list[str],
):
    export_response = _export_financial_report(report, report_code, report_title, report_filters)
    if export_response is not None:
        return export_response
    display_columns = _compute_display_columns(report, report_code, saved_view)
    all_columns = _compute_all_columns(report, report_code)
    allow_column_selection = report_code in {"account-movement", "account-summary"}
    group_by = request.args.get("group_by") or _preferred_group_by_from_view(report_code, saved_view)
    display_headers = {column: _column_label(column, report.ledger_currency) for column in display_columns}
    all_column_headers = {column: _column_label(column, report.ledger_currency) for column in all_columns}
    source_rows = [dict(row.values) for row in report.rows]
    source_rows = _build_hierarchical_financial_rows(report_code, source_rows, report_filters.company)
    row_metadata = _build_row_metadata(source_rows, report_filters)
    display_rows = _build_display_rows(source_rows, row_metadata, display_columns, report.ledger_currency)
    grouped_rows = _apply_grouping(display_rows, source_rows, report, report_code, saved_view)
    display_totals = {key: _format_cell(key, value, report.ledger_currency) for key, value in report.totals.items()}
    return render_template(
        "reportes/financial_report.html",
        titulo=f"{report_title} - {APPNAME}",
        report_code=report_code,
        report_title=report_title,
        rows=report.rows,
        columns=display_columns,
        display_headers=display_headers,
        all_column_headers=all_column_headers,
        display_rows=grouped_rows,
        totals=display_totals,
        total_rows=report.total_rows,
        page=report.page,
        page_size=report.page_size,
        ledger_currency=report.ledger_currency,
        context_summary=_build_context_summary(report, report_filters),
        report_filters=report_filters,
        right_align_columns=_RIGHT_ALIGN_COLUMNS,
        is_balanced=_is_report_balanced(report.totals),
        saved_view=saved_view,
        saved_views=saved_views,
        selected_columns=display_columns,
        all_columns=all_columns,
        allow_column_selection=allow_column_selection,
        group_by=group_by,
    )


def _render_operational_framework(
    report_code: str,
    report_title: str,
    report,
    *,
    module_home_endpoint: str,
    module_home_label: str,
    filter_mode: str,
    filter_state: dict[str, object],
    context_summary: dict[str, str],
):
    export_response = _export_operational_report(report, report_code, report_title, filter_state)
    if export_response is not None:
        return export_response
    rows = getattr(report, "rows", [])
    totals_raw = getattr(report, "totals", {})
    ledger_currency = getattr(report, "ledger_currency", None)
    columns = getattr(report, "columns", None) or (list(rows[0].values.keys()) if rows else [])
    display_headers = {column: _column_label(column, ledger_currency) for column in columns}
    display_rows = [
        {column: _format_cell(column, row.values.get(column), ledger_currency) for column in columns} for row in rows
    ]
    totals = {key: _format_cell(key, value, ledger_currency) for key, value in totals_raw.items()}
    return render_template(
        "reportes/operational_report.html",
        titulo=f"{report_title} - {APPNAME}",
        report_title=report_title,
        report_code=report_code,
        columns=columns,
        display_headers=display_headers,
        display_rows=display_rows,
        totals=totals,
        filter_mode=filter_mode,
        filter_state=filter_state,
        context_summary=context_summary,
        right_align_columns=_RIGHT_ALIGN_COLUMNS,
        module_home_url=url_for(module_home_endpoint),
        module_home_label=module_home_label,
    )


@reportes.route("/reports/account-summary")
@login_required
@modulo_activo("accounting")
@verifica_acceso("accounting")
def account_summary():
    """Resumen de movimientos por cuenta (Sábana analítica)."""
    filters, selected_view, saved_views = _resolve_view_context("account-summary", _financial_filters())
    report = get_account_summary_report(filters) if _should_run_financial_report() else _empty_financial_report()
    return _render_financial_report(
        "account-summary",
        _("Resumen de Movimiento por Cuenta"),
        report,
        filters,
        selected_view,
        saved_views,
    )


@reportes.route("/reports/account-movement")
@login_required
@modulo_activo("accounting")
@verifica_acceso("accounting")
def account_movement():
    """Report account movement detail."""
    filters, selected_view, saved_views = _resolve_view_context("account-movement", _financial_filters())
    report = get_account_movement_detail(filters) if _should_run_financial_report() else _empty_financial_report()
    if request.args.get("export") in {"csv", "xlsx"}:
        export_report = get_account_movement_detail(cast(FinancialReportFilters, replace(filters, export_all=True, page=1)))
        return _render_financial_report(
            "account-movement",
            _("Detalle de Movimiento Contable"),
            export_report,
            filters,
            selected_view,
            saved_views,
        )
    return _render_financial_report(
        "account-movement",
        _("Detalle de Movimiento Contable"),
        report,
        filters,
        selected_view,
        saved_views,
    )


@reportes.route("/reports/trial-balance")
@login_required
@modulo_activo("accounting")
@verifica_acceso("accounting")
def trial_balance():
    """Report trial balance."""
    filters, selected_view, saved_views = _resolve_view_context("trial-balance", _financial_filters())
    report = get_trial_balance_report(filters) if _should_run_financial_report() else _empty_financial_report()
    return _render_financial_report(
        "trial-balance",
        _("Balanza de Comprobación"),
        report,
        filters,
        selected_view,
        saved_views,
    )


@reportes.route("/reports/income-statement")
@login_required
@modulo_activo("accounting")
@verifica_acceso("accounting")
def income_statement():
    """Report income statement."""
    filters, selected_view, saved_views = _resolve_view_context("income-statement", _financial_filters())
    report = get_income_statement_report(filters) if _should_run_financial_report() else _empty_financial_report()
    return _render_financial_report(
        "income-statement",
        _("Estado de Resultado"),
        report,
        filters,
        selected_view,
        saved_views,
    )


@reportes.route("/reports/balance-sheet")
@login_required
@modulo_activo("accounting")
@verifica_acceso("accounting")
def balance_sheet():
    """Report balance sheet."""
    filters, selected_view, saved_views = _resolve_view_context("balance-sheet", _financial_filters())
    report = get_balance_sheet_report(filters) if _should_run_financial_report() else _empty_financial_report()
    return _render_financial_report(
        "balance-sheet",
        _("Balance General"),
        report,
        filters,
        selected_view,
        saved_views,
    )


@reportes.route("/reports/subledger")
@login_required
@modulo_activo("accounting")
def subledger():
    """Report AR/AP subledger by document."""
    company = request.args.get("company", "cacao")
    party_type = request.args.get("party_type", "customer")
    report = get_ar_ap_subledger(
        SubledgerFilters(
            company=company,
            party_type=party_type,
            party_id=request.args.get("party_id") or None,
            as_of_date=_date_arg("as_of_date"),
        )
    )
    return render_template(
        REPORT_TABLE_HTML,
        titulo="Subledger AR/AP - " + APPNAME,
        report_title="Subledger AR/AP",
        rows=report.rows,
        totals=report.totals,
    )


@reportes.route("/reports/aging")
@login_required
@modulo_activo("accounting")
def aging():
    """Report AR/AP aging."""
    company = request.args.get("company", "cacao")
    party_type = request.args.get("party_type", "customer")
    report = get_aging_report(
        AgingFilters(
            company=company,
            party_type=party_type,
            party_id=request.args.get("party_id") or None,
            as_of_date=_date_arg("as_of_date") or date.today(),
        )
    )
    return render_template(
        REPORT_TABLE_HTML,
        titulo="Aging AR/AP - " + APPNAME,
        report_title="Aging AR/AP",
        rows=report.rows,
        totals=report.totals,
    )


@reportes.route("/reports/kardex")
@login_required
@modulo_activo("inventory")
def kardex():
    """Report inventory kardex."""
    company = _resolve_company(request.args.get("company", "cacao"))
    filters = KardexFilters(
        company=company,
        item_code=request.args.get("item_code") or None,
        warehouse=request.args.get("warehouse") or None,
        date_from=_date_arg("date_from"),
        date_to=_date_arg("date_to"),
    )
    report = get_kardex(filters)
    return _render_operational_framework(
        "kardex",
        _("Kardex"),
        report,
        module_home_endpoint="inventario.inventario_",
        module_home_label=_("Inventario"),
        filter_mode="kardex",
        filter_state={
            "company": company,
            "item_code": filters.item_code or "",
            "warehouse": filters.warehouse or "",
            "date_from": filters.date_from.isoformat() if filters.date_from else "",
            "date_to": filters.date_to.isoformat() if filters.date_to else "",
        },
        context_summary=_operational_context_summary(
            report,
            company=company,
            date_from=filters.date_from.isoformat() if filters.date_from else None,
            date_to=filters.date_to.isoformat() if filters.date_to else None,
        ),
    )


@reportes.route("/reports/inventory-existence")
@login_required
@modulo_activo("inventory")
def inventory_existence():
    """Genera reporte de existencias a una fecha clave."""
    company = _resolve_company(request.args.get("company", "cacao"))
    as_of_date = _date_arg("as_of_date")
    filters = KardexFilters(
        company=company,
        item_code=request.args.get("item_code") or None,
        warehouse=request.args.get("warehouse") or None,
        date_to=as_of_date,
    )
    report = get_inventory_existence(filters)
    return _render_operational_framework(
        "inventory-existence",
        _("Existencia de Inventario"),
        report,
        module_home_endpoint="inventario.inventario_",
        module_home_label=_("Inventario"),
        filter_mode="inventory_existence",
        filter_state={
            "company": company,
            "item_code": filters.item_code or "",
            "warehouse": filters.warehouse or "",
            "as_of_date": as_of_date.isoformat() if as_of_date else "",
        },
        context_summary=_operational_context_summary(
            report,
            company=company,
            as_of_date=as_of_date.isoformat() if as_of_date else None,
        ),
    )


@reportes.route("/reports/reconciliations")
@login_required
@modulo_activo("accounting")
def reconciliations():
    """Report reconciliations."""
    report = get_reconciliation_report(
        company=request.args.get("company", "cacao"),
        as_of_date=_date_arg("as_of_date"),
    )
    return render_template(
        REPORT_TABLE_HTML,
        titulo="Reconciliaciones - " + APPNAME,
        report_title="Reconciliaciones",
        rows=report.rows,
        totals=report.totals,
    )


def _operational_filters() -> OperationalReportFilters:
    return OperationalReportFilters(
        company=request.args.get("company", "cacao"),
        date_from=_date_arg("date_from"),
        date_to=_date_arg("date_to"),
        party_id=request.args.get("party_id") or None,
        item_code=request.args.get("item_code") or None,
        warehouse=request.args.get("warehouse") or None,
    )


def _render_operational_report(report_name: str, report):
    return render_template(
        REPORT_TABLE_HTML,
        titulo=report_name + " - " + APPNAME,
        report_title=report_name,
        rows=report.rows,
        totals=report.totals,
    )


def _operational_context_summary(report, **values: object) -> dict[str, str]:
    total_rows = getattr(report, "total_rows", len(getattr(report, "rows", [])))
    summary = {"company": str(values.get("company") or "—"), "records": str(total_rows)}
    for key, value in values.items():
        if key == "company":
            continue
        summary[key] = "—" if value in (None, "") else str(value)
    return summary


@reportes.route("/reports/bank-movement")
@login_required
@modulo_activo(("cash", "banking"))
def bank_movement():
    """Genera reporte de detalle de movimiento bancario."""
    company = _resolve_company(request.args.get("company", "cacao"))
    filters = BankingFilters(
        company=company,
        bank_account_id=request.args.get("bank_account_id") or None,
        date_from=_date_arg("date_from"),
        date_to=_date_arg("date_to"),
    )
    report = get_bank_movement_detail(filters)
    return _render_operational_framework(
        "bank-movement",
        _("Detalle de Movimiento Bancario"),
        report,
        module_home_endpoint="bancos.bancos_",
        module_home_label=_("Bancos"),
        filter_mode="bank_movement",
        filter_state={
            "company": company,
            "bank_account_id": filters.bank_account_id or "",
            "date_from": filters.date_from.isoformat() if filters.date_from else "",
            "date_to": filters.date_to.isoformat() if filters.date_to else "",
        },
        context_summary=_operational_context_summary(
            report,
            company=company,
            date_from=filters.date_from.isoformat() if filters.date_from else None,
            date_to=filters.date_to.isoformat() if filters.date_to else None,
        ),
    )


@reportes.route("/reports/bank-balance-summary")
@login_required
@modulo_activo(("cash", "banking"))
def bank_balance_summary():
    """Genera reporte de resumen de saldos bancarios."""
    company = _resolve_company(request.args.get("company", "cacao"))
    filters = BankingFilters(
        company=company,
        bank_account_id=request.args.get("bank_account_id") or None,
        as_of_date=_date_arg("as_of_date"),
    )
    report = get_bank_balance_summary(filters)
    return _render_operational_framework(
        "bank-balance-summary",
        _("Resumen de Saldos Bancarios"),
        report,
        module_home_endpoint="bancos.bancos_",
        module_home_label=_("Bancos"),
        filter_mode="bank_balance_summary",
        filter_state={
            "company": company,
            "bank_account_id": filters.bank_account_id or "",
            "as_of_date": filters.as_of_date.isoformat() if filters.as_of_date else "",
        },
        context_summary=_operational_context_summary(
            report,
            company=company,
            as_of_date=filters.as_of_date.isoformat() if filters.as_of_date else None,
        ),
    )


@reportes.route("/reports/accounts-payable")
@login_required
@modulo_activo("purchases")
def accounts_payable():
    """Genera reporte de cuentas por pagar por proveedor a fecha clave."""
    company = _resolve_company(request.args.get("company", "cacao"))
    as_of_date = _date_arg("as_of_date")
    party_id = request.args.get("party_id") or None
    report = get_ar_ap_subledger(
        SubledgerFilters(company=company, party_type="supplier", party_id=party_id, as_of_date=as_of_date)
    )
    return _render_operational_framework(
        "accounts-payable",
        _("Cuentas por Pagar"),
        report,
        module_home_endpoint="compras.compras_",
        module_home_label=_("Compras"),
        filter_mode="accounts_payable",
        filter_state={
            "company": company,
            "party_id": party_id or "",
            "as_of_date": as_of_date.isoformat() if as_of_date else "",
        },
        context_summary=_operational_context_summary(
            report,
            company=company,
            party_type=_("Proveedor"),
            as_of_date=as_of_date.isoformat() if as_of_date else None,
        ),
    )


@reportes.route("/reports/ap-aging")
@login_required
@modulo_activo("purchases")
def ap_aging():
    """Genera aging de cuentas por pagar."""
    company = _resolve_company(request.args.get("company", "cacao"))
    as_of_date = _date_arg("as_of_date") or date.today()
    party_id = request.args.get("party_id") or None
    report = get_aging_report(AgingFilters(company=company, party_type="supplier", party_id=party_id, as_of_date=as_of_date))
    return _render_operational_framework(
        "ap-aging",
        _("Aging de Cuentas por Pagar"),
        report,
        module_home_endpoint="compras.compras_",
        module_home_label=_("Compras"),
        filter_mode="ap_aging",
        filter_state={"company": company, "party_id": party_id or "", "as_of_date": as_of_date.isoformat()},
        context_summary=_operational_context_summary(
            report, company=company, party_type=_("Proveedor"), as_of_date=as_of_date.isoformat()
        ),
    )


@reportes.route("/reports/accounts-receivable")
@login_required
@modulo_activo("sales")
def accounts_receivable():
    """Genera reporte de cuentas por cobrar por cliente a fecha clave."""
    company = _resolve_company(request.args.get("company", "cacao"))
    as_of_date = _date_arg("as_of_date")
    party_id = request.args.get("party_id") or None
    report = get_ar_ap_subledger(
        SubledgerFilters(company=company, party_type="customer", party_id=party_id, as_of_date=as_of_date)
    )
    return _render_operational_framework(
        "accounts-receivable",
        _("Cuentas por Cobrar"),
        report,
        module_home_endpoint="ventas.ventas_",
        module_home_label=_("Ventas"),
        filter_mode="accounts_receivable",
        filter_state={
            "company": company,
            "party_id": party_id or "",
            "as_of_date": as_of_date.isoformat() if as_of_date else "",
        },
        context_summary=_operational_context_summary(
            report,
            company=company,
            party_type=_("Cliente"),
            as_of_date=as_of_date.isoformat() if as_of_date else None,
        ),
    )


@reportes.route("/reports/ar-aging")
@login_required
@modulo_activo("sales")
def ar_aging():
    """Genera aging de cuentas por cobrar."""
    company = _resolve_company(request.args.get("company", "cacao"))
    as_of_date = _date_arg("as_of_date") or date.today()
    party_id = request.args.get("party_id") or None
    report = get_aging_report(AgingFilters(company=company, party_type="customer", party_id=party_id, as_of_date=as_of_date))
    return _render_operational_framework(
        "ar-aging",
        _("Aging de Cuentas por Cobrar"),
        report,
        module_home_endpoint="ventas.ventas_",
        module_home_label=_("Ventas"),
        filter_mode="ar_aging",
        filter_state={"company": company, "party_id": party_id or "", "as_of_date": as_of_date.isoformat()},
        context_summary=_operational_context_summary(
            report, company=company, party_type=_("Cliente"), as_of_date=as_of_date.isoformat()
        ),
    )


@reportes.route("/reports/purchases-by-supplier")
@login_required
@modulo_activo("purchases")
def purchases_by_supplier():
    """Genera reporte de compras agrupadas por proveedor."""
    return _render_operational_report("Compras por Proveedor", get_purchases_by_supplier(_operational_filters()))


@reportes.route("/reports/purchases-by-item")
@login_required
@modulo_activo("purchases")
def purchases_by_item():
    """Genera reporte de compras agrupadas por articulo."""
    return _render_operational_report("Compras por Item", get_purchases_by_item(_operational_filters()))


@reportes.route("/reports/sales-by-customer")
@login_required
@modulo_activo("sales")
def sales_by_customer():
    """Genera reporte de ventas agrupadas por cliente."""
    return _render_operational_report("Ventas por Cliente", get_sales_by_customer(_operational_filters()))


@reportes.route("/reports/sales-by-item")
@login_required
@modulo_activo("sales")
def sales_by_item():
    """Genera reporte de ventas agrupadas por articulo."""
    return _render_operational_report("Ventas por Item", get_sales_by_item(_operational_filters()))


@reportes.route("/reports/gross-margin")
@login_required
@modulo_activo("sales")
def gross_margin():
    """Genera reporte de margen bruto por ventas."""
    return _render_operational_report("Margen Bruto", get_gross_margin(_operational_filters()))


@reportes.route("/reports/stock-balance")
@login_required
@modulo_activo("inventory")
def stock_balance():
    """Genera reporte de balance de stock por articulo y bodega."""
    return _render_operational_report("Stock Balance", get_stock_balance(_operational_filters()))


@reportes.route("/reports/inventory-valuation")
@login_required
@modulo_activo("inventory")
def inventory_valuation():
    """Genera reporte de valoracion del inventario."""
    return _render_operational_report("Valoracion de Inventario", get_inventory_valuation(_operational_filters()))


@reportes.route("/reports/batches")
@login_required
@modulo_activo("inventory")
def batches():
    """Genera reporte de lotes de inventario."""
    return _render_operational_report("Lotes", get_batch_report(_operational_filters()))


@reportes.route("/reports/serials")
@login_required
@modulo_activo("inventory")
def serials():
    """Genera reporte de numeros de serie de inventario."""
    return _render_operational_report("Seriales", get_serial_report(_operational_filters()))
