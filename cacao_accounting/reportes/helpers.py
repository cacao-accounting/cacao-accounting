# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2025 - 2026 William José Moreno Reyes

"""Reportes operativos de subledgers, aging, Kardex y reconciliaciones."""

from __future__ import annotations

import csv
from datetime import date
from decimal import Decimal
from decimal import DecimalException
from io import BytesIO, StringIO
from typing import Any

from flask import Blueprint, render_template, request, send_file, url_for
from flask_login import current_user
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from werkzeug.routing import BuildError

from cacao_accounting.auth.permisos import Permisos
from cacao_accounting.database import Accounts, AccountingPeriod, Book, Entity, database
from cacao_accounting.database.helpers import obtener_id_modulo_por_nombre
from cacao_accounting.reportes.services import (
    FinancialReportFilters,
    OperationalReportFilters,
    PaginatedReport,
)
from cacao_accounting.version import APPNAME

try:  # pragma: no cover - fallback defensivo para contextos sin Flask-Babel inicializado.
    from flask_babel import gettext as _babel_gettext
except ImportError:  # pragma: no cover

    def _(value: str) -> str:
        return value

else:

    def _(value: str) -> str:
        try:
            return _babel_gettext(value)
        except (KeyError, RuntimeError):
            return value


reportes = Blueprint("reportes", __name__, template_folder="templates")

REPORT_TABLE_HTML = "reportes/report_table.html"

_COLUMN_LABELS = {
    "posting_date": "Posting Date",
    "accounting_period": "Period",
    "document_no": "Voucher",
    "voucher_type": "Type",
    "account_code": "Account",
    "account_name": "Account Name",
    "account_type": "Account Type",
    "classification": "Section",
    "debit": "Debit",
    "credit": "Credit",
    "running_balance": "Final Balance",
    "currency": "Currency",
    "ledger": "Ledger",
    "company": "Company",
    "opening_balance": "Opening Balance",
    "ending_balance": "Final Balance",
    "cost_center": "Cost Center",
    "unit": "Unit",
    "project": "Project",
    "party_type": "Party Type",
    "party_id": "Party",
    "created_by": "User",
    "created": "Creation Date",
    "created_at": "Creation Date",
    "movement_count": "Movements",
    "first_movement": "First Movement",
    "last_movement": "Last Movement",
    "line_comment": "Reference",
    "reference_type": "Reference Type",
    "is_reversal": "Is Reversal",
    "reversal_of": "Reversal Of",
    "status": "Status",
    "voucher_status": "Status",
    "section": "Section",
    "amount": "Amount",
    "bank_account": "Bank Account",
    "party_name": "Party",
    "certificate_no": "Certificate",
    "payment_no": "Payment",
    "supplier_name": "Supplier",
    "supplier_tax_id": "Supplier Tax ID",
    "concept": "Withholding Concept",
    "base_amount": "Tax Base",
    "rate": "Rate",
    "withheld_amount": "Withheld Amount",
    "payment_type": "Payment Type",
    "incoming_amount": "Incoming Amount",
    "outgoing_amount": "Outgoing Amount",
    "receipts_amount": "Receipts",
    "payments_amount": "Payments",
    "account_no": "Account Number",
    "item_name": "Item Name",
    "balance_qty": "Balance Qty",
    "incoming_qty": "Incoming Qty",
    "outgoing_qty": "Outgoing Qty",
    "value_change": "Value Change",
    "original_amount": "Original Amount",
    "paid_amount": "Paid Amount",
    "outstanding_amount": "Outstanding Amount",
    "days": "Days",
    "bucket": "Bucket",
    "remarks": "Remarks",
}
_MONEY_COLUMNS = {
    "debit",
    "credit",
    "difference",
    "opening_balance",
    "ending_balance",
    "running_balance",
    "amount",
    "assets",
    "liabilities",
    "equity",
    "period_profit",
    "income",
    "cost",
    "expense",
    "gross_profit",
    "net_profit",
    "unclassified_amount",
    "incoming_amount",
    "outgoing_amount",
    "receipts_amount",
    "payments_amount",
    "original_amount",
    "base_amount",
    "withheld_amount",
    "paid_amount",
    "outstanding_amount",
    "value_change",
    "stock_value",
    "remaining_stock_value",
}
_RIGHT_ALIGN_COLUMNS = _MONEY_COLUMNS | {"level", "incoming_qty", "outgoing_qty", "balance_qty", "actual_qty", "days"}
_ALWAYS_VISIBLE_COLUMNS = {
    "debit",
    "credit",
    "difference",
    "account_code",
    "account_name",
    "section",
    "amount",
    "opening_balance",
    "ending_balance",
}
_EMPTY_CELL_VALUE = "—"
_FINANCIAL_FILTER_FIELDS = (
    "company",
    "ledger",
    "accounting_period",
    "voucher_number",
    "account_code",
    "account_from",
    "account_to",
    "cost_center_code",
    "unit_code",
    "project_code",
    "party_type",
    "party_id",
    "voucher_type",
    "status",
    "show_cancellations",
    "include_running_balance",
    "page_size",
    "sort_by",
    "sort_dir",
    "group_by",
)


def _to_decimal_or_zero(value: object) -> Decimal:
    try:
        return value if isinstance(value, Decimal) else Decimal(str(value))
    except DecimalException:
        return Decimal("0")


def _format_number(value: object) -> str:
    try:
        amount = value if isinstance(value, Decimal) else Decimal(str(value))
    except DecimalException:
        return _EMPTY_CELL_VALUE
    formatted = f"{abs(amount):,.2f}"
    return f"({formatted})" if amount < 0 else formatted


def _column_label(column: str, ledger_currency: str | None) -> str:
    label = _(_COLUMN_LABELS.get(column, column.replace("_", " ").title()))
    if column in _MONEY_COLUMNS and ledger_currency:
        return f"{label} ({ledger_currency})"
    return label


def _format_cell(column: str, value: object, ledger_currency: str | None) -> str:
    if value is None or value == "":
        return _EMPTY_CELL_VALUE
    if isinstance(value, dict):
        return " / ".join(f"{curr} {_format_number(val)}" for curr, val in value.items())
    if column in _MONEY_COLUMNS:
        return _format_number(value)
    if column == "posting_date" and isinstance(value, date):
        return value.isoformat()
    if column == "voucher_status":
        status_value = str(value).lower()
        if status_value == "cancelled":
            return _("Cancelado")
        if status_value == "reversal":
            return _("Reversión")
        return _("Contabilizado")
    if column == "section":
        section_labels = {
            "assets": _("ACTIVOS"),
            "liabilities": _("PASIVOS"),
            "equity": _("PATRIMONIO"),
            "income": _("INGRESOS"),
            "cost": _("COSTOS"),
            "expense": _("GASTOS"),
            "gross_profit": _("UTILIDAD BRUTA"),
            "net_profit": _("UTILIDAD NETA"),
            "unclassified": _("SIN CLASIFICACIÓN"),
            "operating_adjustments": _("AJUSTES AL RESULTADO"),
            "total_operating": _("TOTAL ACTIVIDADES DE OPERACIÓN"),
            "investing": _("ACTIVIDADES DE INVERSIÓN"),
            "total_investing": _("TOTAL ACTIVIDADES DE INVERSIÓN"),
            "financing": _("ACTIVIDADES DE FINANCIAMIENTO"),
            "total_financing": _("TOTAL ACTIVIDADES DE FINANCIAMIENTO"),
            "net_change_cash": _("VARIACIÓN NETA DE EFECTIVO"),
            "cash_opening": _("EFECTIVO AL INICIO DEL PERÍODO"),
            "cash_closing": _("EFECTIVO AL FINAL DEL PERÍODO"),
        }
        return section_labels.get(str(value), str(value))
    return str(value)


def _build_context_summary(report, report_filters: FinancialReportFilters) -> dict[str, str]:
    ledger_label = report_filters.ledger or "—"
    if report_filters.ledger and report.ledger_currency:
        ledger_label = f"{report_filters.ledger} ({report.ledger_currency})"
    if report_filters.status == "cancelled":
        status_label = _("Cancelado")
    elif report_filters.include_cancellations:
        status_label = _("Todos")
    else:
        status_label = _("Contabilizado")
    period_label = _period_display_label(report_filters)
    return {
        "company": report_filters.company,
        "ledger": ledger_label,
        "period": period_label,
        "status": status_label,
        "records": str(report.total_rows),
    }


def _period_display_label(report_filters: FinancialReportFilters) -> str:
    """Etiqueta de presentación del período o rango seleccionado."""
    if getattr(report_filters, "period_from", None) or getattr(report_filters, "period_to", None):
        from cacao_accounting.reportes.periods import resolve_period_range

        period_range = resolve_period_range(
            report_filters.company, getattr(report_filters, "period_from", None), getattr(report_filters, "period_to", None)
        )
        if period_range is not None:
            return period_range.label
    return report_filters.accounting_period or "—"


def _is_report_balanced(raw_totals: dict[str, Decimal]) -> bool:
    difference = raw_totals.get("difference")
    if difference is None:
        return False
    try:
        return Decimal(str(difference)) == Decimal("0")
    except DecimalException:
        return False


def _unclassified_accounts_count(raw_totals: dict[str, object]) -> int:
    """Devuelve cuántas cuentas sin clasificación advierte el reporte."""
    value = raw_totals.get("unclassified_accounts", 0)
    try:
        return max(int(value), 0)  # type: ignore[call-overload]
    except (TypeError, ValueError):
        return 0


def _build_unclassified_warning(count: int) -> str | None:
    """Construye el mensaje de advertencia por cuentas excluidas del reporte."""
    if count <= 0:
        return None
    message = _(
        "%(count)s cuenta(s) con clasificación vacía o desconocida fueron excluidas de este reporte."
        " Revise la sección SIN CLASIFICACIÓN y corrija el catálogo contable."
    )
    try:
        return message % {"count": count}
    except (KeyError, TypeError, ValueError):  # pragma: no cover - mensaje sin marcadores
        return message


def _report_form_key(report_code: str) -> str:
    return f"reports.financial.{report_code}"


def _load_report_view_options(report_code: str) -> list[str]:
    return ["default"]


def _extract_filter_payload() -> dict[str, str]:
    payload: dict[str, str] = {}
    for key in _FINANCIAL_FILTER_FIELDS:
        value = request.args.get(key)
        if value:
            payload[key] = value
    visible_columns = request.args.getlist("visible_columns")
    if visible_columns:
        payload["visible_columns"] = ",".join(visible_columns)
    return payload


def _restore_filters_from_view(filters: FinancialReportFilters, report_code: str, view_key: str) -> FinancialReportFilters:
    return filters


def _safe_page_size(raw: object, default: int) -> int:
    try:
        if raw is None or isinstance(raw, str):
            value = default
        else:
            value = int(raw)  # type: ignore[call-overload]
        return max(value, 1)
    except (TypeError, ValueError):
        return max(default, 1)


def _str_or_none(value: object) -> str | None:
    return str(value or "") or None


def _handle_saved_view_action(report_code: str, filters: FinancialReportFilters) -> tuple[FinancialReportFilters, str]:
    view_key = (request.args.get("saved_view") or "default").strip() or "default"
    return filters, view_key


def _resolve_view_context(report_code: str, filters: FinancialReportFilters) -> tuple[FinancialReportFilters, str, list[str]]:
    resolved_filters, selected_view = _handle_saved_view_action(report_code, filters)
    return resolved_filters, selected_view, _load_report_view_options(report_code)


def _preferred_columns_from_view(report_code: str, view_key: str) -> list[str]:
    return []


def _preferred_group_by_from_view(report_code: str, view_key: str) -> str:
    return ""


def _resolve_company(company_code: str) -> str:
    requested_company = company_code or "cacao"
    company_exists = database.session.execute(
        database.select(Entity.code).where(Entity.code == requested_company)
    ).scalar_one_or_none()
    if company_exists is not None:
        return requested_company
    default_company = database.session.execute(
        database.select(Entity.code).order_by(Entity.default.desc(), Entity.code.asc())
    ).scalar_one_or_none()
    return default_company or "cacao"


def _default_ledger_for_company(company_code: str) -> str | None:
    permisos = Permisos(modulo=obtener_id_modulo_por_nombre("accounting"), usuario=current_user.id)
    if not permisos.tiene_acceso_compania(company_code):
        return None
    return (
        database.session.execute(
            database.select(Book.code)
            .where(Book.entity == company_code)
            .order_by(Book.default.desc(), Book.is_primary.desc(), Book.code.asc())
        )
        .scalars()
        .first()
    )


def _default_period_for_company(company_code: str, target_date: date | None = None) -> str | None:
    effective_date = target_date or date.today()
    period_name = (
        database.session.execute(
            database.select(AccountingPeriod.name)
            .where(
                AccountingPeriod.entity == company_code,
                AccountingPeriod.enabled.is_(True),
                AccountingPeriod.start <= effective_date,
                AccountingPeriod.end >= effective_date,
            )
            .order_by(AccountingPeriod.start.desc())
        )
        .scalars()
        .first()
    )
    if period_name:
        return period_name
    return (
        database.session.execute(
            database.select(AccountingPeriod.name)
            .where(AccountingPeriod.entity == company_code, AccountingPeriod.enabled.is_(True))
            .order_by(AccountingPeriod.start.desc())
        )
        .scalars()
        .first()
    )


def _build_drill_down_url(
    values: dict[str, object], company: str, ledger: str | None, period: str | None = None
) -> str | None:
    account_code = values.get("account_code")
    if account_code in (None, "", _EMPTY_CELL_VALUE):
        return None
    kwargs: dict[str, Any] = {}
    if ledger:
        kwargs["ledger"] = ledger
    if period:
        kwargs["accounting_period"] = period
    return url_for("reportes.account_movement", company=company, account_code=str(account_code), **kwargs)


def _build_voucher_url(values: dict[str, object]) -> str | None:
    """Resolve a document detail route without bypassing its own ACL checks."""
    voucher_type = str(values.get("voucher_type") or "").lower()
    voucher_id = str(values.get("voucher_id") or values.get("document_id") or values.get("document_no") or "").strip()
    endpoints = {
        "journal_entry": ("contabilidad.ver_comprobante", "identifier"),
        "payment_entry": ("bancos.bancos_pago", "payment_id"),
        "sales_invoice": ("ventas.ventas_factura_venta", "invoice_id"),
        "sales_order": ("ventas.ventas_orden_venta", "order_id"),
        "delivery_note": ("ventas.ventas_entrega", "note_id"),
        "sales_quotation": ("ventas.ventas_cotizacion", "quotation_id"),
        "purchase_invoice": ("compras.compras_factura_compra", "invoice_id"),
        "purchase_order": ("compras.compras_orden_compra", "order_id"),
        "purchase_receipt": ("compras.compras_recepcion", "receipt_id"),
        "stock_entry": ("inventario.inventario_entrada", "entry_id"),
    }
    endpoint = endpoints.get(voucher_type)
    if endpoint and voucher_id:
        endpoint_name, identifier_name = endpoint
        try:
            return url_for(endpoint_name, **{identifier_name: voucher_id})  # type: ignore[arg-type]
        except (BuildError, RuntimeError):
            # Unit/report builders also run inside an application context
            # without an active request. Keep the persisted identifier and
            # use the same detail paths that Flask generates in a request.
            fallback_paths = {
                "journal_entry": "/accounting/journal/{id}",
                "payment_entry": "/payment/{id}",
                "sales_invoice": "/sales/sales-invoice/{id}",
                "sales_order": "/sales/sales-order/{id}",
                "delivery_note": "/sales/delivery-note/{id}",
                "sales_quotation": "/sales/sales-quotation/{id}",
                "purchase_invoice": "/buying/purchase-invoice/{id}",
                "purchase_order": "/buying/purchase-order/{id}",
                "purchase_receipt": "/buying/purchase-receipt/{id}",
                "stock_entry": "/inventory/stock-entry/{id}",
            }
            return fallback_paths[voucher_type].format(id=voucher_id)
    return None


def _build_hierarchical_financial_rows(
    report_code: str, source_rows: list[dict[str, object]], company: str
) -> list[dict[str, object]]:
    if report_code not in {"trial-balance", "income-statement", "balance-sheet", "cash-flow"}:
        return source_rows
    sections_order, section_nodes, section_non_account_rows = _collect_section_nodes(source_rows)
    _enrich_section_nodes(section_nodes, company)
    return _flatten_hierarchical_rows(sections_order, section_nodes, section_non_account_rows)


def _ensure_parent_nodes_exist(
    nodes: dict[str, dict[str, object]],
    account_code: str,
    row_section: object,
) -> None:
    """Garantiza que todos los nodos padre existan en la coleccion."""
    code_parts = account_code.split(".")
    for index in range(1, len(code_parts)):
        parent_code = ".".join(code_parts[:index])
        parent_node = nodes.setdefault(
            parent_code,
            {
                "section": row_section,
                "account_code": parent_code,
                "account_name": None,
            },
        )
        if row_section and not parent_node.get("section"):
            parent_node["section"] = row_section


def _collect_section_nodes(
    source_rows: list[dict[str, object]],
) -> tuple[list[str], dict[str, dict[str, dict[str, object]]], dict[str, list[dict[str, object]]]]:
    sections_order: list[str] = []
    section_nodes: dict[str, dict[str, dict[str, object]]] = {}
    section_non_account_rows: dict[str, list[dict[str, object]]] = {}
    for row in source_rows:
        section = str(row.get("section") or "__all__")
        if section not in sections_order:
            sections_order.append(section)
        account_code = str(row.get("account_code") or "").strip()
        if not account_code:
            section_non_account_rows.setdefault(section, []).append(dict(row))
            continue
        nodes = section_nodes.setdefault(section, {})
        existing_node = nodes.get(account_code, {})
        node = {**existing_node, **dict(row)}
        node["account_code"] = account_code
        nodes[account_code] = node
        _ensure_parent_nodes_exist(nodes, account_code, row.get("section"))
    return sections_order, section_nodes, section_non_account_rows


def _build_children_map(codes: set[str]) -> dict[str, list[str]]:
    """Construye mapa de relaciones padre-hijo para códigos de cuenta."""
    children_map: dict[str, list[str]] = {}
    for code in codes:
        parent = ".".join(code.split(".")[:-1])
        if parent:
            children_map.setdefault(parent, []).append(code)
    return children_map


def _compute_numeric_fields(nodes: dict[str, dict[str, object]]) -> set[str]:
    """Calcula los campos numéricos presentes en los nodos."""
    return {
        field
        for row in nodes.values()
        for field, value in row.items()
        if field in _MONEY_COLUMNS and isinstance(value, (int, float, Decimal, str))
    }


def _get_account_names(account_codes: list[str], company: str) -> dict[str, str]:
    """Obtiene nombres de cuentas desde la base de datos."""
    return {
        account.code: account.name
        for account in database.session.execute(
            database.select(Accounts).where(Accounts.entity == company, Accounts.code.in_(account_codes))
        ).scalars()
    }


def _enrich_node_metadata(
    node: dict[str, object],
    node_code: str,
    account_name: str | None,
    children_map: dict[str, list[str]],
) -> None:
    """Enriquece un nodo individual con metadatos calculados."""
    node["account_name"] = node.get("account_name") or account_name or node_code
    node["level"] = node_code.count(".") + 1
    node["is_group"] = bool(children_map.get(node_code))


def _enrich_section_nodes(section_nodes: dict[str, dict[str, dict[str, object]]], company: str) -> None:
    for section, nodes in section_nodes.items():
        if not nodes:
            continue
        account_codes = list(nodes)
        account_names = _get_account_names(account_codes, company)
        numeric_fields = _compute_numeric_fields(nodes)
        children_map = _build_children_map(set(nodes.keys()))
        _propagate_child_amounts_to_parents(nodes, numeric_fields)
        for node_code, node in nodes.items():
            _enrich_node_metadata(node, node_code, account_names.get(node_code), children_map)


def _propagate_child_amounts_to_parents(nodes: dict[str, dict[str, object]], numeric_fields: set[str]) -> None:
    for code in sorted(nodes.keys(), key=lambda value: value.count("."), reverse=True):
        parent = ".".join(code.split(".")[:-1])
        if not parent or parent not in nodes:
            continue
        for field in numeric_fields:
            parent_amount = _to_decimal_or_zero(nodes[parent].get(field))
            child_amount = _to_decimal_or_zero(nodes[code].get(field))
            nodes[parent][field] = parent_amount + child_amount


def _find_root_codes(nodes: dict[str, dict[str, object]]) -> list[str]:
    """Encuentra los códigos raíz que no tienen padre en el mismo conjunto."""
    return sorted(
        [code for code in nodes if ".".join(code.split(".")[:-1]) not in nodes],
        key=str,
    )


def _flatten_section(
    section: str,
    nodes: dict[str, dict[str, object]],
    non_account_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    """Flattenea una seccion individual."""
    result: list[dict[str, object]] = []
    if not nodes:
        return list(non_account_rows)
    ordered_children_map = _build_children_map(set(nodes.keys()))
    root_codes = _find_root_codes(nodes)
    result.extend(_flatten_nodes_by_root(nodes, ordered_children_map, root_codes))
    # Los resúmenes de sección (por ejemplo, utilidad del período) se muestran
    # después del detalle para no interrumpir la jerarquía de cuentas.
    result.extend(non_account_rows)
    return result


def _flatten_hierarchical_rows(
    sections_order: list[str],
    section_nodes: dict[str, dict[str, dict[str, object]]],
    section_non_account_rows: dict[str, list[dict[str, object]]],
) -> list[dict[str, object]]:
    flattened_rows: list[dict[str, object]] = []
    for section in sections_order:
        nodes = section_nodes.get(section, {})
        non_account = section_non_account_rows.get(section, [])
        flattened_rows.extend(_flatten_section(section, nodes, non_account))
    return flattened_rows


def _flatten_nodes_by_root(
    nodes: dict[str, dict[str, object]],
    ordered_children_map: dict[str, list[str]],
    root_codes: list[str],
) -> list[dict[str, object]]:
    result: list[dict[str, object]] = []

    def append_node(code: str) -> None:
        result.append(dict(nodes[code]))
        for child_code in sorted(ordered_children_map.get(code, []), key=str):
            append_node(child_code)

    for root_code in root_codes:
        append_node(root_code)
    return result


def _resolve_row_level(row: dict[str, object], account_code: str) -> int:
    level_value = row.get("level")
    if isinstance(level_value, int):
        return level_value
    try:
        if level_value is not None:
            return int(str(level_value))
    except (TypeError, ValueError):
        pass
    return account_code.count(".") + 1 if account_code else 0


def _date_arg(name: str) -> date | None:
    value = request.args.get(name)
    return date.fromisoformat(value) if value else None


def _period_params() -> tuple[str | None, str | None]:
    """Devuelve los identificadores de período (inicio y fin) desde la petición.

    Acepta tanto ``accounting_period_from``/``accounting_period_to`` como la
    variante corta ``period_from``/``period_to`` para compatibilidad.
    """
    period_from = request.args.get("accounting_period_from") or request.args.get("period_from")
    period_to = request.args.get("accounting_period_to") or request.args.get("period_to")
    return period_from, period_to


def _period_picker_payload(company: str, period_from: str | None = None, period_to: str | None = None) -> dict[str, object]:
    """Construye el contexto del selector de rango de períodos en los templates.

    Devuelve la lista de períodos disponibles y el rango activo (``period_from``
    y ``period_to`` como ids). Cuando no hay criterio explícito se pre-selecciona
    el período contable actual para que el selector y el backend coincidan.
    """
    from cacao_accounting.reportes.periods import current_period_for_company, list_periods_for_company

    active_from, active_to = _period_params()
    period_from = period_from or active_from
    period_to = period_to or active_to
    periods = [
        {"id": str(period.id), "name": period.name, "is_closed": bool(period.is_closed)}
        for period in list_periods_for_company(company)
    ]
    if not period_from and not period_to:
        current = current_period_for_company(company)
        if current is not None:
            period_from = period_to = str(current.id)
    return {
        "periods": periods,
        "period_from": period_from or "",
        "period_to": period_to or period_from or "",
    }


def _resolve_as_of_date(company: str) -> date | None:
    """Resuelve la fecha de corte desde el rango de períodos o el parámetro manual.

    El corte de los reportes "as of" es el último día del período final del
    rango seleccionado: el backend lo deriva de ``AccountingPeriod`` y nunca
    confía en fechas calculadas por el navegador. Si el llamador envía un
    ``as_of_date`` junto con cualquier criterio de período, la fecha manual
    debe coincidir con el extremo resuelto o la petición se rechaza. Sin
    ningún criterio de período se conserva la compat con URL legadas.
    """
    manual = _date_arg("as_of_date")
    period_from, period_to = _period_params()
    has_period_criterion = bool(period_from or period_to or request.args.get("accounting_period"))
    if has_period_criterion:
        from cacao_accounting.reportes.periods import (
            reject_manual_date_overrides,
            resolve_period_range,
        )

        if not (period_from or period_to):
            legacy_name = request.args.get("accounting_period")
            if legacy_name:
                period_from = period_to = _resolve_period_id_by_name(company, legacy_name)
        period_range = resolve_period_range(company, period_from, period_to)
        if period_range is not None:
            reject_manual_date_overrides(request.args, period_range)
            return manual if manual is not None else period_range.period_end
    if manual is not None:
        return manual
    from cacao_accounting.reportes.periods import resolve_period_range

    default_range = resolve_period_range(company, None, None)
    return default_range.period_end if default_range is not None else None


def _resolve_period_id_by_name(company: str, period_name: str) -> str | None:
    """Resuelve el id de un ``AccountingPeriod`` por nombre dentro de la compañía.

    Devuelve ``None`` si no existe un período con ese nombre para evitar que un
    id inválido llegue a ``resolve_period_range`` y aborte la petición.
    """
    from cacao_accounting.database import AccountingPeriod, database
    from sqlalchemy import select

    return database.session.execute(
        select(AccountingPeriod.id).where(AccountingPeriod.entity == company, AccountingPeriod.name == period_name)
    ).scalar_one_or_none()


def _resolve_date_bounds(company: str) -> tuple[date | None, date | None]:
    """Resuelve ``(date_from, date_to)`` desde el rango de períodos o parámetros manuales.

    Si el llamador envía fechas manuales junto con cualquier criterio de
    período, ambos extremos deben coincidir con el rango resuelto. Sin
    ningún criterio de período, las fechas manuales se respetan para no
    romper compat con URL legadas.
    """
    date_from = _date_arg("date_from")
    date_to = _date_arg("date_to")
    period_from, period_to = _period_params()
    has_period_criterion = bool(period_from or period_to or request.args.get("accounting_period"))
    if has_period_criterion and (date_from is not None or date_to is not None):
        from cacao_accounting.reportes.periods import (
            reject_manual_date_overrides,
            resolve_period_range,
        )

        if not (period_from or period_to):
            legacy_name = request.args.get("accounting_period")
            if legacy_name:
                period_from = period_to = _resolve_period_id_by_name(company, legacy_name)
        period_range = resolve_period_range(company, period_from, period_to)
        if period_range is not None:
            reject_manual_date_overrides(request.args, period_range)
            return period_range.period_start, period_range.period_end
    if date_from is not None or date_to is not None:
        return date_from, date_to
    if period_from or period_to:
        from cacao_accounting.reportes.periods import reject_manual_date_overrides, resolve_period_range

        period_range = resolve_period_range(company, period_from, period_to)
        if period_range is not None:
            reject_manual_date_overrides(request.args, period_range)
            return period_range.period_start, period_range.period_end
    from cacao_accounting.reportes.periods import resolve_period_range

    default_range = resolve_period_range(company, None, None)
    if default_range is not None:
        return default_range.period_start, default_range.period_end
    return None, None


def _int_arg(name: str, default: int) -> int:
    value = request.args.get(name, default=str(default))
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _bool_arg(name: str) -> bool:
    return request.args.get(name, "").lower() in {"1", "true", "yes", "on"}


def _financial_filters() -> FinancialReportFilters:
    company_code = _resolve_company(request.args.get("company", "cacao"))
    show_cancellations = _bool_arg("show_cancellations")
    requested_status = request.args.get("status") or "submitted"
    status = None if show_cancellations else requested_status
    ledger = request.args.get("ledger") or _default_ledger_for_company(company_code)
    period_from, period_to = _period_params()
    requested_period = request.args.get("accounting_period")
    accounting_period = requested_period or _default_period_for_company(company_code)
    if requested_period is None and not (period_from or period_to):
        from cacao_accounting.reportes.periods import resolve_period_range

        default_range = resolve_period_range(company_code, None, None)
        if default_range is not None:
            period_from, period_to = default_range.from_id, default_range.to_id
            accounting_period = default_range.to_name
    if period_from or period_to:
        from cacao_accounting.reportes.periods import reject_manual_date_overrides, resolve_period_range

        period_range = resolve_period_range(company_code, period_from, period_to)
        if period_range is not None:
            reject_manual_date_overrides(request.args, period_range)
            if requested_period is None:
                accounting_period = None if not period_range.single_period else period_range.to_name
    return FinancialReportFilters(
        company=company_code,
        ledger=ledger,
        accounting_period=accounting_period,
        period_from=period_from,
        period_to=period_to or period_from,
        voucher_number=request.args.get("voucher_number") or None,
        account_code=request.args.get("account_code") or None,
        account_from=request.args.get("account_from") or None,
        account_to=request.args.get("account_to") or None,
        cost_center_code=request.args.get("cost_center_code") or None,
        unit_code=request.args.get("unit_code") or None,
        project_code=request.args.get("project_code") or None,
        party_type=request.args.get("party_type") or None,
        party_id=request.args.get("party_id") or None,
        voucher_type=request.args.get("voucher_type") or None,
        status=status,
        include_cancellations=show_cancellations or (status == "cancelled"),
        include_running_balance=_bool_arg("include_running_balance"),
        page=max(_int_arg("page", 1), 1),
        page_size=max(_int_arg("page_size", 100), 1),
        sort_by=request.args.get("sort_by", "posting_date"),
        sort_dir=request.args.get("sort_dir", "asc"),
        export_all=False,
        include_descendants=_bool_arg("include_descendants"),
    )


def _should_run_financial_report() -> bool:
    """Evita cargar datos al abrir la vista sin aplicar filtros explícitos."""
    if request.args.get("apply_filters") in {"1", "true", "yes", "on"}:
        return True
    return "export" in request.args


def _empty_financial_report() -> PaginatedReport:
    return PaginatedReport(rows=[], totals={}, columns=[], total_rows=0, page=1, page_size=100, ledger_currency=None)


def _report_to_matrix(report) -> tuple[list[str], list[list[object]]]:
    rows = getattr(report, "rows", [])
    columns = getattr(report, "columns", None) or (list(rows[0].values.keys()) if rows else [])
    data_rows = [[row.values.get(column) for column in columns] for row in rows]
    return columns, data_rows


def _write_operational_report_xlsx(report, report_code: str, title: str, filter_payload: dict[str, object]) -> bytes:
    columns, rows = _report_to_matrix(report)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report_code[:31]
    sheet.append([title])
    sheet.append([_("Fecha de generación"), date.today().isoformat()])
    sheet.append([_("Usuario"), getattr(current_user, "user", "")])
    sheet.append([])
    if columns:
        sheet.append([_column_label(column, report.ledger_currency) for column in columns])
        sheet.freeze_panes = "A5"
    for row in rows:
        sheet.append([_format_cell(column, row[index], report.ledger_currency) for index, column in enumerate(columns)])
    if report.totals:
        sheet.append([])
    for total_name, total_value in report.totals.items():
        sheet.append(
            [
                _("TOTAL"),
                _column_label(total_name, report.ledger_currency),
                _format_cell(total_name, total_value, report.ledger_currency),
            ]
        )
    for column_cells in sheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        max_length = max((len(value) for value in values), default=10)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 60)
    for column in range(1, sheet.max_column + 1):
        sheet.cell(row=4, column=column).alignment = Alignment(horizontal="center")

    filters_sheet = workbook.create_sheet(_("Filtros"))
    filters_sheet.append([_("Filtro"), _("Valor")])
    for key, value in filter_payload.items():
        if value in (None, "", False):
            continue
        filters_sheet.append([_(key.replace("_", " ").title()), str(value)])
    filters_sheet.freeze_panes = "A2"
    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def _export_operational_report(report, report_code: str, title: str, filter_payload: dict[str, object]):
    export_format = request.args.get("export")
    if export_format not in {"csv", "xlsx"}:
        return None

    columns, rows = _report_to_matrix(report)
    if export_format == "csv":
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(columns)
        writer.writerows(rows)
        return send_file(
            BytesIO(buffer.getvalue().encode("utf-8")),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"{report_code}.csv",
        )

    xlsx_content = _write_operational_report_xlsx(report, report_code, title, filter_payload)
    return send_file(
        BytesIO(xlsx_content),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{report_code}.xlsx",
    )


def _write_financial_report_xlsx(report, report_code: str, title: str, report_filters: FinancialReportFilters) -> bytes:
    columns, rows = _report_to_matrix(report)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report_code[:31]
    sheet.append([title])
    sheet.append([_("Fecha de generación"), date.today().isoformat()])
    sheet.append([_("Usuario"), getattr(current_user, "user", "")])
    sheet.append([])
    if columns:
        localized_headers = [_column_label(column, report.ledger_currency) for column in columns]
        sheet.append(localized_headers)
        sheet.freeze_panes = "A5"
    for row in rows:
        sheet.append([_format_cell(column, row[index], report.ledger_currency) for index, column in enumerate(columns)])
    if report.totals:
        sheet.append([])
    for total_name, total_value in report.totals.items():
        sheet.append(
            [
                _("TOTAL"),
                _column_label(total_name, report.ledger_currency),
                _format_cell(total_name, total_value, report.ledger_currency),
            ]
        )
    for column_cells in sheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        max_length = max((len(value) for value in values), default=10)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 60)
    for column in range(1, sheet.max_column + 1):
        sheet.cell(row=4, column=column).alignment = Alignment(horizontal="center")

    filters_sheet = workbook.create_sheet(_("Filtros"))
    filters_sheet.append([_("Filtro"), _("Valor")])
    for key in _FINANCIAL_FILTER_FIELDS:
        value = getattr(report_filters, key, None)
        if value in (None, "", False):
            continue
        filters_sheet.append([_(key.replace("_", " ").title()), str(value)])
    filters_sheet.freeze_panes = "A2"
    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def _export_financial_report(report, report_code: str, title: str, report_filters: FinancialReportFilters):
    export_format = request.args.get("export")
    if export_format not in {"csv", "xlsx"}:
        return None

    columns, rows = _report_to_matrix(report)
    if export_format == "csv":
        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(columns)
        writer.writerows(rows)
        return send_file(
            BytesIO(buffer.getvalue().encode("utf-8")),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"{report_code}.csv",
        )

    xlsx_content = _write_financial_report_xlsx(report, report_code, title, report_filters)
    return send_file(
        BytesIO(xlsx_content),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{report_code}.xlsx",
    )


def _compute_display_columns(report, report_code: str, saved_view: str) -> list[str]:
    selected_columns = request.args.getlist("visible_columns")
    if not selected_columns:
        selected_columns = _preferred_columns_from_view(report_code, saved_view)
    columns = report.columns or []
    if selected_columns:
        columns = [column for column in columns if column in selected_columns]
    if report_code == "trial-balance":
        columns = [column for column in columns if column != "level"]
    display_columns = [
        column
        for column in columns
        if any((row.values.get(column) not in (None, "", "—") for row in report.rows)) or column in _ALWAYS_VISIBLE_COLUMNS
    ]
    return display_columns if display_columns else columns


def _compute_all_columns(report, report_code: str) -> list[str]:
    extra_columns = ["reference_type", "is_reversal", "reversal_of"]
    all_columns = list(dict.fromkeys([*(report.columns or []), *extra_columns]))
    if report_code == "trial-balance":
        all_columns = [column for column in all_columns if column != "level"]
    return all_columns


def _build_row_metadata(
    source_rows: list[dict[str, object]],
    report_filters: FinancialReportFilters,
) -> list[dict[str, object]]:
    child_counts: dict[str, int] = {}
    for row in source_rows:
        account_code = str(row.get("account_code") or "")
        if not account_code:
            continue
        parent_code = ".".join(account_code.split(".")[:-1])
        if parent_code:
            child_counts[parent_code] = child_counts.get(parent_code, 0) + 1
    row_metadata = []
    for row in source_rows:
        account_code = str(row.get("account_code") or "")
        parent_code = ".".join(account_code.split(".")[:-1]) if account_code else ""
        row_metadata.append(
            {
                "code": account_code,
                "parent": parent_code,
                "has_children": bool(child_counts.get(account_code)),
                "level": _resolve_row_level(row, account_code),
                "drilldown_url": _build_drill_down_url(
                    row, report_filters.company, report_filters.ledger, report_filters.accounting_period
                ),
                "voucher_url": _build_voucher_url(row),
                "is_group": bool(row.get("is_group")),
            }
        )
    return row_metadata


def _build_display_rows(
    source_rows: list[dict[str, object]],
    row_metadata: list[dict[str, object]],
    display_columns: list[str],
    ledger_currency: str | None,
) -> list[dict[str, object]]:
    display_rows: list[dict[str, object]] = []
    for index, row in enumerate(source_rows):
        formatted_row: dict[str, object] = {
            column: _format_cell(column, row.get(column), ledger_currency) for column in display_columns
        }
        formatted_row["__meta"] = row_metadata[index]
        display_rows.append(formatted_row)
    return display_rows


def _apply_grouping(
    display_rows: list[dict[str, object]],
    source_rows: list[dict[str, object]],
    report,
    report_code: str,
    saved_view: str,
) -> list[dict[str, object]]:
    group_by = request.args.get("group_by") or _preferred_group_by_from_view(report_code, saved_view)
    if not (report_code == "account-movement" and group_by and group_by in (report.columns or [])):
        return display_rows
    return _group_rows_by_field(display_rows, source_rows, group_by, report.ledger_currency)


def _group_rows_by_field(
    display_rows: list[dict[str, object]],
    source_rows: list[dict[str, object]],
    group_by: str,
    ledger_currency: str | None,
) -> list[dict[str, object]]:
    grouped_rows: list[dict[str, object]] = []
    current_group: str | None = None
    group_debit = Decimal("0")
    group_credit = Decimal("0")
    for index, row in enumerate(display_rows):
        raw_group_value = source_rows[index].get(group_by)
        group_value = _format_cell(group_by, raw_group_value, ledger_currency)
        if group_value != current_group:
            if current_group is not None:
                grouped_rows.append(_build_group_subtotal_row(group_debit, group_credit, ledger_currency))
            group_title = _(group_by.replace("_", " ").title())
            grouped_rows.append({"__row_type": "group", "__group_title": f"{group_title}: {group_value}"})
            current_group = group_value
            group_debit = Decimal("0")
            group_credit = Decimal("0")
        group_debit = _add_to_decimal_or_zero(group_debit, row.get("debit"))
        group_credit = _add_to_decimal_or_zero(group_credit, row.get("credit"))
        grouped_rows.append(row)
    if current_group is not None:
        grouped_rows.append(_build_group_subtotal_row(group_debit, group_credit, ledger_currency))
    return grouped_rows


def _build_group_subtotal_row(debit: Decimal, credit: Decimal, ledger_currency: str | None) -> dict[str, object]:
    return {
        "__row_type": "group_subtotal",
        "__group_title": _("Subtotal"),
        "debit": _format_cell("debit", debit, ledger_currency),
        "credit": _format_cell("credit", credit, ledger_currency),
    }


def _add_to_decimal_or_zero(current: Decimal, value: object) -> Decimal:
    try:
        return current + Decimal(str(value or "0").replace(",", "").replace("(", "-").replace(")", ""))
    except DecimalException:
        return current


def _render_financial_report(
    report_code: str,
    report_title: str,
    report,
    report_filters: FinancialReportFilters,
    saved_view: str,
    saved_views: list[str],
):
    export_response = _export_financial_report(report, report_code, report_title, report_filters)
    if export_response is not None:
        return export_response
    display_columns = _compute_display_columns(report, report_code, saved_view)
    all_columns = _compute_all_columns(report, report_code)
    allow_column_selection = report_code in {"account-movement", "account-summary"}
    group_by = request.args.get("group_by") or _preferred_group_by_from_view(report_code, saved_view)
    display_headers = {column: _column_label(column, report.ledger_currency) for column in display_columns}
    all_column_headers = {column: _column_label(column, report.ledger_currency) for column in all_columns}
    source_rows = [dict(row.values) for row in report.rows]
    source_rows = _build_hierarchical_financial_rows(report_code, source_rows, report_filters.company)
    row_metadata = _build_row_metadata(source_rows, report_filters)
    display_rows = _build_display_rows(source_rows, row_metadata, display_columns, report.ledger_currency)
    grouped_rows = _apply_grouping(display_rows, source_rows, report, report_code, saved_view)
    display_totals = {key: _format_cell(key, value, report.ledger_currency) for key, value in report.totals.items()}
    unclassified_count = _unclassified_accounts_count(report.totals)
    period_picker = _period_picker_payload(
        report_filters.company, getattr(report_filters, "period_from", None), getattr(report_filters, "period_to", None)
    )
    return render_template(
        "reportes/financial_report.html",
        titulo=f"{report_title} - {APPNAME}",
        report_code=report_code,
        report_title=report_title,
        rows=report.rows,
        columns=display_columns,
        display_headers=display_headers,
        all_column_headers=all_column_headers,
        display_rows=grouped_rows,
        totals=display_totals,
        total_rows=report.total_rows,
        page=report.page,
        page_size=report.page_size,
        ledger_currency=report.ledger_currency,
        context_summary=_build_context_summary(report, report_filters),
        report_filters=report_filters,
        right_align_columns=_RIGHT_ALIGN_COLUMNS,
        is_balanced=_is_report_balanced(report.totals),
        unclassified_accounts=unclassified_count,
        unclassified_warning=_build_unclassified_warning(unclassified_count),
        saved_view=saved_view,
        saved_views=saved_views,
        selected_columns=display_columns,
        all_columns=all_columns,
        allow_column_selection=allow_column_selection,
        group_by=group_by,
        periods=period_picker["periods"],
        period_from=period_picker["period_from"],
        period_to=period_picker["period_to"],
    )


def _render_operational_framework(
    report_code: str,
    report_title: str,
    report,
    *,
    module_home_endpoint: str,
    module_home_label: str,
    filter_mode: str,
    filter_state: dict[str, object],
    context_summary: dict[str, str],
):
    export_response = _export_operational_report(report, report_code, report_title, filter_state)
    if export_response is not None:
        return export_response
    rows = list(getattr(report, "rows", []))
    totals_raw = getattr(report, "totals", {})
    ledger_currency = getattr(report, "ledger_currency", None)
    columns = getattr(report, "columns", None) or (list(rows[0].values.keys()) if rows else [])
    requested_page = max(request.args.get("page", 1, type=int) or 1, 1)
    requested_page_size = min(max(request.args.get("page_size", 100, type=int) or 100, 1), 500)
    report_total_rows = getattr(report, "total_rows", 0) or 0
    report_page_size = getattr(report, "page_size", 0) or 0
    is_service_paginated = report_total_rows > len(rows) and report_page_size > 0
    page = getattr(report, "page", requested_page) if is_service_paginated else requested_page
    page_size = report_page_size if is_service_paginated else requested_page_size
    total_rows = report_total_rows or len(rows)
    page_rows = rows if is_service_paginated else rows[(page - 1) * page_size : page * page_size]
    display_headers = {column: _column_label(column, ledger_currency) for column in columns}
    display_rows = [
        {column: _format_cell(column, row.values.get(column), ledger_currency) for column in columns} for row in page_rows
    ]
    drill_down_urls = [
        {column: _build_voucher_url(row.values) for column in ("document_no", "voucher_id")} for row in page_rows
    ]
    totals = {key: _format_cell(key, value, ledger_currency) for key, value in totals_raw.items()}
    query = request.args.to_dict()

    def _page_url(target_page: int) -> str:
        query["page"] = str(target_page)
        query["page_size"] = str(page_size)
        endpoint = request.endpoint or "reportes.account_summary"
        return str(url_for(endpoint, **query))  # type: ignore[arg-type]

    total_pages = max((total_rows + page_size - 1) // page_size, 1)
    company = str(filter_state.get("company") or request.args.get("company", "cacao"))
    period_picker = _period_picker_payload(
        company,
        str(filter_state.get("accounting_period_from") or ""),
        str(filter_state.get("accounting_period_to") or ""),
    )
    return render_template(
        "reportes/operational_report.html",
        titulo=f"{report_title} - {APPNAME}",
        report_title=report_title,
        report_code=report_code,
        columns=columns,
        display_headers=display_headers,
        display_rows=display_rows,
        drill_down_urls=drill_down_urls,
        totals=totals,
        page=page,
        total_pages=total_pages,
        total_rows=total_rows,
        previous_page_url=_page_url(page - 1) if page > 1 else None,
        next_page_url=_page_url(page + 1) if page < total_pages else None,
        filter_mode=filter_mode,
        filter_state=filter_state,
        context_summary=context_summary,
        right_align_columns=_RIGHT_ALIGN_COLUMNS,
        module_home_url=url_for(module_home_endpoint),
        module_home_label=module_home_label,
        periods=period_picker["periods"],
        period_from=period_picker["period_from"],
        period_to=period_picker["period_to"],
    )


def _operational_filters() -> OperationalReportFilters:
    period_from, period_to = _period_params()
    company = request.args.get("company", "cacao")
    date_from, date_to = _resolve_date_bounds(company)
    return OperationalReportFilters(
        company=company,
        date_from=date_from,
        date_to=date_to,
        party_id=request.args.get("party_id") or None,
        item_code=request.args.get("item_code") or None,
        warehouse=request.args.get("warehouse") or None,
        period_from=period_from,
        period_to=period_to or period_from,
    )


def _render_operational_report(report_name: str, report):
    company = request.args.get("company", "cacao")
    period_picker = _period_picker_payload(company)
    return render_template(
        REPORT_TABLE_HTML,
        titulo=report_name + " - " + APPNAME,
        report_title=report_name,
        rows=report.rows,
        totals=report.totals,
        total_rows=getattr(report, "total_rows", len(getattr(report, "rows", []))),
        periods=period_picker["periods"],
        period_from=period_picker["period_from"],
        period_to=period_picker["period_to"],
    )


def _operational_context_summary(report, **values: object) -> dict[str, str]:
    total_rows = getattr(report, "total_rows", len(getattr(report, "rows", [])))
    summary = {"company": str(values.get("company") or "—"), "records": str(total_rows)}
    for key, value in values.items():
        if key == "company":
            continue
        summary[key] = "—" if value in (None, "") else str(value)
    return summary
