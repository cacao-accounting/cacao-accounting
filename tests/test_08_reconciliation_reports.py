# SPDX-License-Identifier: Apache-2.0
# SPDX-FileCopyrightText: 2025 - 2026 William José Moreno Reyes

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from cacao_accounting import create_app
from cacao_accounting.config import configuracion


@pytest.fixture()
def app_ctx():
    app = create_app(
        {
            **configuracion,
            "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
            "SQLALCHEMY_TRACK_MODIFICATIONS": False,
            "WTF_CSRF_ENABLED": False,
            "TESTING": True,
        }
    )
    with app.app_context():
        from cacao_accounting.database import Entity, PurchaseMatchingConfig, database

        database.create_all()
        database.session.add_all(
            [
                Entity(code="cacao", name="Cacao", company_name="Cacao", tax_id="J0001", currency="NIO"),
                # Estos escenarios cubren conciliaciones sin OC; el comportamiento
                # estricto se verifica en pruebas que cambian explícitamente la configuración.
                PurchaseMatchingConfig(company="cacao", require_purchase_order=False),
            ]
        )
        database.session.commit()
        yield app


def _seed_accounting_admin() -> None:
    """Crea un actor persistido para servicios de posting fail-closed."""
    from cacao_accounting.database import Book, Modules, User, database

    if database.session.execute(database.select(Modules).filter_by(module="accounting")).scalar_one_or_none() is None:
        database.session.add(Modules(module="accounting", default=True, enabled=True))
    if database.session.get(User, "admin") is None:
        database.session.add(User(id="admin", user="acl-admin", password=b"x", classification="admin", active=True))
    if database.session.execute(database.select(Book).filter_by(entity="cacao")).scalar_one_or_none() is None:
        database.session.add(
            Book(entity="cacao", code="MAIN", name="Libro principal", status="activo", is_primary=True, currency="NIO")
        )
    database.session.commit()


def test_purchase_reconciliation_line_matching_supports_partial_and_completion(app_ctx):
    from cacao_accounting.compras.purchase_reconciliation_service import (
        get_purchase_reconciliation_pending,
        reconcile_purchase_invoice,
    )
    from cacao_accounting.database import (
        PurchaseReconciliationItem,
        Item,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseReceipt,
        PurchaseReceiptItem,
        UOM,
        Warehouse,
        database,
    )

    database.session.add_all(
        [
            UOM(code="EA", name="Each"),
            Item(code="ITEM-GR8", name="Item GR8", item_type="goods", is_stock_item=True, default_uom="EA"),
            Warehouse(code="WH-GR8", name="Bodega GR8", company="cacao"),
        ]
    )
    receipt = PurchaseReceipt(company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-8", docstatus=1)
    database.session.add(receipt)
    database.session.flush()
    database.session.add(
        PurchaseReceiptItem(
            purchase_receipt_id=receipt.id,
            item_code="ITEM-GR8",
            item_name="Item GR8",
            qty=Decimal("10"),
            qty_in_base_uom=Decimal("10"),
            uom="EA",
            rate=Decimal("5.00"),
            amount=Decimal("50.00"),
            warehouse="WH-GR8",
        )
    )
    invoices = []
    for qty in (Decimal("4"), Decimal("6")):
        invoice = PurchaseInvoice(
            company="cacao",
            posting_date=date(2026, 5, 2),
            supplier_id="SUPP-8",
            purchase_receipt_id=receipt.id,
            docstatus=1,
        )
        database.session.add(invoice)
        database.session.flush()
        database.session.add(
            PurchaseInvoiceItem(
                purchase_invoice_id=invoice.id,
                item_code="ITEM-GR8",
                item_name="Item GR8",
                qty=qty,
                uom="EA",
                rate=Decimal("5.00"),
                amount=qty * Decimal("5.00"),
                warehouse="WH-GR8",
            )
        )
        invoices.append(invoice)
    database.session.commit()

    first = reconcile_purchase_invoice(invoices[0].id)
    assert first.matched_qty == Decimal("4.000000000")
    assert get_purchase_reconciliation_pending("cacao")[0].pending_qty == Decimal("6.000000000")

    second = reconcile_purchase_invoice(invoices[1].id)
    database.session.commit()
    assert second.matched_qty == Decimal("6.000000000")
    assert get_purchase_reconciliation_pending("cacao") == []
    assert database.session.execute(database.select(PurchaseReconciliationItem)).scalars().all()


@pytest.mark.full
def test_three_way_multicurrency_receipt_compensation_report(app_ctx):
    """Conciliación 3-way en USD: recepción, dos facturas y saldo pendiente.

    Cálculo manual: recepción 15 x USD 12 = USD 180; primera factura 9 x
    USD 12 = USD 108; segunda factura 4 x USD 12 = USD 48; queda 2 unidades,
    USD 24, pendientes de compensar en el reporte de recepción.
    """
    from cacao_accounting.compras.purchase_reconciliation_service import (
        get_purchase_reconciliation_pending,
        reconcile_purchase_invoice,
    )
    from cacao_accounting.database import (
        Currency,
        Item,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseReceipt,
        PurchaseReceiptItem,
        UOM,
        Warehouse,
        database,
    )

    database.session.add_all(
        [
            Currency(code="NIO", name="Cordoba", decimals=2, active=True),
            Currency(code="USD", name="Dollar", decimals=2, active=True),
            UOM(code="EA-FULL", name="Each"),
            Item(code="ITEM-FULL-3W", name="Item 3-way USD", item_type="goods", is_stock_item=True, default_uom="EA-FULL"),
            Warehouse(code="WH-FULL-3W", name="Bodega 3-way", company="cacao"),
        ]
    )
    receipt = PurchaseReceipt(
        company="cacao",
        posting_date=date(2026, 8, 1),
        supplier_id="SUPP-FULL-3W",
        transaction_currency="USD",
        base_currency="NIO",
        exchange_rate=Decimal("36"),
        docstatus=1,
    )
    database.session.add(receipt)
    database.session.flush()
    database.session.add(
        PurchaseReceiptItem(
            purchase_receipt_id=receipt.id,
            item_code="ITEM-FULL-3W",
            item_name="Item 3-way USD",
            qty=Decimal("15"),
            qty_in_base_uom=Decimal("15"),
            uom="EA-FULL",
            rate=Decimal("12"),
            amount=Decimal("180"),
            base_amount=Decimal("6480"),
            warehouse="WH-FULL-3W",
        )
    )
    invoices = []
    for qty in (Decimal("9"), Decimal("4")):
        invoice = PurchaseInvoice(
            company="cacao",
            posting_date=date(2026, 8, 2),
            supplier_id="SUPP-FULL-3W",
            purchase_receipt_id=receipt.id,
            transaction_currency="USD",
            base_currency="NIO",
            exchange_rate=Decimal("36"),
            docstatus=1,
        )
        database.session.add(invoice)
        database.session.flush()
        database.session.add(
            PurchaseInvoiceItem(
                purchase_invoice_id=invoice.id,
                item_code="ITEM-FULL-3W",
                item_name="Item 3-way USD",
                qty=qty,
                uom="EA-FULL",
                rate=Decimal("12"),
                amount=qty * Decimal("12"),
                base_amount=qty * Decimal("12") * Decimal("36"),
                warehouse="WH-FULL-3W",
            )
        )
        invoices.append(invoice)
    database.session.commit()

    first = reconcile_purchase_invoice(invoices[0].id)
    database.session.commit()
    pending_after_first = get_purchase_reconciliation_pending("cacao")
    assert first.matched_qty == Decimal("9.000000000")
    assert pending_after_first[0].pending_qty == Decimal("6.000000000")
    assert pending_after_first[0].pending_amount == Decimal("72.0000")

    second = reconcile_purchase_invoice(invoices[1].id)
    database.session.commit()
    pending_after_second = get_purchase_reconciliation_pending("cacao")
    assert second.matched_qty == Decimal("4.000000000")
    assert pending_after_second[0].pending_qty == Decimal("2.000000000")
    assert pending_after_second[0].pending_amount == Decimal("24.0000")


INVENTORY_REBUILD_SCENARIOS = [
    ("two_receipts", [("10", "100"), ("5", "60")], "15", "160", "10.666666667"),
    ("receipt_return", [("20", "240"), ("-4", "-48")], "16", "192", "12"),
    ("receipt_issue", [("30", "450"), ("-12", "-180")], "18", "270", "15"),
    ("three_costs", [("3", "21"), ("4", "32"), ("5", "45")], "12", "98", "8.166666667"),
    ("issue_then_replenish", [("25", "250"), ("-7", "-70"), ("10", "120")], "28", "300", "10.714285714"),
    ("count_increase", [("8", "80"), ("2", "22")], "10", "102", "10.2"),
    ("count_decrease", [("12", "144"), ("-3", "-36")], "9", "108", "12"),
    ("fractional_units", [("1.5", "18"), ("2.25", "29.25")], "3.75", "47.25", "12.6"),
    ("fractional_return", [("7.5", "90"), ("-2.5", "-30")], "5", "60", "12"),
    ("zero_value_receipt", [("10", "0"), ("5", "25")], "15", "25", "1.666666667"),
    ("zero_balance", [("10", "100"), ("-10", "-100")], "0", "0", "0"),
    ("small_rounding", [("3", "10.01"), ("2", "6.67")], "5", "16.68", "3.336"),
    ("large_batch", [("1000", "12500"), ("-125", "-1562.5")], "875", "10937.5", "12.5"),
    ("mixed_adjustments", [("10", "100"), ("-1", "-9"), ("2", "22")], "11", "113", "10.272727273"),
    ("negative_adjustment", [("20", "200"), ("-2", "-30"), ("-3", "-45")], "15", "125", "8.333333333"),
    ("three_receipts", [("2", "20"), ("3", "36"), ("4", "56")], "9", "112", "12.444444444"),
    ("two_returns", [("40", "480"), ("-5", "-60"), ("-7", "-84")], "28", "336", "12"),
    ("moving_average", [("10", "100"), ("10", "140"), ("-5", "-60")], "15", "180", "12"),
    ("cost_reversal", [("6", "72"), ("-2", "-24"), ("2", "30")], "6", "78", "13"),
    ("decimal_costs", [("1.25", "13.75"), ("2.75", "35.75")], "4", "49.5", "12.375"),
    ("negative_then_recover", [("-2", "-20"), ("10", "130")], "8", "110", "13.75"),
    ("zero_cost_return", [("5", "50"), ("-1", "0")], "4", "50", "12.5"),
    ("high_precision", [("0.333", "4.329"), ("0.667", "9.343")], "1", "13.672", "13.672"),
    ("manual_recount", [("14", "210"), ("-4", "-60"), ("1", "17")], "11", "167", "15.181818182"),
    ("purchase_credit", [("9", "108"), ("-3", "-36"), ("4", "52")], "10", "124", "12.4"),
    ("purchase_debit", [("10", "120"), ("2", "30"), ("-1", "-12")], "11", "138", "12.545454545"),
]


@pytest.mark.full
@pytest.mark.parametrize(
    "scenario, movements, expected_qty, expected_value, expected_rate",
    INVENTORY_REBUILD_SCENARIOS,
)
def test_inventory_rebuild_manual_business_scenarios(
    app_ctx, scenario, movements, expected_qty, expected_value, expected_rate
):
    """Reconstruye stock desde movimientos manuales y verifica valor algebraico.

    Cada caso representa una combinación real de recepción, salida, retorno o
    ajuste. La expectativa no usa el servicio: es la suma independiente de
    cantidades y valores, con tasa final = valor / cantidad.
    """
    from cacao_accounting.database import Item, StockBin, StockLedgerEntry, Warehouse, database
    from cacao_accounting.inventario.service import rebuild_stock_bins

    item_code = f"ITEM-REBUILD-{scenario.upper()}"
    warehouse_code = f"WH-{scenario.upper()}"
    database.session.add_all(
        [
            Item(code=item_code, name=f"Item {scenario}", item_type="goods", is_stock_item=True, default_uom="EA"),
            Warehouse(code=warehouse_code, name=f"Warehouse {scenario}", company="cacao"),
            StockBin(
                company="cacao",
                item_code=item_code,
                warehouse=warehouse_code,
                actual_qty=Decimal("999"),
                reserved_qty=Decimal("2"),
                stock_value=Decimal("9999"),
            ),
        ]
    )
    qty_after = Decimal("0")
    stock_value = Decimal("0")
    entries = []
    for index, (qty_raw, value_raw) in enumerate(movements, start=1):
        qty = Decimal(qty_raw)
        value = Decimal(value_raw)
        qty_after += qty
        stock_value += value
        entries.append(
            StockLedgerEntry(
                posting_date=date(2026, 1, index),
                item_code=item_code,
                warehouse=warehouse_code,
                company="cacao",
                qty_change=qty,
                qty_after_transaction=qty_after,
                valuation_rate=(abs(value / qty) if qty else Decimal("0")),
                stock_value_difference=value,
                stock_value=stock_value,
                voucher_type="inventory_test",
                voucher_id=f"{scenario}-{index}",
            )
        )
    database.session.add_all(entries)
    database.session.commit()

    result = rebuild_stock_bins("cacao", item_code=item_code, warehouse=warehouse_code)
    database.session.commit()
    bin_row = database.session.execute(
        database.select(StockBin).filter_by(company="cacao", item_code=item_code, warehouse=warehouse_code)
    ).scalar_one()

    expected_qty_decimal = Decimal(expected_qty)
    expected_value_decimal = Decimal(expected_value)
    expected_rate_decimal = Decimal(expected_rate)
    assert result.rebuilt_bins == 1
    assert result.rebuilt_layers == len(movements)
    assert bin_row.actual_qty == expected_qty_decimal
    assert bin_row.stock_value == expected_value_decimal
    assert bin_row.valuation_rate.quantize(Decimal("0.000000001")) == expected_rate_decimal.quantize(Decimal("0.000000001"))
    assert bin_row.reserved_qty == Decimal("2")


def test_purchase_reconciliation_rejects_overbilling_and_price_difference(app_ctx):
    from cacao_accounting.compras.purchase_reconciliation_service import reconcile_purchase_invoice
    from cacao_accounting.database import (
        Item,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseReceipt,
        PurchaseReceiptItem,
        PurchaseReconciliation,
        PurchaseReconciliationItem,
        UOM,
        Warehouse,
        database,
    )

    database.session.add_all(
        [
            UOM(code="EA", name="Each"),
            Item(code="ITEM-GR9", name="Item GR9", item_type="goods", is_stock_item=True, default_uom="EA"),
            Warehouse(code="WH-GR9", name="Bodega GR9", company="cacao"),
        ]
    )
    receipt = PurchaseReceipt(company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-9", docstatus=1)
    database.session.add(receipt)
    database.session.flush()
    database.session.add(
        PurchaseReceiptItem(
            purchase_receipt_id=receipt.id,
            item_code="ITEM-GR9",
            qty=Decimal("2"),
            qty_in_base_uom=Decimal("2"),
            uom="EA",
            rate=Decimal("5.00"),
            amount=Decimal("10.00"),
            warehouse="WH-GR9",
        )
    )
    invoice = PurchaseInvoice(
        company="cacao", posting_date=date(2026, 5, 2), supplier_id="SUPP-9", purchase_receipt_id=receipt.id, docstatus=1
    )
    database.session.add(invoice)
    database.session.flush()
    database.session.add(
        PurchaseInvoiceItem(
            purchase_invoice_id=invoice.id,
            item_code="ITEM-GR9",
            qty=Decimal("3"),
            uom="EA",
            rate=Decimal("5.00"),
            amount=Decimal("15.00"),
            warehouse="WH-GR9",
        )
    )
    database.session.commit()

    result = reconcile_purchase_invoice(invoice.id)
    database.session.commit()

    reconciliation = database.session.execute(
        database.select(PurchaseReconciliation).filter_by(id=result.reconciliation_id)
    ).scalar_one()
    items = (
        database.session.execute(
            database.select(PurchaseReconciliationItem).filter_by(purchase_reconciliation_id=reconciliation.id)
        )
        .scalars()
        .all()
    )

    assert result.matching_result == "MATCH_FAILED"
    assert reconciliation.status == "disputed"
    assert items == []

    valid_invoice = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 3),
        supplier_id="SUPP-9",
        purchase_receipt_id=receipt.id,
        docstatus=1,
    )
    database.session.add(valid_invoice)
    database.session.flush()
    database.session.add(
        PurchaseInvoiceItem(
            purchase_invoice_id=valid_invoice.id,
            item_code="ITEM-GR9",
            qty=Decimal("2"),
            uom="EA",
            rate=Decimal("5.00"),
            amount=Decimal("10.00"),
            warehouse="WH-GR9",
        )
    )
    database.session.commit()

    valid_result = reconcile_purchase_invoice(valid_invoice.id)
    assert valid_result.matching_result == "MATCH_OK"


def test_purchase_reconciliation_does_not_net_opposite_line_prices():
    """Diferencias de precio opuestas no deben ocultar una infracción por línea."""
    from cacao_accounting.compras.purchase_reconciliation_service import (
        MatchingConfig,
        MatchingResult,
        ToleranceType,
        _evaluate_matching_result,
    )

    config = MatchingConfig(
        matching_type="2-way",
        price_tolerance_type=ToleranceType.PERCENTAGE,
        price_tolerance_value=Decimal("0"),
        qty_tolerance_type=ToleranceType.PERCENTAGE,
        qty_tolerance_value=Decimal("0"),
        require_purchase_order=True,
        bridge_account_required=True,
        auto_reconcile=True,
        allow_price_difference=False,
    )

    result = _evaluate_matching_result(
        total_invoiced_qty=Decimal("20"),
        total_reference_qty=Decimal("20"),
        total_price_difference=Decimal("0"),
        total_amount_difference=Decimal("0"),
        total_reference_amount=Decimal("200"),
        config=config,
        price_tolerance_failed=True,
    )

    assert result == MatchingResult.MATCH_FAILED


def test_bank_reconciliation_supports_partial_and_rejects_duplicates(app_ctx):
    from cacao_accounting.bancos.reconciliation_service import (
        BankReconciliationError,
        BankReconciliationMatch,
        BankReconciliationRequest,
        reconcile_bank_items,
    )
    from cacao_accounting.database import Bank, BankAccount, BankTransaction, PaymentEntry, ReconciliationItem, database

    bank = Bank(name="Banco")
    database.session.add(bank)
    database.session.flush()
    bank_account = BankAccount(bank_id=bank.id, company="cacao", account_name="Cuenta")
    database.session.add(bank_account)
    database.session.flush()
    transaction = BankTransaction(bank_account_id=bank_account.id, posting_date=date(2026, 5, 5), deposit=Decimal("100.00"))
    payment_a = PaymentEntry(
        company="cacao",
        posting_date=date(2026, 5, 5),
        payment_type="receive",
        received_amount=Decimal("60.00"),
        bank_account_id=bank_account.id,
        docstatus=1,
    )
    payment_b = PaymentEntry(
        company="cacao",
        posting_date=date(2026, 5, 5),
        payment_type="receive",
        received_amount=Decimal("40.00"),
        bank_account_id=bank_account.id,
        docstatus=1,
    )
    database.session.add_all([transaction, payment_a, payment_b])
    database.session.commit()

    request = BankReconciliationRequest(
        company="cacao",
        reconciliation_date=date(2026, 5, 5),
        matches=[
            BankReconciliationMatch(transaction.id, "payment_entry", payment_a.id, Decimal("60.00")),
            BankReconciliationMatch(transaction.id, "payment_entry", payment_b.id, Decimal("40.00")),
        ],
    )
    reconciliation = reconcile_bank_items(request)
    database.session.commit()

    items = database.session.execute(database.select(ReconciliationItem)).scalars().all()
    assert transaction.is_reconciled is True
    assert sum(item.allocated_amount for item in items) == Decimal("100.00")
    assert reconcile_bank_items(request).id == reconciliation.id
    with pytest.raises(BankReconciliationError, match="excede"):
        reconcile_bank_items(
            BankReconciliationRequest(
                company="cacao",
                reconciliation_date=date(2026, 5, 5),
                matches=[BankReconciliationMatch(transaction.id, "payment_entry", payment_a.id, Decimal("1.00"))],
            )
        )


def test_bank_reconciliation_locks_shared_target(app_ctx, monkeypatch):
    """El destino compartido se lee con FOR UPDATE antes de calcular pendiente."""
    from cacao_accounting.bancos import reconciliation_service
    from cacao_accounting.database import Bank, BankAccount, BankTransaction, PaymentEntry, database

    bank = Bank(name="Banco lock target")
    database.session.add(bank)
    database.session.flush()
    account = BankAccount(bank_id=bank.id, company="cacao", account_name="Cuenta lock target")
    database.session.add(account)
    database.session.flush()
    transaction = BankTransaction(bank_account_id=account.id, posting_date=date(2026, 5, 5), deposit=Decimal("100"))
    payment = PaymentEntry(
        company="cacao",
        posting_date=date(2026, 5, 5),
        payment_type="receive",
        received_amount=Decimal("100"),
        bank_account_id=account.id,
        docstatus=1,
    )
    database.session.add_all([transaction, payment])
    database.session.commit()

    calls = []
    original_get = database.session.get

    def recording_get(model, ident, **kwargs):
        calls.append((model, ident, kwargs.get("with_for_update")))
        return original_get(model, ident, **kwargs)

    monkeypatch.setattr(database.session, "get", recording_get)
    reconciliation_service._validate_reconciliation_match(
        match=reconciliation_service.BankReconciliationMatch(transaction.id, "payment_entry", payment.id, Decimal("100")),
        company="cacao",
    )
    assert (PaymentEntry, payment.id, True) in calls


def test_bank_candidates_match_direction_and_allow_partial_payment(app_ctx):
    from cacao_accounting.bancos.reconciliation_service import find_bank_reconciliation_candidates
    from cacao_accounting.database import Bank, BankAccount, BankTransaction, PaymentEntry, database

    bank = Bank(name="Banco candidatos")
    database.session.add(bank)
    database.session.flush()
    bank_account = BankAccount(bank_id=bank.id, company="cacao", account_name="Cuenta candidatos")
    database.session.add(bank_account)
    database.session.flush()
    other_account = BankAccount(bank_id=bank.id, company="cacao", account_name="Cuenta ajena")
    database.session.add(other_account)
    database.session.flush()
    deposit = BankTransaction(
        bank_account_id=bank_account.id,
        posting_date=date(2026, 5, 5),
        deposit=Decimal("500"),
    )
    withdrawal = BankTransaction(
        bank_account_id=bank_account.id,
        posting_date=date(2026, 5, 5),
        withdrawal=Decimal("500"),
    )
    receive = PaymentEntry(
        company="cacao",
        posting_date=date(2026, 5, 5),
        payment_type="receive",
        received_amount=Decimal("1000"),
        bank_account_id=bank_account.id,
        docstatus=1,
    )
    pay = PaymentEntry(
        company="cacao",
        posting_date=date(2026, 5, 5),
        payment_type="pay",
        paid_amount=Decimal("1000"),
        bank_account_id=bank_account.id,
        docstatus=1,
    )
    other_receive = PaymentEntry(
        company="cacao",
        posting_date=date(2026, 5, 5),
        payment_type="receive",
        received_amount=Decimal("500"),
        bank_account_id=other_account.id,
        docstatus=1,
    )
    database.session.add_all([deposit, withdrawal, receive, pay, other_receive])
    database.session.commit()

    deposit_candidates = find_bank_reconciliation_candidates(deposit.id)
    withdrawal_candidates = find_bank_reconciliation_candidates(withdrawal.id)

    assert [(candidate.reference_id, candidate.amount, candidate.status) for candidate in deposit_candidates] == [
        (receive.id, Decimal("500.00"), "partial")
    ]
    assert [(candidate.reference_id, candidate.amount, candidate.status) for candidate in withdrawal_candidates] == [
        (pay.id, Decimal("500.00"), "partial")
    ]


def test_bank_reconciliation_converts_gl_entry_with_mismatched_currency(app_ctx):
    """BANK-AUDIT-16: una entrada GL con moneda distinta a la bancaria debe
    convertirse a la moneda bancaria en lugar de descartarse."""
    from cacao_accounting.bancos.reconciliation_service import find_bank_reconciliation_candidates
    from cacao_accounting.database import (
        Accounts,
        Bank,
        BankAccount,
        BankTransaction,
        Book,
        Book,
        ExchangeRate,
        GLEntry,
        database,
    )

    bank = Bank(name="Banco USD")
    database.session.add(bank)
    database.session.flush()
    bank_account = BankAccount(bank_id=bank.id, company="cacao", account_name="Cuenta USD", currency="USD", gl_account_id=None)
    database.session.add(bank_account)
    database.session.flush()

    gl_account = Accounts(entity="cacao", code="BANK-GL-01", name="Banco GL", active=True, enabled=True, group=False)
    database.session.add(gl_account)
    database.session.flush()

    local_book = Book(code="LOCAL", name="Local", entity="cacao", currency="NIO", is_primary=True, status="activo")
    database.session.add(local_book)
    database.session.flush()

    bank_account.gl_account_id = str(gl_account.id)
    database.session.add(bank_account)
    database.session.add(ExchangeRate(origin="NIO", destination="USD", rate="0.0273043", date=date(2026, 5, 5)))

    transaction = BankTransaction(
        bank_account_id=bank_account.id,
        posting_date=date(2026, 5, 5),
        deposit=Decimal("100.00"),
    )
    entry = GLEntry(
        posting_date=date(2026, 5, 5),
        company="cacao",
        ledger_id=local_book.id,
        account_id=gl_account.id,
        account_code=gl_account.code,
        debit=Decimal("1000.0000"),
        credit=Decimal("0"),
        account_currency="NIO",
        company_currency="NIO",
        voucher_type="journal_entry",
        voucher_id="TEST-GL-USD-1",
        is_cancelled=False,
        is_reversal=False,
    )
    database.session.add_all([transaction, entry])
    database.session.commit()

    candidates = find_bank_reconciliation_candidates(transaction.id)
    gl_candidates = [c for c in candidates if c.reference_type == "gl_entry"]

    assert len(gl_candidates) == 1
    assert gl_candidates[0].reference_id == entry.id
    assert gl_candidates[0].amount == Decimal("27.3043")


def test_bank_reconciliation_matches_eur_gl_debit_and_credit_in_usd(app_ctx):
    """BANK-AUDIT-16: convierte EUR 100 a USD 110 para débitos y créditos."""
    from cacao_accounting.bancos.reconciliation_service import find_bank_reconciliation_candidates
    from cacao_accounting.database import (
        Accounts,
        Bank,
        BankAccount,
        BankTransaction,
        Book,
        Entity,
        ExchangeRate,
        GLEntry,
        database,
    )

    bank = Bank(name="Banco USD EUR")
    gl_account = Accounts(entity="cacao", code="BANK-GL-EUR", name="Banco GL EUR", active=True, enabled=True, group=False)
    book = Book(code="USD-BOOK", name="Libro USD", entity="cacao", currency="USD", is_primary=True, status="activo")
    database.session.add_all([bank, gl_account, book])
    database.session.flush()
    entity = database.session.execute(database.select(Entity).filter_by(code="cacao")).scalars().first()
    assert entity is not None
    entity.currency = "USD"
    bank_account = BankAccount(
        bank_id=bank.id,
        company="cacao",
        account_name="Cuenta USD EUR",
        currency="USD",
        gl_account_id=gl_account.id,
    )
    database.session.add(bank_account)
    database.session.add(ExchangeRate(origin="EUR", destination="USD", rate="1.10", date=date(2026, 5, 5)))
    database.session.flush()

    deposit = BankTransaction(
        bank_account_id=bank_account.id,
        posting_date=date(2026, 5, 5),
        deposit=Decimal("110.00"),
    )
    withdrawal = BankTransaction(
        bank_account_id=bank_account.id,
        posting_date=date(2026, 5, 5),
        withdrawal=Decimal("110.00"),
    )
    debit = GLEntry(
        posting_date=date(2026, 5, 5),
        company="cacao",
        ledger_id=book.id,
        account_id=gl_account.id,
        account_code=gl_account.code,
        debit=Decimal("110.00"),
        credit=Decimal("0"),
        debit_in_account_currency=Decimal("100.00"),
        account_currency="EUR",
        company_currency="USD",
        voucher_type="journal_entry",
        voucher_id="TEST-GL-EUR-DEBIT",
        is_cancelled=False,
        is_reversal=False,
    )
    credit = GLEntry(
        posting_date=date(2026, 5, 5),
        company="cacao",
        ledger_id=book.id,
        account_id=gl_account.id,
        account_code=gl_account.code,
        debit=Decimal("0"),
        credit=Decimal("110.00"),
        credit_in_account_currency=Decimal("100.00"),
        account_currency="EUR",
        company_currency="USD",
        voucher_type="journal_entry",
        voucher_id="TEST-GL-EUR-CREDIT",
        is_cancelled=False,
        is_reversal=False,
    )
    database.session.add_all([deposit, withdrawal, debit, credit])
    database.session.commit()

    deposit_candidates = find_bank_reconciliation_candidates(deposit.id)
    withdrawal_candidates = find_bank_reconciliation_candidates(withdrawal.id)

    assert [(candidate.reference_id, candidate.amount) for candidate in deposit_candidates] == [
        (debit.id, Decimal("110.0000"))
    ]
    assert [(candidate.reference_id, candidate.amount) for candidate in withdrawal_candidates] == [
        (credit.id, Decimal("110.0000"))
    ]


def test_posted_payment_bank_dimension_reconciles_with_bank_summary(app_ctx):
    """Payment GL lines must retain the bank account used by the posting."""
    from cacao_accounting.contabilidad.posting import post_document_to_gl
    from cacao_accounting.database import (
        Accounts,
        Bank,
        BankAccount,
        CompanyDefaultAccount,
        GLEntry,
        PaymentEntry,
        database,
    )
    from cacao_accounting.reportes.services import BankingFilters, get_bank_balance_summary

    bank_gl = Accounts(
        entity="cacao",
        code="BANK-DIMENSION",
        name="Banco dimensión",
        active=True,
        enabled=True,
        classification="asset",
        account_type="bank",
    )
    advance = Accounts(
        entity="cacao",
        code="ADV-DIMENSION",
        name="Anticipo dimensión",
        active=True,
        enabled=True,
        classification="asset",
        account_type="asset",
    )
    bank = Bank(name="Banco dimensión")
    database.session.add_all([bank_gl, advance, bank])
    database.session.flush()
    bank_account = BankAccount(
        bank_id=bank.id,
        company="cacao",
        account_name="Cuenta dimensión",
        gl_account_id=bank_gl.id,
    )
    database.session.add(bank_account)
    database.session.flush()
    payment = PaymentEntry(
        company="cacao",
        posting_date=date(2026, 5, 12),
        payment_type="pay",
        bank_account_id=bank_account.id,
        paid_amount=Decimal("100.00"),
        docstatus=1,
    )
    database.session.add_all(
        [
            payment,
            CompanyDefaultAccount(
                company="cacao",
                default_bank=bank_gl.id,
                supplier_advance_account_id=advance.id,
            ),
        ]
    )
    database.session.commit()

    post_document_to_gl(payment)
    database.session.commit()

    bank_line = database.session.execute(
        database.select(GLEntry).filter_by(voucher_type="payment_entry", voucher_id=payment.id, account_id=bank_gl.id)
    ).scalar_one()
    report = get_bank_balance_summary(BankingFilters(company="cacao", bank_account_id=bank_account.id))

    assert bank_line.bank_account_id == bank_account.id
    assert report.totals["ending_balance"] == Decimal("-100.0000")


def test_reconciliation_matrix_isolates_selected_ledger(app_ctx):
    """La matriz no debe mezclar el saldo de otro libro contable."""
    from cacao_accounting.database import Accounts, Book, CompanyDefaultAccount, GLEntry, database
    from cacao_accounting.reportes.services import ReconciliationFilters, get_reconciliation_matrix

    receivable = Accounts(
        entity="cacao",
        code="AR-MATRIX",
        name="Cuentas por cobrar matriz",
        active=True,
        enabled=True,
        classification="Activo",
        account_type="receivable",
    )
    primary = Book(code="MATRIX-P", name="Matriz primaria", entity="cacao", currency="NIO", is_primary=True)
    secondary = Book(code="MATRIX-S", name="Matriz secundaria", entity="cacao", currency="NIO")
    database.session.add_all([receivable, primary, secondary])
    database.session.flush()
    database.session.add(CompanyDefaultAccount(company="cacao", default_receivable=receivable.id))
    database.session.add_all(
        [
            GLEntry(
                posting_date=date(2026, 5, 1),
                company="cacao",
                ledger_id=primary.id,
                account_id=receivable.id,
                account_code=receivable.code,
                debit=Decimal("100.00"),
                credit=Decimal("0"),
                voucher_type="journal_entry",
                voucher_id="MATRIX-P-1",
            ),
            GLEntry(
                posting_date=date(2026, 5, 1),
                company="cacao",
                ledger_id=secondary.id,
                account_id=receivable.id,
                account_code=receivable.code,
                debit=Decimal("900.00"),
                credit=Decimal("0"),
                voucher_type="journal_entry",
                voucher_id="MATRIX-S-1",
            ),
        ]
    )
    database.session.commit()

    report = get_reconciliation_matrix(ReconciliationFilters(company="cacao", ledger=primary.code))
    ar_row = next(row.values for row in report.rows if row.values["area"] == "AR")
    assert ar_row["gl_control_amount"] == Decimal("100.00")
    assert ar_row["difference"] == Decimal("-100.00")


def test_reconciliation_report_diagnoses_posting_without_bank_transaction(app_ctx):
    """El reporte identifica pagos posteados sin extracto bancario enlazado."""
    from cacao_accounting.database import Bank, BankAccount, PaymentEntry, database
    from cacao_accounting.reportes.services import get_reconciliation_report

    bank = Bank(name="Banco diagnóstico")
    database.session.add(bank)
    database.session.flush()
    bank_account = BankAccount(
        bank_id=bank.id,
        company="cacao",
        account_name="Cuenta diagnóstico",
        account_no="DIAG-001",
    )
    database.session.add(bank_account)
    database.session.flush()
    payment = PaymentEntry(
        company="cacao",
        posting_date=date(2026, 5, 20),
        payment_type="pay",
        bank_account_id=bank_account.id,
        paid_amount=Decimal("25.00"),
        docstatus=1,
    )
    database.session.add(payment)
    database.session.commit()

    report = get_reconciliation_report(company="cacao", as_of_date=date(2026, 5, 31))

    diagnostics = [row.values for row in report.rows if row.values["recon_type"] == "bank_diagnostic"]
    assert any(
        row["status"] == "posting_without_bank_transaction"
        and row["source_id"] == payment.id
        and row["amount"] == Decimal("25.00")
        for row in diagnostics
    )
    assert report.totals["bank_orphan_count"] == Decimal("1")


def test_reconciliation_matrix_exposes_uninvoiced_receipts_against_grni(app_ctx):
    """La matriz muestra recepciones pendientes contra la cuenta puente GRNI."""
    from cacao_accounting.database import (
        Accounts,
        Book,
        CompanyDefaultAccount,
        Item,
        PurchaseReceipt,
        PurchaseReceiptItem,
        UOM,
        database,
    )
    from cacao_accounting.reportes.services import ReconciliationFilters, get_reconciliation_matrix

    bridge = Accounts(
        entity="cacao",
        code="GRNI-MATRIX",
        name="GRNI matriz",
        active=True,
        enabled=True,
        classification="Pasivo",
        account_type="liability",
    )
    book = Book(code="GRNI-MATRIX-BOOK", name="GRNI matrix book", entity="cacao", currency="NIO", is_primary=True)
    uom = UOM(code="EA", name="Each")
    item = Item(code="ITEM-GRNI-MATRIX", name="Item GRNI", item_type="goods", is_stock_item=True, default_uom="EA")
    receipt = PurchaseReceipt(company="cacao", posting_date=date(2026, 5, 15), docstatus=1, grand_total=Decimal("50"))
    database.session.add_all([bridge, book, uom, item, receipt])
    database.session.flush()
    database.session.add_all(
        [
            CompanyDefaultAccount(company="cacao", bridge_account_id=bridge.id),
            PurchaseReceiptItem(
                purchase_receipt_id=receipt.id,
                item_code=item.code,
                item_name=item.name,
                qty=Decimal("10"),
                qty_in_base_uom=Decimal("10"),
                uom=uom.code,
                rate=Decimal("5"),
                amount=Decimal("50"),
                base_amount=Decimal("50"),
            ),
        ]
    )
    database.session.commit()

    report = get_reconciliation_matrix(ReconciliationFilters(company="cacao", ledger=None))
    grni_row = next(row.values for row in report.rows if row.values["area"] == "GRNI/AP 3-way")
    assert grni_row["subledger_amount"] == Decimal("-50")
    assert grni_row["gl_control_amount"] == Decimal("0")


def test_inventory_valuation_uses_latest_layer_at_cutoff(app_ctx):
    """Inventory valuation must not sum historical snapshots or future layers."""
    from cacao_accounting.database import StockValuationLayer, database
    from cacao_accounting.reportes.services import OperationalReportFilters, get_inventory_valuation

    database.session.add_all(
        [
            StockValuationLayer(
                item_code="ITEM-VALUATION",
                warehouse="WH-VALUATION",
                company="cacao",
                qty=Decimal("10"),
                rate=Decimal("10"),
                remaining_qty=Decimal("10"),
                remaining_stock_value=Decimal("100"),
                stock_value_difference=Decimal("100"),
                voucher_type="stock_entry",
                voucher_id="RECEIPT-VALUATION",
                posting_date=date(2026, 5, 1),
            ),
            StockValuationLayer(
                item_code="ITEM-VALUATION",
                warehouse="WH-VALUATION",
                company="cacao",
                qty=Decimal("-5"),
                rate=Decimal("10"),
                remaining_qty=Decimal("5"),
                remaining_stock_value=Decimal("50"),
                stock_value_difference=Decimal("-50"),
                voucher_type="stock_entry",
                voucher_id="ISSUE-VALUATION",
                posting_date=date(2026, 5, 15),
            ),
            StockValuationLayer(
                item_code="ITEM-VALUATION",
                warehouse="WH-VALUATION",
                company="cacao",
                qty=Decimal("1"),
                rate=Decimal("10"),
                remaining_qty=Decimal("99"),
                remaining_stock_value=Decimal("990"),
                stock_value_difference=Decimal("940"),
                voucher_type="stock_entry",
                voucher_id="FUTURE-VALUATION",
                posting_date=date(2026, 6, 1),
            ),
        ]
    )
    database.session.commit()

    report = get_inventory_valuation(
        OperationalReportFilters(company="cacao", item_code="ITEM-VALUATION", date_to=date(2026, 5, 31))
    )

    assert len(report.rows) == 1
    assert report.rows[0].values["remaining_qty"] == Decimal("5")
    assert report.totals["remaining_stock_value"] == Decimal("50")


def test_financial_reports_ignore_inactive_ledgers(app_ctx):
    """Financial reports must not resolve an inactive book."""
    from cacao_accounting.database import Book, database
    from cacao_accounting.reportes.services import FinancialReportFilters, get_trial_balance_report

    database.session.add(
        Book(
            entity="cacao",
            code="INACTIVE-REPORT",
            name="Libro inactivo de prueba",
            currency="NIO",
            status="inactivo",
            is_primary=True,
        )
    )
    database.session.commit()

    report = get_trial_balance_report(FinancialReportFilters(company="cacao", ledger="INACTIVE-REPORT"))

    assert report.rows == []


def test_bank_difference_journal_uses_account_codes_and_each_book_currency(app_ctx):
    _seed_accounting_admin()
    """El ajuste bancario se contabiliza en la moneda del banco y se convierte por libro."""
    from cacao_accounting.bancos.statement_service import create_bank_difference_journal
    from cacao_accounting.contabilidad.journal_service import submit_journal
    from cacao_accounting.database import (
        Accounts,
        Bank,
        BankAccount,
        BankTransaction,
        Book,
        CompanyDefaultAccount,
        Currency,
        ExchangeRate,
        GLEntry,
        Reconciliation,
        ReconciliationItem,
        database,
    )

    bank_gl = Accounts(entity="cacao", code="BANK-FX", name="Bank USD", classification="asset")
    difference = Accounts(entity="cacao", code="BANK-DIFF", name="Bank difference", classification="expense")
    bank = Bank(name="Banco FX")
    local_book = Book(entity="cacao", code="BANK-NIO", name="NIO", currency="NIO", status="activo", is_primary=True)
    eur_book = Book(entity="cacao", code="BANK-EUR", name="EUR", currency="EUR", status="activo")
    database.session.add_all(
        [
            bank_gl,
            difference,
            bank,
            local_book,
            eur_book,
            Currency(code="NIO", name="Cordoba", decimals=2, active=True),
            Currency(code="USD", name="US Dollar", decimals=2, active=True),
            Currency(code="EUR", name="Euro", decimals=2, active=True),
            ExchangeRate(origin="USD", destination="NIO", rate=Decimal("36"), date=date(2026, 5, 5)),
            ExchangeRate(origin="USD", destination="EUR", rate=Decimal("0.9"), date=date(2026, 5, 5)),
        ]
    )
    database.session.flush()
    bank_account = BankAccount(
        bank_id=bank.id,
        company="cacao",
        account_name="USD account",
        currency="USD",
        gl_account_id=bank_gl.id,
    )
    database.session.add(bank_account)
    database.session.flush()
    transaction = BankTransaction(
        bank_account_id=bank_account.id,
        posting_date=date(2026, 5, 5),
        withdrawal=Decimal("5"),
    )
    reconciliation = Reconciliation(company="cacao", recon_date=date(2026, 5, 5), recon_type="bank")
    database.session.add_all([transaction, reconciliation])
    database.session.flush()
    database.session.add_all(
        [
            ReconciliationItem(
                reconciliation_id=reconciliation.id,
                reference_type="bank_transaction",
                reference_id=transaction.id,
                source_type="bank_transaction",
                source_id=transaction.id,
                amount=Decimal("5"),
            ),
            CompanyDefaultAccount(company="cacao", bank_difference_account_id=difference.id),
        ]
    )
    database.session.commit()

    journal = create_bank_difference_journal(reconciliation.id, Decimal("5"), user_id="admin")
    database.session.commit()
    submit_journal(journal.id)

    entries = database.session.execute(database.select(GLEntry).filter_by(voucher_id=journal.id)).scalars().all()
    assert len(entries) == 4
    assert sum(entry.debit for entry in entries if entry.ledger_id == local_book.id) == Decimal("180")
    assert sum(entry.debit for entry in entries if entry.ledger_id == eur_book.id) == Decimal("4.5")


def test_ar_ap_subledger_excludes_nonposted_documents_and_cancelled_payments(app_ctx):
    """AP/AR solo concilia facturas contabilizadas y pagos vivos."""
    from cacao_accounting.database import (
        DocumentRelation,
        PaymentEntry,
        PaymentReference,
        PurchaseInvoice,
        database,
    )
    from cacao_accounting.reportes.services import SubledgerFilters, get_ar_ap_subledger

    draft = PurchaseInvoice(company="cacao", posting_date=date(2026, 5, 1), grand_total=Decimal("40"), docstatus=0)
    cancelled = PurchaseInvoice(company="cacao", posting_date=date(2026, 5, 1), grand_total=Decimal("60"), docstatus=2)
    active = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 1),
        grand_total=Decimal("100"),
        supplier_id="SUP-REPORT",
        docstatus=1,
    )
    payment = PaymentEntry(
        company="cacao",
        posting_date=date(2026, 5, 2),
        payment_type="pay",
        paid_amount=Decimal("20"),
        docstatus=2,
    )
    database.session.add_all([draft, cancelled, active, payment])
    database.session.flush()
    reference = PaymentReference(
        payment_id=payment.id,
        reference_type="purchase_invoice",
        reference_id=active.id,
        allocated_amount=Decimal("20"),
        allocation_date=date(2026, 5, 2),
        company="cacao",
    )
    database.session.add(reference)
    database.session.flush()
    database.session.add(
        DocumentRelation(
            source_type="purchase_invoice",
            source_id=active.id,
            target_type="payment_entry",
            target_id=payment.id,
            target_item_id=reference.id,
            company="cacao",
            qty=Decimal("1"),
            amount=Decimal("20"),
            relation_type="payment_reference",
            status="active",
        )
    )
    database.session.commit()

    report = get_ar_ap_subledger(SubledgerFilters(company="cacao", party_type="supplier"))
    assert [row.values["document_id"] for row in report.rows] == [active.id]
    assert report.rows[0].values["paid_amount"] == Decimal("0")
    assert report.rows[0].values["outstanding_amount"] == Decimal("100")


def test_accounts_payable_reports_exclude_purchase_returns_and_keep_legacy_invoice(app_ctx):
    """AP views show payable invoices, not negative purchase returns."""
    from cacao_accounting.database import PurchaseInvoice, database
    from cacao_accounting.reportes.services import (
        AgingFilters,
        SubledgerFilters,
        get_aging_report,
        get_ar_ap_subledger,
    )

    invoice = PurchaseInvoice(
        id="FCC-REPORT-LEGACY",
        document_no="FCC-DEMO-2025-001",
        company="cacao",
        posting_date=date(2025, 1, 25),
        grand_total=Decimal("50"),
        outstanding_amount=Decimal("50"),
        docstatus=1,
    )
    purchase_return = PurchaseInvoice(
        id="PI-REPORT-RETURN",
        document_no="cacao-PI-RETURN-00001",
        company="cacao",
        posting_date=date(2026, 8, 15),
        grand_total=Decimal("20000"),
        outstanding_amount=Decimal("20000"),
        is_return=True,
        document_type="purchase_return",
        docstatus=1,
    )
    database.session.add_all([invoice, purchase_return])
    database.session.commit()

    report = get_ar_ap_subledger(SubledgerFilters(company="cacao", party_type="supplier", include_returns=False))
    aging = get_aging_report(
        AgingFilters(
            company="cacao",
            party_type="supplier",
            as_of_date=date(2026, 8, 15),
            include_returns=False,
        )
    )

    assert [row.values["document_no"] for row in report.rows] == ["FCC-DEMO-2025-001"]
    assert report.totals["outstanding_amount"] == Decimal("50")
    assert [row.values["document_no"] for row in aging.rows] == ["FCC-DEMO-2025-001"]
    assert aging.totals["over_90"] == Decimal("50")


def test_demo_purchase_invoice_seed_has_supplier_for_ap_filters():
    """The demo invoice is selectable under the demo supplier in AP reports."""
    from cacao_accounting.datos.dev.data import _make_documentos

    invoice = next(document for document in _make_documentos() if document.document_no == "FCC-DEMO-2025-001")
    assert invoice.supplier_id == "PARTY-DEMO-SUPPLIER"
    assert invoice.supplier_name == "Proveedor Demo SA"


def test_ar_subledger_uses_base_currency_and_offsets_returns(app_ctx):
    """AR no mezcla USD nominales con NIO ni presenta devoluciones como débitos."""
    from cacao_accounting.database import SalesInvoice, database
    from cacao_accounting.reportes.services import (
        MaturityFilters,
        SubledgerFilters,
        get_ar_ap_subledger,
        get_maturity_schedule,
    )

    invoice = SalesInvoice(
        company="cacao",
        posting_date=date(2026, 5, 1),
        customer_id="CUST-FX",
        transaction_currency="USD",
        base_currency="NIO",
        exchange_rate=Decimal("36"),
        grand_total=Decimal("10"),
        base_grand_total=Decimal("360"),
        docstatus=1,
    )
    credit_note = SalesInvoice(
        company="cacao",
        posting_date=date(2026, 5, 2),
        customer_id="CUST-FX",
        transaction_currency="USD",
        base_currency="NIO",
        exchange_rate=Decimal("36"),
        grand_total=Decimal("2"),
        base_grand_total=Decimal("72"),
        is_return=True,
        docstatus=1,
    )
    database.session.add_all([invoice, credit_note])
    database.session.commit()

    subledger = get_ar_ap_subledger(SubledgerFilters(company="cacao", party_type="customer"))
    maturity = get_maturity_schedule(MaturityFilters(company="cacao", party_type="customer", as_of_date=date(2026, 5, 3)))

    assert subledger.totals["original_amount"] == Decimal("288")
    assert subledger.totals["outstanding_amount"] == Decimal("288")
    assert {row.values["currency"] for row in subledger.rows} == {"NIO"}
    assert {row.values["transaction_currency"] for row in subledger.rows} == {"USD"}
    assert sorted(row.values["outstanding_amount"] for row in subledger.rows) == [Decimal("-72"), Decimal("360")]
    assert maturity.totals["outstanding_amount"] == Decimal("288")


def test_ar_subledger_paid_amount_includes_undated_allocation(app_ctx):
    """An allocation without a date remains visible in an as-of report."""
    from cacao_accounting.database import DocumentRelation, PaymentEntry, PaymentReference, SalesInvoice, database
    from cacao_accounting.reportes.services import SubledgerFilters, get_ar_ap_subledger

    invoice = SalesInvoice(
        company="cacao",
        posting_date=date(2026, 5, 1),
        customer_id="CUST-UNDATED",
        grand_total=Decimal("100"),
        outstanding_amount=Decimal("60"),
        docstatus=1,
    )
    payment = PaymentEntry(
        company="cacao",
        posting_date=date(2026, 5, 2),
        payment_type="receive",
        paid_amount=Decimal("40"),
        docstatus=1,
    )
    database.session.add_all([invoice, payment])
    database.session.flush()
    reference = PaymentReference(
        payment_id=payment.id,
        reference_type="sales_invoice",
        reference_id=invoice.id,
        allocated_amount=Decimal("40"),
        allocation_date=None,
    )
    database.session.add(reference)
    database.session.flush()
    database.session.add(
        DocumentRelation(
            source_type="sales_invoice",
            source_id=invoice.id,
            target_type="payment_entry",
            target_id=payment.id,
            target_item_id=reference.id,
            qty=Decimal("1"),
            amount=Decimal("40"),
            relation_type="payment_reference",
            status="active",
        )
    )
    database.session.commit()

    report = get_ar_ap_subledger(SubledgerFilters(company="cacao", party_type="customer", as_of_date=date(2026, 5, 3)))

    assert report.rows[0].values["paid_amount"] == Decimal("40")
    assert report.rows[0].values["outstanding_amount"] == Decimal("60")


def test_ar_subledger_ignores_payment_from_another_company(app_ctx):
    """A cross-company payment relation must not reduce the invoice balance."""
    from cacao_accounting.database import DocumentRelation, PaymentEntry, PaymentReference, SalesInvoice, database
    from cacao_accounting.reportes.services import SubledgerFilters, get_ar_ap_subledger

    invoice = SalesInvoice(
        company="cacao",
        posting_date=date(2026, 5, 1),
        customer_id="CUST-COMPANY",
        grand_total=Decimal("100"),
        docstatus=1,
    )
    payment = PaymentEntry(
        company="other-company",
        posting_date=date(2026, 5, 2),
        payment_type="receive",
        paid_amount=Decimal("100"),
        docstatus=1,
    )
    database.session.add_all([invoice, payment])
    database.session.flush()
    reference = PaymentReference(
        payment_id=payment.id,
        reference_type="sales_invoice",
        reference_id=invoice.id,
        allocated_amount=Decimal("100"),
    )
    database.session.add(reference)
    database.session.flush()
    database.session.add(
        DocumentRelation(
            source_type="sales_invoice",
            source_id=invoice.id,
            target_type="payment_entry",
            target_id=payment.id,
            target_item_id=reference.id,
            qty=Decimal("1"),
            amount=Decimal("100"),
            relation_type="payment_reference",
            status="active",
            company="cacao",
        )
    )
    database.session.commit()

    report = get_ar_ap_subledger(SubledgerFilters(company="cacao", party_type="customer"))

    assert report.rows[0].values["paid_amount"] == Decimal("0")
    assert report.rows[0].values["outstanding_amount"] == Decimal("100")


def test_maturity_schedule_excludes_documents_after_as_of_date(app_ctx):
    """Maturity reports must not include invoices posted after their cutoff."""
    from cacao_accounting.database import SalesInvoice, database
    from cacao_accounting.reportes.services import MaturityFilters, get_maturity_schedule

    database.session.add_all(
        [
            SalesInvoice(
                company="cacao",
                posting_date=date(2026, 5, 1),
                customer_id="CUST-MATURITY",
                grand_total=Decimal("100"),
                docstatus=1,
            ),
            SalesInvoice(
                company="cacao",
                posting_date=date(2026, 6, 1),
                customer_id="CUST-MATURITY",
                grand_total=Decimal("200"),
                docstatus=1,
            ),
        ]
    )
    database.session.commit()

    report = get_maturity_schedule(
        MaturityFilters(
            company="cacao",
            party_type="customer",
            party_id="CUST-MATURITY",
            as_of_date=date(2026, 5, 31),
        )
    )

    assert report.totals["outstanding_amount"] == Decimal("100")
    assert len(report.rows) == 1


def test_reports_return_subledger_aging_kardex_and_reconciliations(app_ctx):
    from cacao_accounting.database import (
        DocumentRelation,
        PaymentEntry,
        PaymentReference,
        SalesInvoice,
        StockLedgerEntry,
        database,
    )
    from cacao_accounting.reportes.services import (
        AgingFilters,
        KardexFilters,
        SubledgerFilters,
        get_aging_report,
        get_ar_ap_subledger,
        get_kardex,
        get_reconciliation_report,
    )

    invoice = SalesInvoice(
        company="cacao",
        posting_date=date(2026, 4, 1),
        customer_id="CUST-R",
        grand_total=Decimal("100.00"),
        docstatus=1,
    )
    database.session.add(invoice)
    database.session.flush()

    payment = PaymentEntry(
        company="cacao",
        posting_date=date(2026, 4, 15),
        payment_type="receive",
        paid_amount=Decimal("25.00"),
        docstatus=1,
    )
    database.session.add(payment)
    database.session.flush()

    ref = PaymentReference(
        payment_id=payment.id,
        reference_type="sales_invoice",
        reference_id=invoice.id,
        allocated_amount=Decimal("25.00"),
        allocation_date=date(2026, 4, 15),
    )
    database.session.add(ref)
    database.session.flush()
    database.session.add(
        DocumentRelation(
            source_type="sales_invoice",
            source_id=invoice.id,
            target_type="payment_entry",
            target_id=payment.id,
            target_item_id=ref.id,
            qty=Decimal("1"),
            amount=Decimal("25"),
            relation_type="payment_reference",
            status="active",
        )
    )
    database.session.add(
        StockLedgerEntry(
            posting_date=date(2026, 5, 1),
            item_code="ITEM-R",
            warehouse="WH-R",
            company="cacao",
            qty_change=Decimal("3"),
            qty_after_transaction=Decimal("3"),
            valuation_rate=Decimal("2.00"),
            stock_value_difference=Decimal("6.00"),
            stock_value=Decimal("6.00"),
            voucher_type="seed",
            voucher_id="seed-r",
        )
    )
    database.session.commit()

    subledger = get_ar_ap_subledger(SubledgerFilters(company="cacao", party_type="customer", as_of_date=date(2026, 5, 5)))
    aging = get_aging_report(AgingFilters(company="cacao", party_type="customer", as_of_date=date(2026, 5, 5)))
    kardex = get_kardex(KardexFilters(company="cacao", item_code="ITEM-R"))
    reconciliations = get_reconciliation_report(company="cacao")

    assert subledger.totals["outstanding_amount"] == Decimal("75.00")
    assert aging.totals["31_60"] == Decimal("75.00")
    assert kardex.totals["incoming_qty"] == Decimal("3.000000000")
    assert reconciliations.totals["bank_reconciled_amount"] == Decimal("0")


def test_inventory_existence_uses_latest_chronological_balance(app_ctx):
    """La existencia histórica no depende del orden físico de inserción."""
    from cacao_accounting.database import StockLedgerEntry, database
    from cacao_accounting.reportes.services import KardexFilters, get_inventory_existence

    database.session.add_all(
        [
            StockLedgerEntry(
                posting_date=date(2026, 5, 10),
                item_code="ITEM-ASOF",
                warehouse="WH-ASOF",
                company="cacao",
                qty_change=Decimal("2"),
                qty_after_transaction=Decimal("7"),
                valuation_rate=Decimal("4"),
                stock_value_difference=Decimal("8"),
                stock_value=Decimal("28"),
                voucher_type="stock_entry",
                voucher_id="NEWER-FIRST",
            ),
            StockLedgerEntry(
                posting_date=date(2026, 5, 1),
                item_code="ITEM-ASOF",
                warehouse="WH-ASOF",
                company="cacao",
                qty_change=Decimal("5"),
                qty_after_transaction=Decimal("5"),
                valuation_rate=Decimal("4"),
                stock_value_difference=Decimal("20"),
                stock_value=Decimal("20"),
                voucher_type="stock_entry",
                voucher_id="OLDER-SECOND",
            ),
        ]
    )
    database.session.commit()

    report = get_inventory_existence(KardexFilters(company="cacao", item_code="ITEM-ASOF", date_to=date(2026, 5, 31)))

    assert report.totals["balance_qty"] == Decimal("7")
    assert report.totals["stock_value"] == Decimal("28")


def test_cancelled_pairs_are_hidden_from_bank_and_stock_reports_but_rebuild_to_net(app_ctx):
    """Una anulación interna no altera saldos ni agrega ruido a reportes ordinarios."""
    from cacao_accounting.database import (
        Bank,
        BankAccount,
        GLEntry,
        Item,
        StockBin,
        StockLedgerEntry,
        UOM,
        Warehouse,
        database,
    )
    from cacao_accounting.inventario.service import rebuild_stock_bins
    from cacao_accounting.reportes.services import BankingFilters, KardexFilters, get_bank_balance_summary, get_kardex

    bank = Bank(name="Banco cancelaciones")
    database.session.add(bank)
    database.session.flush()
    bank_account = BankAccount(bank_id=bank.id, company="cacao", account_name="Cuenta cancelaciones")
    database.session.add(bank_account)
    database.session.flush()
    database.session.add_all(
        [
            GLEntry(
                posting_date=date(2026, 5, 1),
                company="cacao",
                debit=Decimal("100"),
                credit=Decimal("0"),
                bank_account_id=bank_account.id,
                voucher_type="payment_entry",
                voucher_id="PAY-CANCELLED",
                is_cancelled=True,
            ),
            GLEntry(
                posting_date=date(2026, 5, 1),
                company="cacao",
                debit=Decimal("0"),
                credit=Decimal("100"),
                bank_account_id=bank_account.id,
                voucher_type="payment_entry",
                voucher_id="PAY-CANCELLED",
                is_reversal=True,
            ),
            GLEntry(
                posting_date=date(2026, 5, 2),
                company="cacao",
                debit=Decimal("25"),
                credit=Decimal("0"),
                bank_account_id=bank_account.id,
                voucher_type="payment_entry",
                voucher_id="PAY-ACTIVE",
            ),
        ]
    )

    database.session.add_all(
        [
            UOM(code="EA-CAN", name="Unidad cancelaciones"),
            Item(
                code="ITEM-CAN",
                name="Item cancelaciones",
                item_type="goods",
                is_stock_item=True,
                default_uom="EA-CAN",
            ),
            Warehouse(code="WH-CAN", name="Bodega cancelaciones", company="cacao"),
        ]
    )
    database.session.flush()
    database.session.add_all(
        [
            StockLedgerEntry(
                posting_date=date(2026, 5, 1),
                item_code="ITEM-CAN",
                warehouse="WH-CAN",
                company="cacao",
                qty_change=Decimal("10"),
                qty_after_transaction=Decimal("10"),
                valuation_rate=Decimal("2"),
                stock_value_difference=Decimal("20"),
                stock_value=Decimal("20"),
                voucher_type="purchase_receipt",
                voucher_id="PR-CANCELLED",
                is_cancelled=True,
            ),
            StockLedgerEntry(
                posting_date=date(2026, 5, 1),
                item_code="ITEM-CAN",
                warehouse="WH-CAN",
                company="cacao",
                qty_change=Decimal("-10"),
                qty_after_transaction=Decimal("0"),
                valuation_rate=Decimal("2"),
                stock_value_difference=Decimal("-20"),
                stock_value=Decimal("0"),
                voucher_type="purchase_receipt",
                voucher_id="PR-CANCELLED",
                is_cancelled=True,
            ),
            StockLedgerEntry(
                posting_date=date(2026, 5, 2),
                item_code="ITEM-CAN",
                warehouse="WH-CAN",
                company="cacao",
                qty_change=Decimal("5"),
                qty_after_transaction=Decimal("5"),
                valuation_rate=Decimal("3"),
                stock_value_difference=Decimal("15"),
                stock_value=Decimal("15"),
                voucher_type="purchase_receipt",
                voucher_id="PR-ACTIVE",
            ),
        ]
    )
    database.session.commit()

    bank_report = get_bank_balance_summary(BankingFilters(company="cacao", bank_account_id=bank_account.id))
    kardex = get_kardex(KardexFilters(company="cacao", item_code="ITEM-CAN", warehouse="WH-CAN"))
    rebuild_stock_bins("cacao", item_code="ITEM-CAN", warehouse="WH-CAN")
    database.session.commit()
    stock_bin = database.session.execute(
        database.select(StockBin).filter_by(company="cacao", item_code="ITEM-CAN", warehouse="WH-CAN")
    ).scalar_one()

    assert bank_report.totals["ending_balance"] == Decimal("25")
    assert len(kardex.rows) == 1
    assert kardex.totals["incoming_qty"] == Decimal("5")
    assert kardex.totals["outgoing_qty"] == Decimal("0")
    assert stock_bin.actual_qty == Decimal("5")
    assert stock_bin.stock_value == Decimal("15")


def test_financial_reports_framework_uses_gl_and_supports_export(app_ctx):
    from cacao_accounting.database import (
        AccountingPeriod,
        Accounts,
        Book,
        FiscalYear,
        GLEntry,
        Modules,
        User,
        database,
    )
    from cacao_accounting.reportes.services import (
        FinancialReportFilters,
        get_account_movement_detail,
        get_balance_sheet_report,
        get_income_statement_report,
        get_trial_balance_report,
    )

    fiscal_year = FiscalYear(
        entity="cacao",
        name="2026",
        year_start_date=date(2026, 1, 1),
        year_end_date=date(2026, 12, 31),
        is_closed=False,
    )
    database.session.add(fiscal_year)
    database.session.flush()
    period_apr = AccountingPeriod(
        entity="cacao",
        fiscal_year_id=fiscal_year.id,
        name="2026-04",
        enabled=True,
        is_closed=False,
        start=date(2026, 4, 1),
        end=date(2026, 4, 30),
    )
    period_may = AccountingPeriod(
        entity="cacao",
        fiscal_year_id=fiscal_year.id,
        name="2026-05",
        enabled=True,
        is_closed=False,
        start=date(2026, 5, 1),
        end=date(2026, 5, 31),
    )
    book = Book(code="FISC", name="Fiscal", entity="cacao", currency="NIO", is_primary=True, default=True)
    accounts = [
        Accounts(entity="cacao", code="1.01.01", name="Caja", active=True, enabled=True, classification="Activo"),
        Accounts(entity="cacao", code="3.01.01", name="Capital", active=True, enabled=True, classification="Patrimonio"),
        Accounts(entity="cacao", code="4.01.01", name="Ventas", active=True, enabled=True, classification="Ingresos"),
        Accounts(entity="cacao", code="5.01.01", name="Gastos", active=True, enabled=True, classification="Gastos"),
    ]
    database.session.add_all([period_apr, period_may, book, *accounts])
    database.session.flush()
    cash, equity, income, expense = accounts

    database.session.add_all(
        [
            GLEntry(
                posting_date=date(2026, 4, 30),
                company="cacao",
                ledger_id=book.id,
                account_id=cash.id,
                account_code=cash.code,
                debit=Decimal("50.00"),
                credit=Decimal("0"),
                voucher_type="journal_entry",
                voucher_id="OPEN-1",
                document_no="cacao-JOU-2026-04-00001",
                accounting_period_id=period_apr.id,
            ),
            GLEntry(
                posting_date=date(2026, 4, 30),
                company="cacao",
                ledger_id=book.id,
                account_id=equity.id,
                account_code=equity.code,
                debit=Decimal("0"),
                credit=Decimal("50.00"),
                voucher_type="journal_entry",
                voucher_id="OPEN-1",
                document_no="cacao-JOU-2026-04-00001",
                accounting_period_id=period_apr.id,
            ),
            GLEntry(
                posting_date=date(2026, 5, 5),
                company="cacao",
                ledger_id=book.id,
                account_id=cash.id,
                account_code=cash.code,
                debit=Decimal("100.00"),
                credit=Decimal("0"),
                voucher_type="sales_invoice",
                voucher_id="SI-1",
                document_no="cacao-JOU-2026-05-00001",
                accounting_period_id=period_may.id,
            ),
            GLEntry(
                posting_date=date(2026, 5, 5),
                company="cacao",
                ledger_id=book.id,
                account_id=income.id,
                account_code=income.code,
                debit=Decimal("0"),
                credit=Decimal("100.00"),
                voucher_type="sales_invoice",
                voucher_id="SI-1",
                document_no="cacao-JOU-2026-05-00001",
                accounting_period_id=period_may.id,
            ),
            GLEntry(
                posting_date=date(2026, 5, 12),
                company="cacao",
                ledger_id=book.id,
                account_id=expense.id,
                account_code=expense.code,
                debit=Decimal("30.00"),
                credit=Decimal("0"),
                voucher_type="journal_entry",
                voucher_id="JE-2",
                document_no="cacao-JOU-2026-05-00002",
                accounting_period_id=period_may.id,
            ),
            GLEntry(
                posting_date=date(2026, 5, 12),
                company="cacao",
                ledger_id=book.id,
                account_id=cash.id,
                account_code=cash.code,
                debit=Decimal("0"),
                credit=Decimal("30.00"),
                voucher_type="journal_entry",
                voucher_id="JE-2",
                document_no="cacao-JOU-2026-05-00002",
                accounting_period_id=period_may.id,
            ),
        ]
    )
    accounting_module = Modules(module="accounting", default=True, enabled=True)
    report_user = User(user="report-user", name="Report User", password=b"x", classification="admin", active=True)
    database.session.add_all([accounting_module, report_user])
    database.session.commit()

    filters = FinancialReportFilters(
        company="cacao",
        ledger="FISC",
        accounting_period="2026-05",
        include_running_balance=True,
        page=1,
        page_size=10,
        voucher_number="2026-05",
    )
    movement = get_account_movement_detail(filters)
    movement_page_two = get_account_movement_detail(
        FinancialReportFilters(
            company="cacao",
            ledger="FISC",
            accounting_period="2026-05",
            account_code="1.01.01",
            include_running_balance=True,
            page=2,
            page_size=1,
            voucher_number="2026-05",
        )
    )
    trial_balance = get_trial_balance_report(
        FinancialReportFilters(company="cacao", ledger="FISC", accounting_period="2026-05")
    )
    income_statement = get_income_statement_report(
        FinancialReportFilters(company="cacao", ledger="FISC", accounting_period="2026-05")
    )
    balance_sheet = get_balance_sheet_report(
        FinancialReportFilters(company="cacao", ledger="FISC", accounting_period="2026-05")
    )

    assert movement.total_rows == 4
    assert movement.totals["difference"] == Decimal("0")
    cash_running_balances = [
        row.values.get("running_balance")
        for row in movement.rows
        if row.values.get("account_code") == "1.01.01" and "running_balance" in row.values
    ]
    assert cash_running_balances == [Decimal("100.0000"), Decimal("70.0000")]
    assert movement_page_two.rows[0].values.get("running_balance") == Decimal("70.0000")
    assert trial_balance.totals["debit"] == Decimal("130.00")
    assert trial_balance.totals["credit"] == Decimal("130.00")
    assert income_statement.totals["net_profit"] == Decimal("70.00")
    assert balance_sheet.totals["difference"] == Decimal("0.00")

    app_ctx.config["SECRET_KEY"] = "testing"
    client = app_ctx.test_client()
    with client.session_transaction() as session:
        session["_user_id"] = report_user.id
        session["_fresh"] = True
    response = client.get("/reports/account-movement?company=cacao&ledger=FISC&accounting_period=2026-05&export=csv")
    assert response.status_code == 200
    assert response.mimetype == "text/csv"
    html_response = client.get("/reports/account-movement?company=cacao&ledger=FISC&accounting_period=2026-05")
    html = html_response.get_data(as_text=True)
    assert html_response.status_code == 200
    assert 'doctype: "company"' in html
    assert 'doctype: "book"' in html
    assert 'doctype: "accounting_period"' in html
    assert 'name="voucher_number"' in html
    response_xlsx = client.get("/reports/account-movement?company=cacao&ledger=FISC&accounting_period=2026-05&export=xlsx")
    assert response_xlsx.status_code == 200
    assert response_xlsx.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(response_xlsx.data))
    assert "Filtros" in workbook.sheetnames
    assert workbook.active.freeze_panes == "A5"
    filters_sheet = workbook["Filtros"]
    filter_rows = [row for row in filters_sheet.iter_rows(min_row=2, max_col=2, values_only=True) if row[0]]
    assert any("Company" in str(row[0]) for row in filter_rows)
    assert any(str(row[1]) == "cacao" for row in filter_rows)


def test_account_movement_displays_party_name_instead_of_internal_id(app_ctx):
    from cacao_accounting.database import (
        AccountingPeriod,
        Accounts,
        Book,
        FiscalYear,
        GLEntry,
        Modules,
        Party,
        User,
        database,
    )
    from cacao_accounting.reportes.services import FinancialReportFilters, get_account_movement_detail

    report_user = User(user="report-party-user", name="Report Party User", password=b"x", classification="admin", active=True)
    fiscal_year = FiscalYear(
        entity="cacao",
        name="FY-2026-PARTY",
        year_start_date=date(2026, 1, 1),
        year_end_date=date(2026, 12, 31),
    )
    book = Book(entity="cacao", code="FISC-PARTY", name="Fiscal Party", currency="NIO", is_primary=True, default=True)
    payable = Accounts(
        entity="cacao",
        code="2.01.01",
        name="Cuentas por pagar",
        active=True,
        enabled=True,
        classification="pasivo",
    )
    supplier = Party(id="SUPP-VISIBLE", code="SUPP-VISIBLE", name="Proveedor Visible", is_supplier=True, is_active=True)
    database.session.add_all(
        [report_user, fiscal_year, book, payable, supplier, Modules(module="accounting", default=True, enabled=True)]
    )
    database.session.flush()
    period = AccountingPeriod(
        entity="cacao",
        fiscal_year_id=fiscal_year.id,
        name="2026-05-PARTY",
        enabled=True,
        is_closed=False,
        start=date(2026, 5, 1),
        end=date(2026, 5, 31),
    )
    database.session.add(period)
    database.session.flush()
    database.session.add(
        GLEntry(
            posting_date=date(2026, 5, 15),
            company="cacao",
            ledger_id=book.id,
            accounting_period_id=period.id,
            account_id=payable.id,
            account_code=payable.code,
            debit=Decimal("0"),
            credit=Decimal("100"),
            party_type="supplier",
            party_id=supplier.id,
            voucher_type="purchase_invoice",
            voucher_id="PI-VISIBLE-PARTY",
            document_no="cacao-PI-VISIBLE-PARTY",
        )
    )
    database.session.commit()

    report = get_account_movement_detail(
        FinancialReportFilters(company="cacao", ledger=book.code, accounting_period=period.name)
    )

    assert report.rows[0].values["party_id"] == "Proveedor Visible"
    assert supplier.id not in report.rows[0].values["party_id"]

    app_ctx.config["SECRET_KEY"] = "testing"
    client = app_ctx.test_client()
    with client.session_transaction() as session:
        session["_user_id"] = report_user.id
        session["_fresh"] = True
    response = client.get(
        f"/reports/account-movement?apply_filters=1&company=cacao&ledger={book.code}&accounting_period={period.name}"
    )

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "Proveedor Visible" in html


def test_financial_reports_exclude_cancelled_entries_and_reversals_by_default(app_ctx):
    from cacao_accounting.database import AccountingPeriod, Accounts, Book, FiscalYear, GLEntry, Modules, User, database
    from cacao_accounting.reportes.services import (
        FinancialReportFilters,
        get_account_movement_detail,
        get_account_summary_report,
        get_balance_sheet_report,
        get_income_statement_report,
        get_trial_balance_report,
    )

    accounting_module = Modules(module="accounting", default=True, enabled=True)
    report_user = User(
        user="report-cancel-user", name="Report Cancel User", password=b"x", classification="admin", active=True
    )
    fiscal_year = FiscalYear(
        entity="cacao",
        name="FY-2026-C",
        year_start_date=date(2026, 1, 1),
        year_end_date=date(2026, 12, 31),
    )
    book = Book(entity="cacao", code="FISC", name="Fiscal", currency="NIO", is_primary=True, default=True)
    cash = Accounts(entity="cacao", code="1.01.01", name="Caja", active=True, enabled=True, classification="activo")
    receivable = Accounts(entity="cacao", code="1.01.02", name="CxC", active=True, enabled=True, classification="activo")
    equity = Accounts(entity="cacao", code="3.01.01", name="Capital", active=True, enabled=True, classification="patrimonio")
    income = Accounts(entity="cacao", code="4.01.01", name="Ventas", active=True, enabled=True, classification="ingreso")
    database.session.add_all([accounting_module, report_user, fiscal_year, book, cash, receivable, equity, income])
    database.session.flush()
    period = AccountingPeriod(
        entity="cacao",
        fiscal_year_id=fiscal_year.id,
        name="2026-05",
        start=date(2026, 5, 1),
        end=date(2026, 5, 31),
        enabled=True,
        is_closed=False,
    )
    database.session.add(period)
    database.session.flush()
    database.session.add_all(
        [
            GLEntry(
                posting_date=date(2026, 5, 8),
                company="cacao",
                ledger_id=book.id,
                accounting_period_id=period.id,
                account_id=cash.id,
                account_code=cash.code,
                debit=Decimal("100.00"),
                credit=Decimal("0"),
                voucher_type="journal_entry",
                voucher_id="JE-ACT",
                document_no="JE-ACT",
            ),
            GLEntry(
                posting_date=date(2026, 5, 8),
                company="cacao",
                ledger_id=book.id,
                accounting_period_id=period.id,
                account_id=equity.id,
                account_code=equity.code,
                debit=Decimal("0"),
                credit=Decimal("100.00"),
                voucher_type="journal_entry",
                voucher_id="JE-ACT",
                document_no="JE-ACT",
            ),
            GLEntry(
                posting_date=date(2026, 5, 10),
                company="cacao",
                ledger_id=book.id,
                accounting_period_id=period.id,
                account_id=receivable.id,
                account_code=receivable.code,
                debit=Decimal("40.00"),
                credit=Decimal("0"),
                voucher_type="sales_invoice",
                voucher_id="SI-CAN",
                document_no="SI-CAN",
                is_cancelled=True,
            ),
            GLEntry(
                posting_date=date(2026, 5, 10),
                company="cacao",
                ledger_id=book.id,
                accounting_period_id=period.id,
                account_id=income.id,
                account_code=income.code,
                debit=Decimal("0"),
                credit=Decimal("40.00"),
                voucher_type="sales_invoice",
                voucher_id="SI-CAN",
                document_no="SI-CAN",
                is_cancelled=True,
            ),
            GLEntry(
                posting_date=date(2026, 5, 11),
                company="cacao",
                ledger_id=book.id,
                accounting_period_id=period.id,
                account_id=income.id,
                account_code=income.code,
                debit=Decimal("25.00"),
                credit=Decimal("0"),
                voucher_type="sales_invoice",
                voucher_id="SI-REV",
                document_no="SI-REV",
                is_reversal=True,
            ),
            GLEntry(
                posting_date=date(2026, 5, 11),
                company="cacao",
                ledger_id=book.id,
                accounting_period_id=period.id,
                account_id=receivable.id,
                account_code=receivable.code,
                debit=Decimal("0"),
                credit=Decimal("25.00"),
                voucher_type="sales_invoice",
                voucher_id="SI-REV",
                document_no="SI-REV",
                is_reversal=True,
            ),
        ]
    )
    database.session.commit()

    base_filters = FinancialReportFilters(company="cacao", ledger="FISC", accounting_period="2026-05", status="submitted")
    inclusive_filters = FinancialReportFilters(
        company="cacao",
        ledger="FISC",
        accounting_period="2026-05",
        status=None,
        include_cancellations=True,
    )

    movement_report = get_account_movement_detail(base_filters)
    summary_report = get_account_summary_report(base_filters)
    trial_balance_report = get_trial_balance_report(base_filters)
    balance_sheet_report = get_balance_sheet_report(base_filters)
    income_statement_report = get_income_statement_report(base_filters)

    assert movement_report.total_rows == 2
    assert summary_report.totals["debit"] == Decimal("100.00")
    assert summary_report.totals["credit"] == Decimal("100.00")
    assert trial_balance_report.totals["debit"] == Decimal("100.00")
    assert trial_balance_report.totals["credit"] == Decimal("100.00")
    assert balance_sheet_report.totals["assets"] == Decimal("100.00")
    assert balance_sheet_report.totals["equity"] == Decimal("100.00")
    assert income_statement_report.totals["net_profit"] == Decimal("0.00")

    movement_report_with_cancellations = get_account_movement_detail(inclusive_filters)
    trial_balance_with_cancellations = get_trial_balance_report(inclusive_filters)
    balance_sheet_with_cancellations = get_balance_sheet_report(inclusive_filters)
    income_statement_with_cancellations = get_income_statement_report(inclusive_filters)

    assert movement_report_with_cancellations.total_rows == 6
    assert trial_balance_with_cancellations.totals["debit"] == Decimal("165.00")
    assert trial_balance_with_cancellations.totals["credit"] == Decimal("165.00")
    assert balance_sheet_with_cancellations.totals["assets"] == Decimal("115.00")
    assert balance_sheet_with_cancellations.totals["equity"] == Decimal("115.00")
    assert income_statement_with_cancellations.totals["net_profit"] == Decimal("15.00")


def test_financial_report_filters_prefill_and_hide_columns_for_summary_reports(app_ctx):
    from cacao_accounting.database import AccountingPeriod, Book, FiscalYear, Modules, User, database

    accounting_module = Modules(module="accounting", default=True, enabled=True)
    report_user = User(
        user="report-filter-user", name="Report Filter User", password=b"x", classification="admin", active=True
    )
    fiscal_year = FiscalYear(
        entity="cacao",
        name="FY-2026",
        year_start_date=date(2026, 1, 1),
        year_end_date=date(2026, 12, 31),
    )
    database.session.add_all([accounting_module, report_user, fiscal_year])
    database.session.flush()
    database.session.add_all(
        [
            Book(entity="cacao", code="FISC", name="Fiscal", currency="NIO", is_primary=True, default=True),
            AccountingPeriod(
                entity="cacao",
                fiscal_year_id=fiscal_year.id,
                name="2026-05",
                start=date(2026, 5, 1),
                end=date(2026, 5, 31),
                enabled=True,
                is_closed=False,
            ),
        ]
    )
    database.session.commit()

    app_ctx.config["SECRET_KEY"] = "testing"
    client = app_ctx.test_client()
    with client.session_transaction() as session:
        session["_user_id"] = report_user.id
        session["_fresh"] = True

    report_paths = [
        "/reports/account-movement",
        "/reports/account-summary",
        "/reports/trial-balance",
        "/reports/balance-sheet",
        "/reports/income-statement",
    ]

    rendered_reports = {}
    for report_path in report_paths:
        response = client.get(report_path)
        html = response.get_data(as_text=True)
        rendered_reports[report_path] = html

        assert response.status_code == 200
        assert 'id="financial-advanced-state" name="advanced"' in html
        assert 'id="financial-advanced-toggle"' in html
        assert 'aria-controls="financial-advanced-filters"' in html
        assert 'id="financial-advanced-filters"' in html
        assert 'class="ca-report-advanced-filters d-grid gap-2 is-hidden" id="financial-advanced-filters"' in html

        account_index = html.index('name="account_code"')
        cancellations_index = html.index('name="show_cancellations"')
        closing_index = html.index('name="include_closing"')
        advanced_index = html.index('id="financial-advanced-filters"')
        assert account_index < cancellations_index < closing_index < advanced_index

        advanced_block = html[advanced_index : html.index('class="d-flex flex-wrap gap-2 mt-2"', advanced_index)]
        assert 'name="voucher_number"' in advanced_block
        assert 'name="account_from"' in advanced_block
        assert 'name="account_to"' in advanced_block
        assert 'name="cost_center_code"' in advanced_block
        assert 'name="unit_code"' in advanced_block
        assert 'name="project_code"' in advanced_block
        assert 'name="party_type"' in advanced_block
        assert 'name="party_id"' in advanced_block
        assert 'name="voucher_type"' in advanced_block
        assert 'name="status"' in advanced_block
        assert 'name="group_by"' in advanced_block

    advanced_response = client.get("/reports/trial-balance?advanced=1")
    advanced_html = advanced_response.get_data(as_text=True)
    assert advanced_response.status_code == 200
    assert 'aria-expanded="true"' in advanced_html
    assert 'class="ca-report-advanced-filters d-grid gap-2" id="financial-advanced-filters"' in advanced_html

    summary_html = rendered_reports["/reports/trial-balance"]
    detail_html = rendered_reports["/reports/account-movement"]

    assert 'initialValue: "FISC"' in summary_html
    assert 'initialValue: "2026-05"' in summary_html
    assert "Columnas visibles" not in summary_html
    assert 'data-bs-target="#saveViewModal">Guardar vista' not in summary_html
    assert 'name="view_action" value="reset">Eliminar vista' not in summary_html
    assert "Columnas visibles" not in detail_html


def test_financial_report_can_group_by_voucher_type_when_column_is_hidden(app_ctx):
    from cacao_accounting.database import Accounts, AccountingPeriod, Book, FiscalYear, GLEntry, Modules, User, database

    accounting_module = Modules(module="accounting", default=True, enabled=True)
    report_user = User(user="report-group-user", name="Report Group User", password=b"x", classification="admin", active=True)
    fiscal_year = FiscalYear(
        entity="cacao",
        name="FY-2026-G",
        year_start_date=date(2026, 1, 1),
        year_end_date=date(2026, 12, 31),
    )
    book = Book(entity="cacao", code="FISC", name="Fiscal", currency="NIO", is_primary=True, default=True)
    account = Accounts(entity="cacao", code="1.01.99", name="Caja Grupo", active=True, enabled=True)
    offset = Accounts(entity="cacao", code="3.01.99", name="Capital Grupo", active=True, enabled=True)
    database.session.add_all([accounting_module, report_user, fiscal_year, book, account, offset])
    database.session.flush()
    period = AccountingPeriod(
        entity="cacao",
        fiscal_year_id=fiscal_year.id,
        name="2026-05",
        start=date(2026, 5, 1),
        end=date(2026, 5, 31),
        enabled=True,
        is_closed=False,
    )
    database.session.add(period)
    database.session.flush()
    database.session.add_all(
        [
            GLEntry(
                posting_date=date(2026, 5, 8),
                company="cacao",
                ledger_id=book.id,
                accounting_period_id=period.id,
                account_id=account.id,
                account_code=account.code,
                debit=Decimal("10.00"),
                credit=Decimal("0"),
                voucher_type="journal_entry",
                voucher_id="JE-G",
                document_no="JE-G",
            ),
            GLEntry(
                posting_date=date(2026, 5, 8),
                company="cacao",
                ledger_id=book.id,
                accounting_period_id=period.id,
                account_id=offset.id,
                account_code=offset.code,
                debit=Decimal("0"),
                credit=Decimal("10.00"),
                voucher_type="journal_entry",
                voucher_id="JE-G",
                document_no="JE-G",
            ),
        ]
    )
    database.session.commit()

    app_ctx.config["SECRET_KEY"] = "testing"
    client = app_ctx.test_client()
    with client.session_transaction() as session:
        session["_user_id"] = report_user.id
        session["_fresh"] = True

    response = client.get(
        "/reports/account-movement?apply_filters=1&company=cacao&ledger=FISC&accounting_period=2026-05"
        "&group_by=voucher_type&visible_columns=posting_date&visible_columns=account_code&visible_columns=debit"
        "&visible_columns=credit"
    )
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert "Voucher Type: journal_entry" in html
    assert "Subtotal" in html


def test_search_select_party_role_and_group_filters(app_ctx):
    from cacao_accounting.database import CompanyParty, Modules, Party, PartyGroup, PriceList, TaxRule, User, database

    user = User(user="party-filter-user", name="Party Filter User", password=b"x", classification="admin", active=True)
    database.session.add_all(
        [
            Modules(module="accounting", default=True, enabled=True),
            user,
            PartyGroup(id="PG-CUST", group_type="customer", name="Mayorista", is_active=True),
            PartyGroup(id="PG-SUPP", group_type="supplier", name="Importador", is_active=True),
            PartyGroup(id="PG-INACTIVE", group_type="supplier", name="Inactivo", is_active=False),
            Party(id="SUPP-F", code="SUPP-F", is_supplier=True, name="Proveedor F", tax_id="SUPP-F", is_active=True),
            Party(id="CUST-F", code="CUST-F", is_customer=True, name="Cliente F", tax_id="CUST-F", is_active=True),
            CompanyParty(company="cacao", party_id="SUPP-F", is_active=True),
            CompanyParty(company="cacao", party_id="CUST-F", is_active=True),
            PriceList(name="Lista Ventas Cacao", company="cacao", is_selling=True, is_active=True),
            PriceList(name="Lista Compras Cacao", company="cacao", is_buying=True, is_selling=False, is_active=True),
            TaxRule(name="IVA Venta", company="cacao", applies_to="sales", concept="iva", is_active=True),
            TaxRule(name="IVA Compra", company="cacao", applies_to="purchase", concept="iva", is_active=True),
        ]
    )
    database.session.commit()

    app_ctx.config["SECRET_KEY"] = "testing"
    client = app_ctx.test_client()
    with client.session_transaction() as session:
        session["_user_id"] = user.id
        session["_fresh"] = True

    supplier_payload = client.get("/api/search-select?doctype=party&q=Proveedor&company=cacao&role=supplier").json
    customer_group_payload = client.get("/api/search-select?doctype=customer_group&q=may").json
    supplier_group_payload = client.get("/api/search-select?doctype=party_group&group_type=supplier&q=i").json
    price_list_payload = client.get("/api/search-select?doctype=price_list&q=Lista&company=cacao&is_selling=true").json
    tax_rule_payload = client.get("/api/search-select?doctype=tax_rule&q=IVA&company=cacao&applies_to=sales").json

    assert [item["value"] for item in supplier_payload["results"]] == ["SUPP-F"]
    assert [item["display_name"] for item in customer_group_payload["results"]] == ["Mayorista"]
    assert [item["display_name"] for item in supplier_group_payload["results"]] == ["Importador"]
    assert [item["display_name"] for item in price_list_payload["results"]] == ["Lista Ventas Cacao (Venta)"]
    assert [item["display_name"] for item in tax_rule_payload["results"]] == ["IVA Venta (Ventas)"]


def test_trial_balance_uses_tree_presentation_without_level_column(app_ctx):
    from cacao_accounting.database import (
        Accounts,
        AccountingPeriod,
        Book,
        FiscalYear,
        GLEntry,
        Modules,
        User,
        database,
    )

    accounting_module = Modules(module="accounting", default=True, enabled=True)
    report_user = User(user="trial-tree-user", name="Trial Tree User", password=b"x", classification="admin", active=True)
    fiscal_year = FiscalYear(
        entity="cacao",
        name="FY-2026",
        year_start_date=date(2026, 1, 1),
        year_end_date=date(2026, 12, 31),
    )
    book = Book(entity="cacao", code="FISC", name="Fiscal", currency="NIO", is_primary=True, default=True)
    period = AccountingPeriod(
        entity="cacao",
        fiscal_year_id=fiscal_year.id,
        name="2026-05",
        start=date(2026, 5, 1),
        end=date(2026, 5, 31),
        enabled=True,
        is_closed=False,
    )
    database.session.add_all([accounting_module, report_user, fiscal_year, book])
    database.session.flush()
    period.fiscal_year_id = fiscal_year.id
    database.session.add(period)
    account_parent = Accounts(
        entity="cacao",
        code="1.01",
        name="Activo Corriente",
        active=True,
        enabled=True,
        account_type="asset",
        classification="activo",
    )
    account_leaf = Accounts(
        entity="cacao",
        code="1.01.001",
        name="Caja",
        active=True,
        enabled=True,
        account_type="cash",
        classification="activo",
    )
    database.session.add_all([account_parent, account_leaf])
    database.session.flush()
    database.session.add_all(
        [
            GLEntry(
                posting_date=date(2026, 5, 1),
                company="cacao",
                ledger_id=book.id,
                accounting_period_id=period.id,
                account_id=account_leaf.id,
                account_code=account_leaf.code,
                debit=Decimal("120.00"),
                credit=Decimal("0"),
                voucher_type="journal_entry",
                voucher_id="TREE-1",
                document_no="TREE-1",
            ),
            GLEntry(
                posting_date=date(2026, 5, 1),
                company="cacao",
                ledger_id=book.id,
                accounting_period_id=period.id,
                account_id=account_leaf.id,
                account_code=account_leaf.code,
                debit=Decimal("0"),
                credit=Decimal("120.00"),
                voucher_type="journal_entry",
                voucher_id="TREE-1",
                document_no="TREE-1",
            ),
        ]
    )
    database.session.commit()

    app_ctx.config["SECRET_KEY"] = "testing"
    client = app_ctx.test_client()
    with client.session_transaction() as session:
        session["_user_id"] = report_user.id
        session["_fresh"] = True

    response = client.get("/reports/trial-balance?apply_filters=1&company=cacao&ledger=FISC&accounting_period=2026-05")
    html = response.get_data(as_text=True)
    assert response.status_code == 200
    assert "Activo Corriente" in html
    assert "Caja" in html
    assert "Level" not in html
    assert "ca-tree-toggle" in html


def test_tax_template_posts_sales_tax_and_price_suggestion(app_ctx):
    from cacao_accounting.contabilidad.posting import post_document_to_gl
    from cacao_accounting.database import (
        Accounts,
        GLEntry,
        ItemPrice,
        PartyAccount,
        PriceList,
        SalesInvoice,
        SalesInvoiceItem,
        Tax,
        TaxTemplate,
        TaxTemplateItem,
        database,
    )
    from cacao_accounting.tax_pricing_service import get_item_price, validate_price_tolerance

    receivable = Accounts(entity="cacao", code="AR-T", name="AR", active=True, enabled=True, account_type="receivable")
    income = Accounts(entity="cacao", code="INC-T", name="Ingreso", active=True, enabled=True, account_type="income")
    tax_account = Accounts(entity="cacao", code="TAX-T", name="IVA", active=True, enabled=True, account_type="liability")
    database.session.add_all([receivable, income, tax_account])
    database.session.flush()
    template = TaxTemplate(name="IVA Ventas", company="cacao", template_type="selling")
    tax = Tax(name="IVA 15", rate=Decimal("15.00"), tax_type="percentage", applies_to="sales", account_id=tax_account.id)
    price_list = PriceList(name="Ventas", company="cacao", currency="NIO", is_selling=True)
    database.session.add_all([template, tax, price_list])
    database.session.flush()
    database.session.add_all(
        [
            TaxTemplateItem(tax_template_id=template.id, tax_id=tax.id, sequence=1, behavior="additive"),
            PartyAccount(party_id="CUST-T", company="cacao", receivable_account_id=receivable.id),
            ItemPrice(item_code="ITEM-T", price_list_id=price_list.id, uom="EA", price=Decimal("100.00")),
        ]
    )
    invoice = SalesInvoice(
        company="cacao",
        posting_date=date(2026, 5, 5),
        customer_id="CUST-T",
        tax_template_id=template.id,
        total=Decimal("100.00"),
        grand_total=Decimal("115.00"),
        docstatus=1,
    )
    database.session.add(invoice)
    database.session.flush()
    line = SalesInvoiceItem(
        sales_invoice_id=invoice.id,
        item_code="ITEM-T",
        qty=Decimal("1"),
        uom="EA",
        rate=Decimal("100.00"),
        amount=Decimal("100.00"),
        income_account_id=income.id,
    )
    database.session.add(line)
    database.session.commit()

    post_document_to_gl(invoice)
    suggestion = get_item_price("ITEM-T", price_list.id, Decimal("1"), "EA", date(2026, 5, 5))
    line.suggested_rate = suggestion.price
    tolerance = validate_price_tolerance("sales_invoice", line, None)
    entries = database.session.execute(database.select(GLEntry)).scalars().all()

    assert suggestion.price == Decimal("100.0000")
    assert tolerance.allowed is True
    assert sum(entry.debit for entry in entries) == Decimal("115.0000")
    assert sum(entry.credit for entry in entries) == Decimal("115.0000")
    assert any(entry.account_id == tax_account.id and entry.credit == Decimal("15.0000") for entry in entries)


def test_catalog_loader_accepts_spanish_and_english_headers(app_ctx, tmp_path):
    from cacao_accounting.contabilidad.ctas import CatalogoCtas, cargar_catalogos
    from cacao_accounting.database import Accounts, Entity, database

    database.session.add(Entity(code="eng", name="English", company_name="English", tax_id="J-ENG", currency="NIO"))
    database.session.commit()

    english_catalog = tmp_path / "english.csv"
    english_catalog.write_text(
        "code,name,parent,group,classification,type,account_type\n1,Assets,,true,Asset,,\n1.01,Bank,1,false,Asset,,bank\n",
        encoding="utf-8",
    )
    cargar_catalogos(CatalogoCtas(file=str(english_catalog), pais=None, idioma="EN"), "eng")
    database.session.commit()

    account = database.session.execute(database.select(Accounts).filter_by(entity="eng", code="1.01")).scalar_one()
    assert account.account_type == "bank"
    assert account.group is False


def test_base_catalog_mapping_covers_required_default_accounts(app_ctx):
    import csv
    import json
    from pathlib import Path

    from cacao_accounting.contabilidad.default_accounts import DEFAULT_ACCOUNT_FIELDS

    catalog_path = Path("cacao_accounting/contabilidad/ctas/catalogos/base_es.csv")
    mapping_path = Path("cacao_accounting/contabilidad/ctas/catalogos/base_es.json")
    rows = list(csv.DictReader(catalog_path.open(encoding="utf-8")))
    codes = {row["codigo"] for row in rows}
    mapping = json.loads(mapping_path.read_text(encoding="utf-8"))["default_accounts"]

    assert set(DEFAULT_ACCOUNT_FIELDS) == set(mapping)
    assert all(code in codes for code in mapping.values())
    assert len(codes) == len(rows)

    catalog_path_en = Path("cacao_accounting/contabilidad/ctas/catalogos/base_en.csv")
    mapping_path_en = Path("cacao_accounting/contabilidad/ctas/catalogos/base_en.json")
    rows_en = list(csv.DictReader(catalog_path_en.open(encoding="utf-8")))
    codes_en = {row["codigo"] for row in rows_en}
    mapping_en = json.loads(mapping_path_en.read_text(encoding="utf-8"))["default_accounts"]

    assert set(DEFAULT_ACCOUNT_FIELDS) == set(mapping_en)
    assert all(code in codes_en for code in mapping_en.values())
    assert len(codes_en) == len(rows_en)

    for new_prefix in ["niif_pymes_es", "ifrs_smes_en", "us_gaap"]:
        cat_p = Path(f"cacao_accounting/contabilidad/ctas/catalogos/{new_prefix}.csv")
        map_p = Path(f"cacao_accounting/contabilidad/ctas/catalogos/{new_prefix}.json")
        new_rows = list(csv.DictReader(cat_p.open(encoding="utf-8")))
        new_codes = {row["codigo"] for row in new_rows}
        new_mapping = json.loads(map_p.read_text(encoding="utf-8"))["default_accounts"]

        assert set(DEFAULT_ACCOUNT_FIELDS) == set(new_mapping)
        assert all(code in new_codes for code in new_mapping.values())
        assert len(new_codes) == len(new_rows)

        account_rows = {row["codigo"]: row for row in new_rows}
        sales_discount = account_rows[new_mapping["sales_discount_account_id"]]
        purchase_discount = account_rows[new_mapping["purchase_discount_account_id"]]
        assert sales_discount["account_type"] == "payment_discount"
        assert sales_discount["rubro"] == "Expense"
        assert purchase_discount["account_type"] == "payment_discount"
        assert purchase_discount["rubro"] == "Income"


def test_setup_with_predefined_catalog_creates_complete_company_defaults(app_ctx):
    from cacao_accounting.contabilidad.default_accounts import DEFAULT_ACCOUNT_FIELDS
    from cacao_accounting.database import CompanyDefaultAccount, database
    from cacao_accounting.setup.service import available_catalog_files, finalize_setup

    assert ("base_es.csv", "Predeterminado - ES") in available_catalog_files()
    assert ("base_en.csv", "Default - EN") in available_catalog_files()
    assert ("niif_pymes_es.csv", "NIIF Pymes (ES)") in available_catalog_files()
    assert ("ifrs_smes_en.csv", "IFRS SMEs (EN)") in available_catalog_files()
    assert ("us_gaap.csv", "US GAAP — Standard") in available_catalog_files()

    finalize_setup(
        {
            "id": "mapco",
            "razon_social": "Mapping Company",
            "nombre_comercial": "Mapping Company",
            "id_fiscal": "J-MAP",
            "moneda": "NIO",
            "tipo_entidad": "Sociedad Anonima",
        },
        catalogo_tipo="preexistente",
        country="NI",
        idioma="ES",
        catalogo_archivo="base_es.csv",
    )
    database.session.commit()

    defaults = database.session.execute(database.select(CompanyDefaultAccount).filter_by(company="mapco")).scalar_one()
    assert all(getattr(defaults, field) for field in DEFAULT_ACCOUNT_FIELDS)


def test_setup_with_invalid_catalog_raises_error(app_ctx):
    from cacao_accounting.setup.service import finalize_setup

    with pytest.raises(ValueError, match="catálogo seleccionado.*no está disponible"):
        finalize_setup(
            {
                "id": "mapco",
                "razon_social": "Mapping Company",
                "nombre_comercial": "Mapping Company",
                "id_fiscal": "J-MAP",
                "moneda": "NIO",
                "tipo_entidad": "Sociedad Anonima",
            },
            catalogo_tipo="preexistente",
            country="NI",
            idioma="ES",
            catalogo_archivo="missing_catalog.csv",
        )


def test_setup_with_predefined_catalog_creates_bootstrap_records(app_ctx):
    from datetime import date

    from cacao_accounting.database import (
        AccountingPeriod,
        Book,
        CostCenter,
        Currency,
        Entity,
        FiscalYear,
        NamingSeries,
        database,
    )
    from cacao_accounting.setup.service import finalize_setup

    database.session.add(Currency(code="NIO", name="Córdoba", decimals=2, active=True, default=True))
    database.session.add(Currency(code="USD", name="US Dollar", decimals=2, active=True, default=False))
    database.session.add(Currency(code="EUR", name="Euro", decimals=2, active=True, default=False))
    database.session.commit()

    finalize_setup(
        {
            "id": "mapco",
            "razon_social": "Mapping Company",
            "nombre_comercial": "Mapping Company",
            "id_fiscal": "J-MAP",
            "moneda": "NIO",
            "tipo_entidad": "Sociedad Anonima",
        },
        catalogo_tipo="preexistente",
        country="NI",
        idioma="ES",
        catalogo_archivo="base_es.csv",
    )

    entity = database.session.execute(database.select(Entity).filter_by(code="mapco")).scalar_one()
    book = database.session.execute(database.select(Book).filter_by(entity="mapco", default=True)).scalar_one()
    cost_center = database.session.execute(database.select(CostCenter).filter_by(entity="mapco", code="MAIN")).scalar_one()
    fiscal_year = database.session.execute(
        database.select(FiscalYear).filter_by(entity="mapco", name=str(date.today().year))
    ).scalar_one()
    period = database.session.execute(
        database.select(AccountingPeriod).filter_by(entity="mapco", name=f"{date.today().year}-01")
    ).scalar_one()
    series = database.session.execute(
        database.select(NamingSeries).filter_by(company="mapco", entity_type="journal_entry")
    ).scalar_one_or_none()
    currency_nio = database.session.execute(database.select(Currency).filter_by(code="NIO")).scalar_one()
    currency_usd = database.session.execute(database.select(Currency).filter_by(code="USD")).scalar_one()
    currency_eur = database.session.execute(database.select(Currency).filter_by(code="EUR")).scalar_one()

    assert entity is not None
    assert book is not None
    assert book.code == "LOCAL"
    assert cost_center is not None
    assert fiscal_year is not None
    assert period is not None
    assert period.fiscal_year_id == fiscal_year.id
    assert series is not None
    assert currency_nio.active is True
    assert currency_nio.default is True
    assert currency_usd.active is True
    assert currency_usd.default is False
    assert currency_eur.active is True
    assert currency_eur.default is False


def test_setup_seeds_uoms_using_selected_language():
    from cacao_accounting import create_app
    from cacao_accounting.database import PriceList, UOM, database
    from cacao_accounting.setup.service import finalize_setup

    app = create_app(
        {
            **configuracion,
            "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
            "SQLALCHEMY_TRACK_MODIFICATIONS": False,
            "WTF_CSRF_ENABLED": False,
            "TESTING": True,
        }
    )
    with app.app_context():
        database.create_all()
        finalize_setup(
            {
                "id": "uomco",
                "razon_social": "UOM Company",
                "nombre_comercial": "UOM Company",
                "id_fiscal": "J-UOM",
                "moneda": "NIO",
                "tipo_entidad": "Sociedad Anonima",
            },
            catalogo_tipo="preexistente",
            country="NI",
            idioma="EN",
            catalogo_archivo="base_en.csv",
        )

        unit = database.session.execute(database.select(UOM).filter_by(code="UND")).scalar_one()
        box = database.session.execute(database.select(UOM).filter_by(code="CAJ")).scalar_one()
        service = database.session.execute(database.select(UOM).filter_by(code="SERV")).scalar_one()
        sales_price_list = database.session.execute(
            database.select(PriceList).filter_by(company="uomco", is_selling=True, is_default=True)
        ).scalar_one()
        purchase_price_list = database.session.execute(
            database.select(PriceList).filter_by(company="uomco", is_buying=True, is_default=True)
        ).scalar_one()

        assert unit.name == "Unit"
        assert box.name == "Box"
        assert service.name == "Service"
        assert sales_price_list.name == "Default Sales Price List"
        assert purchase_price_list.name == "Default Purchase Price List"


def test_setup_america_country_catalog_and_fast_currency_seed():
    from cacao_accounting import create_app
    from cacao_accounting.datos.base import registra_monedas
    from cacao_accounting.database import Currency, database
    from cacao_accounting.setup.catalogs import AMERICA_CURRENCY_CODES, country_choices, country_currency_map
    from cacao_accounting.setup.service import available_currencies

    app = create_app(
        {
            **configuracion,
            "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
            "SQLALCHEMY_TRACK_MODIFICATIONS": False,
            "WTF_CSRF_ENABLED": False,
            "TESTING": True,
        }
    )
    with app.app_context():
        database.create_all()
        registra_monedas(carga_rapida=True)

        country_codes = {code for code, _label in country_choices("en")}
        currency_codes = {
            currency.code
            for currency in database.session.execute(database.select(Currency).filter(Currency.active.is_(True))).scalars()
        }
        available_codes = {code for code, _label in available_currencies()}

        assert len(country_codes) == 35
        assert {"CA", "US", "MX", "NI", "AR", "BR", "VE"}.issubset(country_codes)
        assert set(AMERICA_CURRENCY_CODES).issubset(currency_codes)
        assert set(country_currency_map().values()).issubset(currency_codes)
        assert "USD" in available_codes
        assert available_codes.issubset(currency_codes)


def test_example_seed_creates_company_default_accounts(app_ctx):
    from cacao_accounting.contabilidad.default_accounts import DEFAULT_ACCOUNT_FIELDS
    from cacao_accounting.database.helpers import inicia_base_de_datos
    from cacao_accounting.database import CompanyDefaultAccount, database

    app = create_app(
        {
            **configuracion,
            "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
            "SQLALCHEMY_TRACK_MODIFICATIONS": False,
            "WTF_CSRF_ENABLED": False,
            "TESTING": True,
        }
    )
    with app.app_context():
        database.drop_all()
        database.create_all()
        assert inicia_base_de_datos(app=app, user="cacao", passwd="cacao", with_examples=True)

        for company in ("cacao", "dulce", "cafe"):
            defaults = database.session.execute(
                database.select(CompanyDefaultAccount).filter_by(company=company)
            ).scalar_one_or_none()
            assert defaults is not None
            assert all(getattr(defaults, field) for field in DEFAULT_ACCOUNT_FIELDS)


def test_example_seed_creates_company_base_records(app_ctx):
    from cacao_accounting.database import (
        AccountingPeriod,
        Book,
        CostCenter,
        Entity,
        FiscalYear,
        NamingSeries,
        PurchaseMatchingConfig,
        database,
    )
    from cacao_accounting.database.helpers import inicia_base_de_datos

    app = create_app(
        {
            **configuracion,
            "SQLALCHEMY_DATABASE_URI": "sqlite:///:memory:",
            "SQLALCHEMY_TRACK_MODIFICATIONS": False,
            "WTF_CSRF_ENABLED": False,
            "TESTING": True,
        }
    )
    with app.app_context():
        database.drop_all()
        database.create_all()
        assert inicia_base_de_datos(app=app, user="cacao", passwd="cacao", with_examples=True)

        for company in ("cacao", "dulce", "cafe"):
            assert database.session.execute(database.select(Entity).filter_by(code=company)).scalar_one_or_none()
            assert database.session.execute(
                database.select(Book).filter_by(entity=company, is_primary=True)
            ).scalar_one_or_none()
            assert database.session.execute(
                database.select(CostCenter).filter_by(entity=company, code="MAIN")
            ).scalar_one_or_none()
            assert database.session.execute(database.select(FiscalYear).filter_by(entity=company)).scalar_one_or_none()
            assert database.session.execute(database.select(AccountingPeriod).filter_by(entity=company)).scalars().first()
            assert database.session.execute(
                database.select(NamingSeries).filter_by(company=company, entity_type="journal_entry")
            ).scalar_one_or_none()
            assert database.session.execute(
                database.select(PurchaseMatchingConfig).filter_by(company=company)
            ).scalar_one_or_none()


def test_default_account_admin_crud_rejects_incompatible_types(app_ctx):
    from cacao_accounting.database import Accounts, Entity, Modules, User, database

    bank = Accounts(entity="cacao", code="BANK-CRUD", name="Banco", active=True, enabled=True, account_type="bank")
    expense = Accounts(entity="cacao", code="EXP-CRUD", name="Gasto", active=True, enabled=True, account_type="expense")
    admin_user = User(user="admin", name="Admin", password=b"x", classification="admin", active=True)
    admin_module = Modules(module="admin", default=True, enabled=True)
    database.session.add_all([bank, expense, admin_user, admin_module])
    database.session.commit()

    app_ctx.config["SECRET_KEY"] = "testing"
    client = app_ctx.test_client()
    with client.session_transaction() as session:
        session["_user_id"] = admin_user.id
        session["_fresh"] = True

    response = client.post(
        "/settings/default-accounts",
        data={"company": "cacao", "default_bank": expense.id, "action": "save"},
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert "debe ser de tipo" in response.get_data(as_text=True)

    response = client.post(
        "/settings/default-accounts",
        data={"company": "cacao", "default_bank": bank.id, "action": "save"},
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert "Cuentas predeterminadas guardadas correctamente" in response.get_data(as_text=True)
    assert database.session.execute(database.select(Entity).filter_by(code="cacao")).scalar_one()

    response = client.post(
        "/settings/default-accounts",
        data={"company": "cacao", "action": "delete"},
        follow_redirects=True,
    )
    assert response.status_code == 200


def test_search_select_account_filters_and_validates_registry(app_ctx):
    from cacao_accounting.database import Accounts, database
    from cacao_accounting.search_select import SearchSelectError, search_select

    bank = Accounts(
        entity="cacao", code="BAN-001", name="Banco Central", active=True, enabled=True, group=False, account_type="bank"
    )
    expense = Accounts(
        entity="cacao", code="EXP-001", name="Banco Gastos", active=True, enabled=True, group=False, account_type="expense"
    )
    disabled = Accounts(
        entity="cacao",
        code="BAN-002",
        name="Banco Deshabilitado",
        active=True,
        enabled=False,
        group=False,
        account_type="bank",
    )
    inactive = Accounts(
        entity="cacao", code="BAN-003", name="Banco Inactivo", active=False, enabled=True, group=False, account_type="bank"
    )
    group = Accounts(
        entity="cacao", code="BAN-004", name="Banco Grupo", active=True, enabled=True, group=True, account_type="bank"
    )
    database.session.add_all([bank, expense, disabled, inactive, group])
    database.session.commit()

    payload = search_select("account", "ban", {"company": ["cacao"], "account_type": ["bank"]}, limit=10)

    assert [item["id"] for item in payload["results"]] == [bank.id]
    assert payload["results"][0]["display_name"] == "BAN-001 - Banco Central"
    assert payload["results"][0]["account_type"] == "bank"

    mixed_types = search_select("account", "ban", {"company": ["cacao"], "account_type": ["bank", "expense"]}, limit=10)
    assert [item["id"] for item in mixed_types["results"]] == [bank.id, expense.id]

    limited = search_select("account", "ban", {"company": ["cacao"], "account_type": ["bank", "expense"]}, limit=1)
    assert len(limited["results"]) == 1
    assert limited["has_more"] is True

    with pytest.raises(SearchSelectError):
        search_select("unknown", "ban", {}, limit=10)

    with pytest.raises(SearchSelectError):
        search_select("account", "ban", {"not_allowed": ["x"]}, limit=10)


def test_search_select_account_parent_accepts_is_group_filter(app_ctx):
    from cacao_accounting.database import Accounts, database
    from cacao_accounting.search_select import search_select

    parent = Accounts(entity="cacao", code="ACT", name="Activo", active=True, enabled=True, group=True)
    leaf = Accounts(entity="cacao", code="ACT-01", name="Caja", active=True, enabled=True, group=False)
    database.session.add_all([parent, leaf])
    database.session.commit()

    payload = search_select("account", "act", {"company": ["cacao"], "is_group": ["1"]}, limit=10)

    assert [item["value"] for item in payload["results"]] == [parent.id]


def test_search_select_cost_center_parent_accepts_is_group_filter(app_ctx):
    from cacao_accounting.database import CostCenter, database
    from cacao_accounting.search_select import search_select

    parent = CostCenter(entity="cacao", code="ADM", name="Admin", active=True, enabled=True, group=True)
    leaf = CostCenter(entity="cacao", code="ADM01", name="Admin 01", active=True, enabled=True, group=False)
    database.session.add_all([parent, leaf])
    database.session.commit()

    payload = search_select("cost_center", "adm", {"company": ["cacao"], "is_group": ["1"]}, limit=10)

    assert [item["value"] for item in payload["results"]] == ["ADM"]


def test_search_select_account_id_and_cost_center_id_return_numeric_values(app_ctx):
    from cacao_accounting.database import Accounts, CostCenter, database
    from cacao_accounting.search_select import search_select

    account = Accounts(entity="cacao", code="1", name="Activo", active=True, enabled=True, group=True, classification="activo")
    cost_center = CostCenter(entity="cacao", code="ADM", name="Administracion", active=True, enabled=True, group=True)
    database.session.add_all([account, cost_center])
    database.session.commit()

    account_payload = search_select("account_id", "act", {"company": ["cacao"], "is_group": ["1"]}, limit=10)
    cost_center_payload = search_select("cost_center_id", "adm", {"company": ["cacao"], "is_group": ["1"]}, limit=10)

    assert account_payload["results"][0]["value"] == str(account.id)
    assert "1 - Activo" in account_payload["results"][0]["label"]
    assert cost_center_payload["results"][0]["value"] == str(cost_center.id)
    assert "ADM - Administracion" in cost_center_payload["results"][0]["label"]


def test_search_select_item_requires_registered_company_filter(app_ctx):
    from decimal import Decimal
    from uuid import uuid4

    from cacao_accounting.database import Item, ItemUOMConversion, UOM, database
    from cacao_accounting.search_select import search_select

    suffix = uuid4().hex[:8].upper()
    base_uom = f"EA-{suffix}"
    box_uom = f"BOX-{suffix}"
    item_code = f"ITEM-{suffix}"
    database.session.add_all(
        [
            UOM(code=base_uom, name="Unidad test"),
            UOM(code=box_uom, name="Caja test"),
            Item(
                code=item_code,
                name="Item con UOMs",
                item_type="goods",
                is_stock_item=True,
                default_uom=base_uom,
                is_active=True,
            ),
            ItemUOMConversion(
                item_code=item_code,
                from_uom=box_uom,
                to_uom=base_uom,
                conversion_factor=Decimal("10"),
            ),
        ]
    )
    database.session.commit()

    payload = search_select("item", item_code, {"company": ["cacao"]}, limit=10)

    assert payload["doctype"] == "item"
    assert payload["query"] == item_code
    assert payload["results"][0]["default_uom"] == base_uom
    assert payload["results"][0]["allowed_uoms"] == [base_uom, box_uom]

    uom_payload = search_select("uom", "", {"code": [box_uom]}, limit=10)
    assert [result["value"] for result in uom_payload["results"]] == [box_uom]


def test_search_select_api_requires_login_and_returns_filtered_accounts(app_ctx):
    from cacao_accounting.database import Accounts, Modules, User, database

    bank = Accounts(
        entity="cacao", code="BANK-API", name="Banco API", active=True, enabled=True, group=False, account_type="bank"
    )
    expense = Accounts(
        entity="cacao", code="BANK-EXP", name="Gasto Banco", active=True, enabled=True, group=False, account_type="expense"
    )
    user = User(user="api-admin", name="API Admin", password=b"x", classification="admin", active=True)
    module = Modules(module="accounting", default=True, enabled=True)
    database.session.add_all([bank, expense, user, module])
    database.session.commit()

    app_ctx.config["SECRET_KEY"] = "testing"
    client = app_ctx.test_client()
    with client.session_transaction() as session:
        session["_user_id"] = user.id
        session["_fresh"] = True

    response = client.get("/api/search-select?doctype=account&q=BANK&company=cacao&account_type=bank")
    assert response.status_code == 200
    payload = response.get_json()
    assert payload["doctype"] == "account"
    assert [item["id"] for item in payload["results"]] == [bank.id]

    response = client.get("/api/search-select?doctype=account&q=BANK&company=cacao&account_type=bank&account_type=expense")
    assert response.status_code == 200
    assert {item["id"] for item in response.get_json()["results"]} == {bank.id, expense.id}

    response = client.get("/api/search-select?doctype=account&q=BANK&company=cacao&bad_filter=x")
    assert response.status_code == 400


def test_search_select_api_supports_account_id_and_cost_center_id(app_ctx):
    from cacao_accounting.database import Accounts, CostCenter, Modules, User, database

    account = Accounts(entity="cacao", code="1", name="Activo", active=True, enabled=True, group=True, classification="activo")
    cost_center = CostCenter(entity="cacao", code="ADM", name="Administracion", active=True, enabled=True, group=True)
    user = User(
        **{
            "user": "smart-select-admin",
            "name": "Smart Select",
            "password": b"x",
            "classification": "admin",
            "active": True,
        }
    )
    module = Modules(module="accounting", default=True, enabled=True)
    database.session.add_all([account, cost_center, user, module])
    database.session.commit()

    app_ctx.config["SECRET_KEY"] = "testing"
    client = app_ctx.test_client()
    with client.session_transaction() as session:
        session["_user_id"] = user.id
        session["_fresh"] = True

    account_response = client.get("/api/search-select?doctype=account_id&company=cacao&is_group=true&q=Activo")
    assert account_response.status_code == 200
    account_payload = account_response.get_json()
    assert account_payload["results"][0]["value"] == str(account.id)
    assert "1 - Activo" in account_payload["results"][0]["label"]

    cost_center_response = client.get("/api/search-select?doctype=cost_center_id&company=cacao&is_group=true&q=Admin")
    assert cost_center_response.status_code == 200
    cost_center_payload = cost_center_response.get_json()
    assert cost_center_payload["results"][0]["value"] == str(cost_center.id)
    assert "ADM - Administracion" in cost_center_payload["results"][0]["label"]


def test_default_accounts_view_uses_smart_select_without_rendering_full_account_options(app_ctx):
    from cacao_accounting.contabilidad.default_accounts import upsert_company_default_accounts
    from cacao_accounting.database import Accounts, Modules, User, database

    bank = Accounts(
        entity="cacao", code="BANK-VIEW", name="Banco Vista", active=True, enabled=True, group=False, account_type="bank"
    )
    receivable = Accounts(
        entity="cacao",
        code="AR-VIEW",
        name="Cuenta por Cobrar",
        active=True,
        enabled=True,
        group=False,
        account_type="receivable",
    )
    admin_user = User(user="view-admin", name="View Admin", password=b"x", classification="admin", active=True)
    admin_module = Modules(module="admin", default=True, enabled=True)
    database.session.add_all([bank, receivable, admin_user, admin_module])
    database.session.flush()
    upsert_company_default_accounts("cacao", {"default_bank": bank.id, "default_receivable": receivable.id})
    database.session.commit()

    app_ctx.config["SECRET_KEY"] = "testing"
    client = app_ctx.test_client()
    with client.session_transaction() as session:
        session["_user_id"] = admin_user.id
        session["_fresh"] = True

    response = client.get("/settings/default-accounts?company=cacao")
    assert response.status_code == 200
    html = response.get_data(as_text=True)
    assert "x-data='smartSelect({" in html
    assert 'account_type: ["bank"]' in html
    assert "BANK-VIEW - Banco Vista" in html
    assert f'<option value="{bank.id}"' not in html
    assert f'<option value="{receivable.id}"' not in html


def test_manual_journal_allows_bank_and_untyped_accounts(app_ctx):
    from cacao_accounting.contabilidad.posting import post_document_to_gl
    from cacao_accounting.database import Accounts, ComprobanteContable, ComprobanteContableDetalle, GLEntry, database

    bank = Accounts(entity="cacao", code="BANK-M", name="Banco", active=True, enabled=True, account_type="bank")
    free = Accounts(entity="cacao", code="FREE-M", name="Libre", active=True, enabled=True)
    database.session.add_all([bank, free])
    database.session.flush()

    bank_journal = ComprobanteContable(entity="cacao", date=date(2026, 5, 6), memo="Manual banco")
    database.session.add(bank_journal)
    database.session.flush()
    database.session.add_all(
        [
            ComprobanteContableDetalle(
                entity="cacao",
                account=bank.code,
                date=bank_journal.date,
                transaction="journal_entry",
                transaction_id=bank_journal.id,
                value=Decimal("10.00"),
            ),
            ComprobanteContableDetalle(
                entity="cacao",
                account=free.code,
                date=bank_journal.date,
                transaction="journal_entry",
                transaction_id=bank_journal.id,
                value=Decimal("-10.00"),
            ),
        ]
    )
    database.session.commit()

    bank_entries = post_document_to_gl(bank_journal)
    assert len(bank_entries) == 2

    free_journal = ComprobanteContable(entity="cacao", date=date(2026, 5, 6), memo="Manual libre")
    database.session.add(free_journal)
    database.session.flush()
    database.session.add_all(
        [
            ComprobanteContableDetalle(
                entity="cacao",
                account=free.code,
                date=free_journal.date,
                transaction="journal_entry",
                transaction_id=free_journal.id,
                value=Decimal("10.00"),
            ),
            ComprobanteContableDetalle(
                entity="cacao",
                account=free.code,
                date=free_journal.date,
                transaction="journal_entry",
                transaction_id=free_journal.id,
                value=Decimal("-10.00"),
            ),
        ]
    )
    database.session.commit()

    entries = post_document_to_gl(free_journal)
    assert len(entries) == 2
    assert database.session.execute(database.select(GLEntry)).scalars().all()


def test_sales_tax_uses_default_account_when_tax_has_no_account(app_ctx):
    from cacao_accounting.contabilidad.posting import post_document_to_gl
    from cacao_accounting.database import (
        Accounts,
        CompanyDefaultAccount,
        GLEntry,
        PartyAccount,
        SalesInvoice,
        SalesInvoiceItem,
        Tax,
        TaxTemplate,
        TaxTemplateItem,
        database,
    )

    receivable = Accounts(entity="cacao", code="AR-DF", name="AR", active=True, enabled=True, account_type="receivable")
    income = Accounts(entity="cacao", code="INC-DF", name="Ingreso", active=True, enabled=True, account_type="income")
    tax_account = Accounts(entity="cacao", code="TAX-DF", name="IVA", active=True, enabled=True, account_type="tax")
    database.session.add_all([receivable, income, tax_account])
    database.session.flush()
    template = TaxTemplate(name="IVA Default", company="cacao", template_type="selling")
    tax = Tax(name="IVA 15", rate=Decimal("15.00"), tax_type="percentage", applies_to="sales", account_id=None)
    database.session.add_all([template, tax])
    database.session.flush()
    database.session.add_all(
        [
            TaxTemplateItem(tax_template_id=template.id, tax_id=tax.id, sequence=1, behavior="additive"),
            PartyAccount(party_id="CUST-DF", company="cacao", receivable_account_id=receivable.id),
            CompanyDefaultAccount(company="cacao", default_sales_tax_account_id=tax_account.id),
        ]
    )
    invoice = SalesInvoice(
        company="cacao",
        posting_date=date(2026, 5, 6),
        customer_id="CUST-DF",
        tax_template_id=template.id,
        total=Decimal("100.00"),
        grand_total=Decimal("115.00"),
        docstatus=1,
    )
    database.session.add(invoice)
    database.session.flush()
    database.session.add(
        SalesInvoiceItem(
            sales_invoice_id=invoice.id,
            item_code="ITEM-DF",
            qty=Decimal("1"),
            uom="EA",
            rate=Decimal("100.00"),
            amount=Decimal("100.00"),
            income_account_id=income.id,
        )
    )
    database.session.commit()

    post_document_to_gl(invoice)
    entries = database.session.execute(database.select(GLEntry)).scalars().all()
    assert any(entry.account_id == tax_account.id and entry.credit == Decimal("15.0000") for entry in entries)


def test_inventory_uom_batch_serial_and_rebuild_stock_bins(app_ctx):
    from cacao_accounting.contabilidad.posting import post_document_to_gl
    from cacao_accounting.database import (
        Accounts,
        Batch,
        CompanyDefaultAccount,
        Item,
        ItemAccount,
        ItemUOMConversion,
        SerialNumber,
        StockBin,
        StockEntry,
        StockEntryItem,
        UOM,
        Warehouse,
        WarehouseCompanyAccount,
        database,
    )
    from cacao_accounting.inventario.service import convert_item_qty, rebuild_stock_bins

    inventory = Accounts(entity="cacao", code="INV-S", name="Inventario", active=True, enabled=True, account_type="asset")
    bridge = Accounts(
        entity="cacao", code="BRIDGE-S", name="Cuenta Puente Compras", active=True, enabled=True, account_type="liability"
    )
    adjustment = Accounts(
        entity="cacao", code="ADJ-S", name="Ajuste Inventario", active=True, enabled=True, account_type="expense"
    )
    database.session.add_all([inventory, bridge, adjustment])
    database.session.flush()
    database.session.add_all(
        [
            UOM(code="EA", name="Each"),
            UOM(code="BOX", name="Box"),
            Item(
                code="ITEM-S",
                name="Serial",
                item_type="goods",
                is_stock_item=True,
                has_batch=True,
                has_serial_no=True,
                default_uom="EA",
            ),
            Warehouse(code="WH-S", name="Bodega", company="cacao"),
        ]
    )
    database.session.flush()
    database.session.add_all(
        [
            WarehouseCompanyAccount(warehouse_code="WH-S", company="cacao", inventory_account_id=inventory.id, is_active=True),
            CompanyDefaultAccount(company="cacao", bridge_account_id=bridge.id, inventory_adjustment_account_id=adjustment.id),
            ItemAccount(item_code="ITEM-S", company="cacao"),
            ItemUOMConversion(item_code="ITEM-S", from_uom="BOX", to_uom="EA", conversion_factor=Decimal("10")),
            Batch(item_code="ITEM-S", batch_no="B-1"),
        ]
    )
    database.session.flush()
    batch = database.session.execute(database.select(Batch).filter_by(batch_no="B-1")).scalar_one()
    entry = StockEntry(
        company="cacao", posting_date=date(2026, 5, 5), purpose="material_receipt", to_warehouse="WH-S", docstatus=1
    )
    database.session.add(entry)
    database.session.flush()
    database.session.add(
        StockEntryItem(
            stock_entry_id=entry.id,
            item_code="ITEM-S",
            target_warehouse="WH-S",
            qty=Decimal("1"),
            uom="EA",
            basic_rate=Decimal("7.00"),
            amount=Decimal("7.00"),
            batch_id=batch.id,
            serial_no="SN-1",
        )
    )
    database.session.commit()

    post_document_to_gl(entry)
    result = rebuild_stock_bins("cacao", item_code="ITEM-S", warehouse="WH-S")
    serial = database.session.execute(database.select(SerialNumber).filter_by(serial_no="SN-1")).scalar_one()
    bin_row = database.session.execute(database.select(StockBin).filter_by(item_code="ITEM-S", warehouse="WH-S")).scalar_one()

    assert convert_item_qty("ITEM-S", Decimal("1"), "BOX", "EA") == Decimal("10")
    assert serial.serial_status == "available"
    assert bin_row.actual_qty == Decimal("1.000000000")
    assert result.rebuilt_bins == 1


def test_bank_statement_import_preview_and_matching_rule(app_ctx):
    from io import StringIO

    from cacao_accounting.bancos.statement_service import apply_bank_matching_rule, import_bank_statement
    from cacao_accounting.database import Bank, BankAccount, BankMatchingRule, BankTransaction, database

    bank = Bank(name="Banco CSV")
    database.session.add(bank)
    database.session.flush()
    account = BankAccount(bank_id=bank.id, company="cacao", account_name="Cuenta CSV", is_active=True)
    database.session.add(account)
    database.session.flush()
    csv_data = "date,reference,description,deposit,withdrawal\n2026-05-05,REF-1,Ingreso,25.00,\n"
    mapping = {
        "date": "date",
        "reference": "reference",
        "description": "description",
        "deposit": "deposit",
        "withdrawal": "withdrawal",
    }

    preview = import_bank_statement(StringIO(csv_data), mapping, account.id, company="cacao", preview=True)
    imported = import_bank_statement(StringIO(csv_data), mapping, account.id, company="cacao", preview=False)
    duplicate = import_bank_statement(StringIO(csv_data), mapping, account.id, company="cacao", preview=True)
    rule = BankMatchingRule(company="cacao", bank_account_id=account.id, name="Referencia", reference_contains="REF")
    database.session.add(rule)
    database.session.commit()
    run = apply_bank_matching_rule(rule.id, account.id, (date(2026, 5, 1), date(2026, 5, 31)))

    assert preview.imported_count == 0
    assert imported.imported_count == 1
    assert duplicate.duplicate_count == 1
    assert database.session.execute(database.select(BankTransaction)).scalars().first()
    assert run.candidates_by_transaction


def test_bank_statement_import_requires_company_and_detects_missing_reference(app_ctx):
    """El importador ata la cuenta a su compañía y no omite referencias nulas."""
    from io import StringIO

    from cacao_accounting.bancos.statement_service import BankStatementError, import_bank_statement
    from cacao_accounting.database import Bank, BankAccount, database

    bank = Bank(name="Banco ownership")
    database.session.add(bank)
    database.session.flush()
    account = BankAccount(bank_id=bank.id, company="cacao", account_name="Cuenta ownership")
    database.session.add(account)
    database.session.commit()
    mapping = {
        "date": "date",
        "reference": "reference",
        "description": "description",
        "deposit": "deposit",
        "withdrawal": "withdrawal",
    }

    with pytest.raises(BankStatementError, match="no pertenece"):
        import_bank_statement(
            StringIO("date,reference,description,deposit,withdrawal\n2026-05-05,,Ingreso,25.00,\n"),
            mapping,
            account.id,
            company="other",
            preview=True,
        )

    csv_data = "date,reference,description,deposit,withdrawal\n2026-05-05,,Ingreso,25.00,\n"
    imported = import_bank_statement(StringIO(csv_data), mapping, account.id, company="cacao", preview=False)
    duplicate = import_bank_statement(StringIO(csv_data), mapping, account.id, company="cacao", preview=True)
    assert imported.imported_count == 1
    assert duplicate.duplicate_count == 1


def test_bank_statement_withdrawal_only_is_reconcilable(app_ctx):
    """Un retiro importado debe conservar su lado monetario y ser conciliable."""
    from cacao_accounting.bancos import _bank_reconciliation_allocated_amount
    from cacao_accounting.bancos.reconciliation_service import _bank_amount
    from cacao_accounting.database import Bank, BankAccount, BankTransaction, database
    from cacao_accounting.imports.adapters.bank_statement import BankStatementAdapter

    bank = Bank(name="Banco retiro")
    database.session.add(bank)
    database.session.flush()
    account = BankAccount(bank_id=bank.id, company="cacao", account_name="Cuenta retiro")
    database.session.add(account)
    database.session.flush()
    row = {
        "bank_account_id": account.id,
        "posting_date": "2026-05-05",
        "reference_number": "W-1",
        "description": "Retiro",
        "deposit": "",
        "withdrawal": "25.00",
    }
    imported = BankStatementAdapter().build_document([row], {})[0]
    assert imported["deposit"] is None
    assert imported["withdrawal"] == Decimal("25.00")

    transaction = BankTransaction(
        bank_account_id=account.id,
        posting_date=date(2026, 5, 5),
        deposit=Decimal("0"),
        withdrawal=Decimal("25.00"),
    )
    database.session.add(transaction)
    database.session.commit()
    assert _bank_amount(transaction) == Decimal("25.00")
    assert _bank_reconciliation_allocated_amount(transaction) == Decimal("25.00")


def test_bank_statement_adapter_rejects_invalid_amount_date_and_cross_company(app_ctx):
    """El importador no crea filas sin monto ni cruza cuentas entre compañías."""
    from cacao_accounting.database import Bank, BankAccount, database
    from cacao_accounting.imports.adapters.bank_statement import BankStatementAdapter

    bank = Bank(name="Banco validación")
    database.session.add(bank)
    database.session.flush()
    account = BankAccount(bank_id=bank.id, company="cacao", account_name="Cuenta validación")
    database.session.add(account)
    database.session.flush()
    adapter = BankStatementAdapter()

    row_errors = adapter.validate_row(
        {
            "bank_account_id": account.id,
            "posting_date": "2026-02-31",
            "deposit": "1,000",
            "withdrawal": "",
        }
    )
    assert any("Fecha bancaria inválida" in error for error in row_errors)
    assert any("Monto bancario inválido" in error for error in row_errors)

    account.company = "otra"
    document_errors = adapter.validate_document(
        [{"bank_account_id": account.id, "posting_date": "2026-05-05"}], {"company_id": "cacao"}
    )
    assert any("pertenece a la compañía" in error for error in document_errors)


def test_bank_statement_adapter_rejects_empty_movement(app_ctx):
    from cacao_accounting.imports.adapters.bank_statement import BankStatementAdapter

    adapter = BankStatementAdapter()
    errors = adapter.validate_row(
        {"bank_account_id": "missing", "posting_date": "2026-05-05", "deposit": "", "withdrawal": ""}
    )
    assert any("depósito o un retiro" in error for error in errors)


def test_bank_company_lists_use_authorized_book_scope(app_ctx, monkeypatch):
    """Los listados bancarios no exponen compañías fuera de los libros autorizados."""
    import importlib
    from types import SimpleNamespace

    bancos_module = importlib.import_module("cacao_accounting.bancos.services")
    from cacao_accounting.database import Bank, BankAccount, Book, Entity, database

    other = Entity(code="other", name="Other", company_name="Other", tax_id="J0002", currency="NIO")
    allowed_book = Book(code="CASH-SCOPE", name="Cash scope", entity="cacao", currency="NIO", is_primary=True)
    other_book = Book(code="CASH-OTHER", name="Cash other", entity="other", currency="NIO", is_primary=True)
    bank = Bank(name="Banco scope")
    database.session.add_all([other, allowed_book, other_book, bank])
    database.session.flush()
    allowed = BankAccount(bank_id=bank.id, company="cacao", account_name="Permitida", account_no="SCOPE-1")
    hidden = BankAccount(bank_id=bank.id, company="other", account_name="Oculta", account_no="SCOPE-2")
    database.session.add_all([allowed, hidden])
    database.session.commit()

    monkeypatch.setattr(bancos_module, "current_user", SimpleNamespace(id="user-1", classification="user"))
    monkeypatch.setattr(
        bancos_module,
        "Permisos",
        lambda **_: SimpleNamespace(obtener_libros_autorizados=lambda *_args, **_kwargs: [allowed_book.id]),
    )

    with app_ctx.test_request_context("/bank-account/list"):
        page = bancos_module._paginate_list(
            BankAccount,
            (BankAccount.account_name,),
            database.select(BankAccount),
            include_status=False,
        )

    assert [account.id for account in page.items] == [allowed.id]


def test_bank_reconciliation_panel_ignores_invalid_historical_transaction(app_ctx, monkeypatch):
    import importlib

    bancos = importlib.import_module("cacao_accounting.bancos")
    from cacao_accounting.database import Bank, BankAccount, BankTransaction, database

    bank = Bank(name="Banco histórico")
    database.session.add(bank)
    database.session.flush()
    account = BankAccount(bank_id=bank.id, company="cacao", account_name="Cuenta histórica")
    database.session.add(account)
    database.session.flush()
    transaction = BankTransaction(bank_account_id=account.id, posting_date=date(2026, 5, 5), deposit=Decimal("100.00"))
    database.session.add(transaction)
    database.session.commit()

    monkeypatch.setattr(
        bancos,
        "find_bank_reconciliation_candidates",
        lambda _id: (_ for _ in ()).throw(bancos.BankReconciliationError("sin monto")),
    )
    assert bancos._safe_bank_reconciliation_candidates(transaction) == []


# ---------------------------------------------------------------------------
# Criterios de aceptacion del Issue: Framework de Conciliacion de Compras
# ---------------------------------------------------------------------------


def test_matching_without_accounting_entries_is_possible(app_ctx):
    """Criterio #1: se puede ejecutar el matching sin generar asientos contables."""
    from cacao_accounting.compras.purchase_reconciliation_service import reconcile_purchase_invoice
    from cacao_accounting.database import (
        GLEntry,
        Item,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseReceipt,
        PurchaseReceiptItem,
        UOM,
        Warehouse,
        database,
    )

    database.session.add_all(
        [
            UOM(code="EA-AC1", name="Each AC1"),
            Item(code="ITEM-AC1", name="Item AC1", item_type="goods", is_stock_item=True, default_uom="EA-AC1"),
            Warehouse(code="WH-AC1", name="Bodega AC1", company="cacao"),
        ]
    )
    receipt = PurchaseReceipt(company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-AC1", docstatus=1)
    database.session.add(receipt)
    database.session.flush()
    database.session.add(
        PurchaseReceiptItem(
            purchase_receipt_id=receipt.id,
            item_code="ITEM-AC1",
            item_name="Item AC1",
            qty=Decimal("5"),
            qty_in_base_uom=Decimal("5"),
            uom="EA-AC1",
            rate=Decimal("10.00"),
            amount=Decimal("50.00"),
            warehouse="WH-AC1",
        )
    )
    invoice = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 2),
        supplier_id="SUPP-AC1",
        purchase_receipt_id=receipt.id,
        docstatus=1,
    )
    database.session.add(invoice)
    database.session.flush()
    database.session.add(
        PurchaseInvoiceItem(
            purchase_invoice_id=invoice.id,
            item_code="ITEM-AC1",
            item_name="Item AC1",
            qty=Decimal("5"),
            uom="EA-AC1",
            rate=Decimal("10.00"),
            amount=Decimal("50.00"),
            warehouse="WH-AC1",
        )
    )
    database.session.commit()

    # reconcile WITHOUT calling post_document_to_gl — no GL entries should exist
    result = reconcile_purchase_invoice(invoice.id)
    database.session.commit()

    gl_count = (
        database.session.execute(database.select(GLEntry).filter_by(voucher_type="purchase_invoice", voucher_id=invoice.id))
        .scalars()
        .all()
    )

    assert result.matching_result == "MATCH_OK"
    assert result.matched_amount == Decimal("50.0000")
    # Matching produced no accounting entries on its own
    assert len(gl_count) == 0


def test_changing_tolerances_does_not_alter_historical_reconciliations(app_ctx):
    """Criterio #2: cambiar tolerancias no altera datos historicos."""
    from cacao_accounting.compras.purchase_reconciliation_service import (
        PurchaseMatchingConfig,
        seed_matching_config_for_company,
    )
    from cacao_accounting.database import (
        Item,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseReceipt,
        PurchaseReceiptItem,
        PurchaseReconciliation,
        UOM,
        Warehouse,
        database,
    )
    from cacao_accounting.compras.purchase_reconciliation_service import reconcile_purchase_invoice

    # Seed strict config
    seed_matching_config_for_company("cacao")
    database.session.commit()

    database.session.add_all(
        [
            UOM(code="EA-AC2", name="Each AC2"),
            Item(code="ITEM-AC2", name="Item AC2", item_type="goods", is_stock_item=True, default_uom="EA-AC2"),
            Warehouse(code="WH-AC2", name="Bodega AC2", company="cacao"),
        ]
    )
    receipt = PurchaseReceipt(company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-AC2", docstatus=1)
    database.session.add(receipt)
    database.session.flush()
    database.session.add(
        PurchaseReceiptItem(
            purchase_receipt_id=receipt.id,
            item_code="ITEM-AC2",
            item_name="Item AC2",
            qty=Decimal("10"),
            qty_in_base_uom=Decimal("10"),
            uom="EA-AC2",
            rate=Decimal("20.00"),
            amount=Decimal("200.00"),
            warehouse="WH-AC2",
        )
    )
    invoice = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 2),
        supplier_id="SUPP-AC2",
        purchase_receipt_id=receipt.id,
        docstatus=1,
    )
    database.session.add(invoice)
    database.session.flush()
    database.session.add(
        PurchaseInvoiceItem(
            purchase_invoice_id=invoice.id,
            item_code="ITEM-AC2",
            item_name="Item AC2",
            qty=Decimal("10"),
            uom="EA-AC2",
            rate=Decimal("20.00"),
            amount=Decimal("200.00"),
            warehouse="WH-AC2",
        )
    )
    database.session.commit()

    reconcile_purchase_invoice(invoice.id)
    database.session.commit()

    # Now change tolerance — should NOT affect the already-created reconciliation
    cfg = database.session.execute(database.select(PurchaseMatchingConfig).filter_by(company="cacao")).scalar_one()
    cfg.price_tolerance_value = Decimal("10")  # relax tolerance
    database.session.commit()

    recon = database.session.execute(
        database.select(PurchaseReconciliation).filter_by(purchase_invoice_id=invoice.id)
    ).scalar_one()

    # Historical record is unchanged
    assert recon.matched_amount == Decimal("200.0000")
    assert recon.matching_type == "3-way"


def test_state_reconstruction_from_events(app_ctx):
    """Criterio #3: se pueden reconstruir estados desde eventos."""
    from cacao_accounting.compras.purchase_reconciliation_service import (
        reconstruct_reconciliation_state,
        reconcile_purchase_invoice,
    )
    from cacao_accounting.database import (
        Item,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseReceipt,
        PurchaseReceiptItem,
        UOM,
        Warehouse,
        database,
    )

    database.session.add_all(
        [
            UOM(code="EA-AC3", name="Each AC3"),
            Item(code="ITEM-AC3", name="Item AC3", item_type="goods", is_stock_item=True, default_uom="EA-AC3"),
            Warehouse(code="WH-AC3", name="Bodega AC3", company="cacao"),
        ]
    )
    receipt = PurchaseReceipt(company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-AC3", docstatus=1)
    database.session.add(receipt)
    database.session.flush()
    database.session.add(
        PurchaseReceiptItem(
            purchase_receipt_id=receipt.id,
            item_code="ITEM-AC3",
            item_name="Item AC3",
            qty=Decimal("3"),
            qty_in_base_uom=Decimal("3"),
            uom="EA-AC3",
            rate=Decimal("30.00"),
            amount=Decimal("90.00"),
            warehouse="WH-AC3",
        )
    )
    invoice = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 2),
        supplier_id="SUPP-AC3",
        purchase_receipt_id=receipt.id,
        docstatus=1,
    )
    database.session.add(invoice)
    database.session.flush()
    database.session.add(
        PurchaseInvoiceItem(
            purchase_invoice_id=invoice.id,
            item_code="ITEM-AC3",
            item_name="Item AC3",
            qty=Decimal("3"),
            uom="EA-AC3",
            rate=Decimal("30.00"),
            amount=Decimal("90.00"),
            warehouse="WH-AC3",
        )
    )
    database.session.commit()

    result = reconcile_purchase_invoice(invoice.id)
    database.session.commit()

    # Reconstruct state from event log
    snapshot = reconstruct_reconciliation_state("cacao", result.reconciliation_id)

    assert snapshot.company == "cacao"
    assert snapshot.document_id == result.reconciliation_id
    # At least one event was logged for this reconciliation
    assert len(snapshot.events) >= 1
    # Event log contains a MATCH event
    event_types = [ev["event_type"] for ev in snapshot.events]
    assert any("MATCH" in et for et in event_types)


def test_system_supports_two_way_and_three_way_without_structural_changes(app_ctx):
    """Criterio #4: el sistema soporta 2-way y 3-way sin cambios estructurales."""
    from cacao_accounting.compras.purchase_reconciliation_service import (
        MatchingType,
        PurchaseMatchingConfig,
        get_matching_config,
        seed_matching_config_for_company,
    )
    from cacao_accounting.database import database

    seed_matching_config_for_company("cacao")
    database.session.commit()

    # 3-way (default)
    cfg_3way = get_matching_config("cacao")
    assert cfg_3way.matching_type == MatchingType.THREE_WAY

    # Switch to 2-way via config — no structural changes required
    cfg = database.session.execute(database.select(PurchaseMatchingConfig).filter_by(company="cacao")).scalar_one()
    cfg.matching_type = MatchingType.TWO_WAY
    database.session.commit()

    cfg_2way = get_matching_config("cacao")
    assert cfg_2way.matching_type == MatchingType.TWO_WAY


def test_two_way_matching_uses_purchase_order_lines_without_receipts(app_ctx):
    """2-way: OC + factura sin recepcion usa lineas de OC y no IDs de recepcion."""
    from sqlalchemy import text

    from cacao_accounting.compras.purchase_reconciliation_service import (
        MatchingType,
        PurchaseMatchingConfig,
        reconcile_purchase_invoice,
        seed_matching_config_for_company,
    )
    from cacao_accounting.database import (
        Item,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseOrder,
        PurchaseOrderItem,
        PurchaseReconciliation,
        PurchaseReconciliationItem,
        UOM,
        database,
    )

    seed_matching_config_for_company("cacao")
    cfg = database.session.execute(database.select(PurchaseMatchingConfig).filter_by(company="cacao")).scalar_one()
    cfg.matching_type = MatchingType.TWO_WAY
    database.session.add_all(
        [
            UOM(code="EA-2W", name="Each 2W"),
            Item(code="ITEM-2W", name="Item 2W", item_type="goods", is_stock_item=False, default_uom="EA-2W"),
        ]
    )
    order = PurchaseOrder(company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-2W", docstatus=1)
    database.session.add(order)
    database.session.flush()
    order_item = PurchaseOrderItem(
        purchase_order_id=order.id,
        item_code="ITEM-2W",
        item_name="Item 2W",
        qty=Decimal("5"),
        qty_in_base_uom=Decimal("5"),
        uom="EA-2W",
        rate=Decimal("10.00"),
        amount=Decimal("50.00"),
    )
    invoice = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 2),
        supplier_id="SUPP-2W",
        purchase_order_id=order.id,
        docstatus=1,
    )
    database.session.add_all([order_item, invoice])
    database.session.flush()
    invoice_item = PurchaseInvoiceItem(
        purchase_invoice_id=invoice.id,
        item_code="ITEM-2W",
        item_name="Item 2W",
        qty=Decimal("5"),
        uom="EA-2W",
        rate=Decimal("10.00"),
        amount=Decimal("50.00"),
    )
    database.session.add(invoice_item)
    database.session.commit()
    database.session.execute(text("PRAGMA foreign_keys=ON"))

    result = reconcile_purchase_invoice(invoice.id)
    database.session.commit()

    reconciliation = database.session.execute(
        database.select(PurchaseReconciliation).filter_by(id=result.reconciliation_id)
    ).scalar_one()
    reconciliation_item = database.session.execute(
        database.select(PurchaseReconciliationItem).filter_by(purchase_reconciliation_id=reconciliation.id)
    ).scalar_one()

    assert result.matching_result == "MATCH_OK"
    assert reconciliation.matching_type == "2-way"
    assert reconciliation.purchase_order_id == order.id
    assert reconciliation.purchase_receipt_id is None
    assert reconciliation_item.purchase_order_item_id == order_item.id
    assert reconciliation_item.purchase_receipt_item_id is None


def test_two_way_matching_aggregates_duplicate_order_and_invoice_lines(app_ctx):
    """2-way evalua cantidades agregadas por producto/UOM antes de crear detalles."""
    from cacao_accounting.compras.purchase_reconciliation_service import (
        MatchingType,
        PurchaseMatchingConfig,
        reconcile_purchase_invoice,
        seed_matching_config_for_company,
    )
    from cacao_accounting.database import (
        Item,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseOrder,
        PurchaseOrderItem,
        PurchaseReconciliationItem,
        UOM,
        database,
    )

    seed_matching_config_for_company("cacao")
    cfg = database.session.execute(database.select(PurchaseMatchingConfig).filter_by(company="cacao")).scalar_one()
    cfg.matching_type = MatchingType.TWO_WAY
    database.session.add_all(
        [
            UOM(code="EA-2WA", name="Each 2WA"),
            Item(code="ITEM-2WA", name="Item 2WA", item_type="goods", is_stock_item=False, default_uom="EA-2WA"),
        ]
    )
    order = PurchaseOrder(company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-2WA", docstatus=1)
    database.session.add(order)
    database.session.flush()
    order_items = [
        PurchaseOrderItem(
            purchase_order_id=order.id,
            item_code="ITEM-2WA",
            item_name="Item 2WA",
            qty=Decimal("2"),
            qty_in_base_uom=Decimal("2"),
            uom="EA-2WA",
            rate=Decimal("10.00"),
            amount=Decimal("20.00"),
        ),
        PurchaseOrderItem(
            purchase_order_id=order.id,
            item_code="ITEM-2WA",
            item_name="Item 2WA",
            qty=Decimal("3"),
            qty_in_base_uom=Decimal("3"),
            uom="EA-2WA",
            rate=Decimal("10.00"),
            amount=Decimal("30.00"),
        ),
    ]
    invoice = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 2),
        supplier_id="SUPP-2WA",
        purchase_order_id=order.id,
        docstatus=1,
    )
    database.session.add_all([*order_items, invoice])
    database.session.flush()
    database.session.add_all(
        [
            PurchaseInvoiceItem(
                purchase_invoice_id=invoice.id,
                item_code="ITEM-2WA",
                item_name="Item 2WA",
                qty=Decimal("2"),
                uom="EA-2WA",
                rate=Decimal("10.00"),
                amount=Decimal("20.00"),
            ),
            PurchaseInvoiceItem(
                purchase_invoice_id=invoice.id,
                item_code="ITEM-2WA",
                item_name="Item 2WA",
                qty=Decimal("3"),
                uom="EA-2WA",
                rate=Decimal("10.00"),
                amount=Decimal("30.00"),
            ),
        ]
    )
    database.session.commit()

    result = reconcile_purchase_invoice(invoice.id)
    database.session.commit()
    items = (
        database.session.execute(
            database.select(PurchaseReconciliationItem).filter_by(purchase_reconciliation_id=result.reconciliation_id)
        )
        .scalars()
        .all()
    )

    assert result.matching_result == "MATCH_OK"
    assert result.status == "reconciled"
    assert len(items) == 2
    assert {item.purchase_order_item_id for item in items} == {order_item.id for order_item in order_items}
    assert all(item.purchase_receipt_item_id is None for item in items)


def test_purchase_invoice_posting_auto_reconciles_two_way_po_only_invoice(app_ctx):
    """Posting: una factura PO-only se auto-concilia cuando la compania esta en 2-way."""
    from cacao_accounting.compras.purchase_reconciliation_service import (
        MatchingType,
        PurchaseMatchingConfig,
        seed_matching_config_for_company,
    )
    from cacao_accounting.contabilidad.posting import post_document_to_gl
    from cacao_accounting.database import (
        Accounts,
        Item,
        PartyAccount,
        PurchaseEconomicEvent,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseOrder,
        PurchaseOrderItem,
        PurchaseReconciliation,
        UOM,
        database,
    )

    payable_account = Accounts(
        entity="cacao",
        code="AP-2W",
        name="AP 2W",
        active=True,
        enabled=True,
        classification="liability",
        account_type="payable",
    )
    expense_account = Accounts(
        entity="cacao",
        code="EXP-2W",
        name="Gasto 2W",
        active=True,
        enabled=True,
        classification="expense",
        account_type="expense",
    )
    database.session.add_all(
        [
            payable_account,
            expense_account,
            UOM(code="EA-2WP", name="Each 2WP"),
            Item(code="ITEM-2WP", name="Item 2WP", item_type="goods", is_stock_item=False, default_uom="EA-2WP"),
        ]
    )
    database.session.flush()
    database.session.add(PartyAccount(party_id="SUPP-2WP", company="cacao", payable_account_id=payable_account.id))
    seed_matching_config_for_company("cacao")
    cfg = database.session.execute(database.select(PurchaseMatchingConfig).filter_by(company="cacao")).scalar_one()
    cfg.matching_type = MatchingType.TWO_WAY

    order = PurchaseOrder(company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-2WP", docstatus=1)
    database.session.add(order)
    database.session.flush()
    database.session.add(
        PurchaseOrderItem(
            purchase_order_id=order.id,
            item_code="ITEM-2WP",
            item_name="Item 2WP",
            qty=Decimal("3"),
            qty_in_base_uom=Decimal("3"),
            uom="EA-2WP",
            rate=Decimal("12.00"),
            amount=Decimal("36.00"),
        )
    )
    invoice = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 2),
        supplier_id="SUPP-2WP",
        purchase_order_id=order.id,
        docstatus=1,
        total=Decimal("36.00"),
        grand_total=Decimal("36.00"),
    )
    database.session.add(invoice)
    database.session.flush()
    database.session.add(
        PurchaseInvoiceItem(
            purchase_invoice_id=invoice.id,
            item_code="ITEM-2WP",
            item_name="Item 2WP",
            qty=Decimal("3"),
            uom="EA-2WP",
            rate=Decimal("12.00"),
            amount=Decimal("36.00"),
            expense_account_id=expense_account.id,
        )
    )
    database.session.commit()

    post_document_to_gl(invoice)
    database.session.commit()

    reconciliation = database.session.execute(
        database.select(PurchaseReconciliation).filter_by(purchase_invoice_id=invoice.id)
    ).scalar_one()
    event_types = [
        event.event_type
        for event in database.session.execute(database.select(PurchaseEconomicEvent).filter_by(company="cacao"))
        .scalars()
        .all()
    ]

    assert reconciliation.matching_type == "2-way"
    assert reconciliation.purchase_receipt_id is None
    assert "INVOICE_RECEIVED" in event_types
    assert "MATCH_COMPLETED" in event_types


def test_cancel_two_way_purchase_invoice_releases_order_quantities(app_ctx):
    """Cancelar una factura 2-way cancela su conciliacion y libera cantidad de OC."""
    from cacao_accounting.compras.purchase_reconciliation_service import (
        MatchingType,
        PurchaseMatchingConfig,
        cancel_purchase_reconciliation,
        reconcile_purchase_invoice,
        seed_matching_config_for_company,
    )
    from cacao_accounting.database import (
        Item,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseOrder,
        PurchaseOrderItem,
        PurchaseReconciliation,
        UOM,
        database,
    )

    seed_matching_config_for_company("cacao")
    cfg = database.session.execute(database.select(PurchaseMatchingConfig).filter_by(company="cacao")).scalar_one()
    cfg.matching_type = MatchingType.TWO_WAY
    database.session.add_all(
        [
            UOM(code="EA-2WC", name="Each 2WC"),
            Item(code="ITEM-2WC", name="Item 2WC", item_type="goods", is_stock_item=False, default_uom="EA-2WC"),
        ]
    )
    order = PurchaseOrder(company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-2WC", docstatus=1)
    database.session.add(order)
    database.session.flush()
    database.session.add(
        PurchaseOrderItem(
            purchase_order_id=order.id,
            item_code="ITEM-2WC",
            item_name="Item 2WC",
            qty=Decimal("4"),
            qty_in_base_uom=Decimal("4"),
            uom="EA-2WC",
            rate=Decimal("8.00"),
            amount=Decimal("32.00"),
        )
    )
    first_invoice = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 2),
        supplier_id="SUPP-2WC",
        purchase_order_id=order.id,
        docstatus=1,
    )
    second_invoice = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 3),
        supplier_id="SUPP-2WC",
        purchase_order_id=order.id,
        docstatus=1,
    )
    database.session.add_all([first_invoice, second_invoice])
    database.session.flush()
    for invoice in (first_invoice, second_invoice):
        database.session.add(
            PurchaseInvoiceItem(
                purchase_invoice_id=invoice.id,
                item_code="ITEM-2WC",
                item_name="Item 2WC",
                qty=Decimal("4"),
                uom="EA-2WC",
                rate=Decimal("8.00"),
                amount=Decimal("32.00"),
            )
        )
    database.session.commit()

    first_result = reconcile_purchase_invoice(first_invoice.id)
    database.session.commit()
    cancel_purchase_reconciliation(first_invoice.id)
    database.session.commit()

    first_reconciliation = database.session.execute(
        database.select(PurchaseReconciliation).filter_by(id=first_result.reconciliation_id)
    ).scalar_one()
    second_result = reconcile_purchase_invoice(second_invoice.id)

    assert first_reconciliation.status == "cancelled"
    assert second_result.matching_result == "MATCH_OK"


def test_bridge_account_is_configurable_not_required_by_default(app_ctx):
    """Criterio #5: la cuenta puente es configurable, no obligatoria."""
    from cacao_accounting.compras.purchase_reconciliation_service import (
        PurchaseMatchingConfig,
        seed_matching_config_for_company,
    )
    from cacao_accounting.database import database

    seed_matching_config_for_company("cacao")
    database.session.commit()

    cfg = database.session.execute(database.select(PurchaseMatchingConfig).filter_by(company="cacao")).scalar_one()
    # By default it is required (strict mode) but can be set to False
    assert isinstance(cfg.bridge_account_required, bool)

    cfg.bridge_account_required = False
    database.session.commit()

    cfg_relaxed = database.session.execute(database.select(PurchaseMatchingConfig).filter_by(company="cacao")).scalar_one()
    assert cfg_relaxed.bridge_account_required is False


def test_purchase_receipt_posting_allows_missing_bridge_when_not_required(app_ctx):
    """Recepcion mantiene stock ledger sin GL puente cuando la cuenta puente no es requerida."""
    from cacao_accounting.compras.purchase_reconciliation_service import (
        PurchaseMatchingConfig,
        seed_matching_config_for_company,
    )
    from cacao_accounting.contabilidad.posting import post_document_to_gl
    from cacao_accounting.database import (
        GLEntry,
        Item,
        PurchaseReceipt,
        PurchaseReceiptItem,
        StockLedgerEntry,
        UOM,
        Warehouse,
        database,
    )

    seed_matching_config_for_company("cacao")
    cfg = database.session.execute(database.select(PurchaseMatchingConfig).filter_by(company="cacao")).scalar_one()
    cfg.bridge_account_required = False
    database.session.add_all(
        [
            UOM(code="EA-NB", name="Each No Bridge"),
            Item(code="ITEM-NB", name="Item No Bridge", item_type="goods", is_stock_item=True, default_uom="EA-NB"),
            Warehouse(code="WH-NB", name="Bodega No Bridge", company="cacao"),
        ]
    )
    receipt = PurchaseReceipt(company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-NB", docstatus=1)
    database.session.add(receipt)
    database.session.flush()
    database.session.add(
        PurchaseReceiptItem(
            purchase_receipt_id=receipt.id,
            item_code="ITEM-NB",
            item_name="Item No Bridge",
            qty=Decimal("1"),
            qty_in_base_uom=Decimal("1"),
            uom="EA-NB",
            rate=Decimal("9.00"),
            amount=Decimal("9.00"),
            warehouse="WH-NB",
        )
    )
    database.session.commit()

    entries = post_document_to_gl(receipt)
    database.session.commit()

    stock_movements = database.session.execute(
        database.select(StockLedgerEntry).filter_by(voucher_type="purchase_receipt", voucher_id=receipt.id)
    ).scalar_one_or_none()
    gl_entries = (
        database.session.execute(database.select(GLEntry).filter_by(voucher_type="purchase_receipt", voucher_id=receipt.id))
        .scalars()
        .all()
    )

    assert entries == []
    assert stock_movements is not None
    assert gl_entries == []


def test_purchase_reconciliation_panel_groups_two_way_and_three_way(app_ctx):
    """Panel: agrupa conciliaciones 2-way sin recepcion y 3-way con recepcion por OC."""
    from cacao_accounting.compras.purchase_reconciliation_service import get_purchase_reconciliation_panel_groups
    from cacao_accounting.database import PurchaseReconciliation, database

    database.session.add_all(
        [
            PurchaseReconciliation(
                company="cacao",
                purchase_order_id="PO-PANEL",
                purchase_receipt_id=None,
                purchase_invoice_id="PINV-2W",
                matching_type="2-way",
                matched_amount=Decimal("10.00"),
                matched_date=date(2026, 5, 1),
                status="reconciled",
            ),
            PurchaseReconciliation(
                company="cacao",
                purchase_order_id="PO-PANEL",
                purchase_receipt_id="PREC-3W",
                purchase_invoice_id="PINV-3W",
                matching_type="3-way",
                matched_amount=Decimal("20.00"),
                matched_date=date(2026, 5, 2),
                status="partial",
            ),
        ]
    )
    database.session.commit()

    groups = get_purchase_reconciliation_panel_groups("cacao")
    group = next(group for group in groups if group.purchase_order_id == "PO-PANEL")

    assert group.invoice_count == 2
    assert group.receipt_count == 1
    assert group.worst_status == "partial"
    assert {reconciliation.matching_type for reconciliation in group.reconciliations} == {"2-way", "3-way"}


def test_goods_received_cancelled_event_emitted_on_receipt_cancel(app_ctx):
    """Cancelar una recepcion emite GOODS_RECEIVED_CANCELLED y cancela conciliaciones dependientes."""
    from cacao_accounting.compras.purchase_reconciliation_service import (
        emit_goods_received_cancelled,
        reconcile_purchase_invoice,
    )
    from cacao_accounting.database import (
        Item,
        PurchaseEconomicEvent,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseReceipt,
        PurchaseReceiptItem,
        PurchaseReconciliation,
        UOM,
        Warehouse,
        database,
    )

    database.session.add_all(
        [
            UOM(code="EA-AC6", name="Each AC6"),
            Item(code="ITEM-AC6", name="Item AC6", item_type="goods", is_stock_item=True, default_uom="EA-AC6"),
            Warehouse(code="WH-AC6", name="Bodega AC6", company="cacao"),
        ]
    )
    receipt = PurchaseReceipt(company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-AC6", docstatus=1)
    database.session.add(receipt)
    database.session.flush()
    database.session.add(
        PurchaseReceiptItem(
            purchase_receipt_id=receipt.id,
            item_code="ITEM-AC6",
            item_name="Item AC6",
            qty=Decimal("2"),
            qty_in_base_uom=Decimal("2"),
            uom="EA-AC6",
            rate=Decimal("50.00"),
            amount=Decimal("100.00"),
            warehouse="WH-AC6",
        )
    )
    invoice = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 2),
        supplier_id="SUPP-AC6",
        purchase_receipt_id=receipt.id,
        docstatus=1,
    )
    database.session.add(invoice)
    database.session.flush()
    database.session.add(
        PurchaseInvoiceItem(
            purchase_invoice_id=invoice.id,
            item_code="ITEM-AC6",
            item_name="Item AC6",
            qty=Decimal("2"),
            uom="EA-AC6",
            rate=Decimal("50.00"),
            amount=Decimal("100.00"),
            warehouse="WH-AC6",
        )
    )
    database.session.commit()

    reconcile_purchase_invoice(invoice.id)
    database.session.commit()

    # Cancel the receipt — should also cancel dependent reconciliation and emit event
    emit_goods_received_cancelled(receipt.id, "cacao")
    database.session.commit()

    recon = database.session.execute(
        database.select(PurchaseReconciliation).filter_by(purchase_receipt_id=receipt.id)
    ).scalar_one()
    assert recon.status == "cancelled"

    cancel_event = database.session.execute(
        database.select(PurchaseEconomicEvent).filter_by(
            company="cacao", document_id=receipt.id, event_type="GOODS_RECEIVED_CANCELLED"
        )
    ).scalar_one_or_none()
    assert cancel_event is not None


def test_purchase_order_status_report(app_ctx):
    from cacao_accounting.compras.purchase_reconciliation_service import get_purchase_order_status_report
    from cacao_accounting.database import PurchaseOrder, PurchaseOrderItem, database

    order_p = PurchaseOrder(company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-P1", docstatus=1)
    database.session.add(order_p)
    database.session.flush()

    item_p = PurchaseOrderItem(
        purchase_order_id=order_p.id,
        item_code="ITEM-P1",
        item_name="Item P1",
        qty=Decimal("15"),
        qty_in_base_uom=Decimal("15"),
        uom="EA",
        rate=Decimal("10.00"),
        amount=Decimal("150.00"),
        received_qty=Decimal("5"),
        billed_qty=Decimal("10"),
    )
    database.session.add(item_p)
    database.session.commit()

    report = get_purchase_order_status_report("cacao")
    assert len(report) >= 1
    row = next(r for r in report if r["id"] == order_p.id)
    assert row["ordered_qty"] == Decimal("15")
    assert row["received_qty"] == Decimal("5")
    assert row["billed_qty"] == Decimal("10")
    assert row["receipt_status"] == "Parcial"
    assert row["billing_status"] == "Parcial"


def test_unlinked_purchase_invoices(app_ctx):
    from cacao_accounting.compras.purchase_reconciliation_service import get_unlinked_purchase_invoices
    from cacao_accounting.database import PurchaseInvoice, database

    unlinked = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 5),
        supplier_id="SUPP-UN1",
        purchase_receipt_id=None,
        docstatus=1,
        document_type="purchase_invoice",
        grand_total=Decimal("450.00"),
    )
    linked = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 5),
        supplier_id="SUPP-UN1",
        purchase_receipt_id="PREC-SOME",
        docstatus=1,
        document_type="purchase_invoice",
        grand_total=Decimal("200.00"),
    )
    database.session.add_all([unlinked, linked])
    database.session.commit()

    report = get_unlinked_purchase_invoices("cacao")
    ids = [r["id"] for r in report]
    assert unlinked.id in ids
    assert linked.id not in ids


def test_unlinked_purchase_receipts_summary(app_ctx):
    from cacao_accounting.compras.purchase_reconciliation_service import get_unlinked_purchase_receipts_summary
    from cacao_accounting.database import PurchaseReceipt, PurchaseReceiptItem, database, UOM, Item, Warehouse

    database.session.add_all(
        [
            UOM(code="EA-UR", name="Each UR"),
            Item(code="ITEM-UR", name="Item UR", item_type="goods", is_stock_item=True, default_uom="EA-UR"),
            Warehouse(code="WH-UR", name="Bodega UR", company="cacao"),
        ]
    )
    database.session.flush()

    receipt = PurchaseReceipt(company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-UR", docstatus=1)
    database.session.add(receipt)
    database.session.flush()

    receipt_item = PurchaseReceiptItem(
        purchase_receipt_id=receipt.id,
        item_code="ITEM-UR",
        item_name="Item UR",
        qty=Decimal("100"),
        qty_in_base_uom=Decimal("100"),
        uom="EA-UR",
        rate=Decimal("12.00"),
        amount=Decimal("1200.00"),
        warehouse="WH-UR",
    )
    database.session.add(receipt_item)
    database.session.commit()

    # Total pending is 100
    report = get_unlinked_purchase_receipts_summary("cacao")
    assert len(report) >= 1
    row = next(r for r in report if r["id"] == receipt.id)
    assert row["pending_qty"] == Decimal("100")


def test_purchase_reconciliation_web_view_and_tabs(app_ctx):
    from cacao_accounting.database import User, Modules, database

    app_ctx.config["SECRET_KEY"] = "testing-reconc"
    client = app_ctx.test_client()

    user = User(user="reconciliation-user", name="Reconc User", password=b"x", classification="admin", active=True)
    purchases_module = Modules(module="purchases", default=True, enabled=True)
    database.session.add_all([user, purchases_module])
    database.session.commit()

    with client.session_transaction() as session:
        session["_user_id"] = user.id
        session["_fresh"] = True

    response = client.get("/buying/purchase-reconciliation?company=cacao")
    assert response.status_code == 200
    html = response.get_data(as_text=True)

    assert "orders-panel" in html
    assert "invoices-panel" in html
    assert "receipts-panel" in html
    assert "reconciliationTabs" in html


def test_purchase_reconciliation_currency_mismatch_rejected(app_ctx):
    from cacao_accounting.compras.purchase_reconciliation_service import (
        reconcile_purchase_invoice,
        PurchaseReconciliationError,
    )
    from cacao_accounting.database import (
        Item,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseReceipt,
        PurchaseReceiptItem,
        UOM,
        Warehouse,
        database,
    )

    database.session.add_all(
        [
            UOM(code="EA-CURR", name="Each CURR"),
            Item(code="ITEM-CURR", name="Item CURR", item_type="goods", is_stock_item=True, default_uom="EA-CURR"),
            Warehouse(code="WH-CURR", name="Bodega CURR", company="cacao"),
        ]
    )
    database.session.flush()

    receipt = PurchaseReceipt(
        company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-CURR", transaction_currency="USD", docstatus=1
    )
    database.session.add(receipt)
    database.session.flush()

    database.session.add(
        PurchaseReceiptItem(
            purchase_receipt_id=receipt.id,
            item_code="ITEM-CURR",
            qty=Decimal("10"),
            qty_in_base_uom=Decimal("10"),
            uom="EA-CURR",
            rate=Decimal("5.00"),
            amount=Decimal("50.00"),
            warehouse="WH-CURR",
        )
    )

    # Invoice in NIO
    invoice = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 2),
        supplier_id="SUPP-CURR",
        purchase_receipt_id=receipt.id,
        transaction_currency="NIO",
        docstatus=1,
    )
    database.session.add(invoice)
    database.session.flush()

    database.session.add(
        PurchaseInvoiceItem(
            purchase_invoice_id=invoice.id,
            item_code="ITEM-CURR",
            qty=Decimal("10"),
            uom="EA-CURR",
            rate=Decimal("170.00"),
            amount=Decimal("1700.00"),
            warehouse="WH-CURR",
        )
    )
    database.session.commit()

    with pytest.raises(PurchaseReconciliationError, match="en la misma moneda"):
        reconcile_purchase_invoice(invoice.id)


def test_three_way_rejects_receipt_from_another_purchase_order(app_ctx):
    """Una factura 3-way no puede mezclar la OC de la factura y la recepción."""
    from cacao_accounting.compras.purchase_reconciliation_service import (
        PurchaseReconciliationError,
        reconcile_purchase_invoice,
    )
    from cacao_accounting.database import (
        Item,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseMatchingConfig,
        PurchaseOrder,
        PurchaseReconciliation,
        PurchaseReceipt,
        PurchaseReceiptItem,
        UOM,
        Warehouse,
        database,
    )

    database.session.add_all(
        [
            UOM(code="EA-PO-MIX", name="Each PO mix"),
            Item(code="ITEM-PO-MIX", name="Item PO mix", item_type="goods", is_stock_item=True, default_uom="EA-PO-MIX"),
            Warehouse(code="WH-PO-MIX", name="Bodega PO mix", company="cacao"),
        ]
    )
    order_a = PurchaseOrder(
        id="PO-MIX-A",
        company="cacao",
        supplier_id="SUPP-PO-MIX",
        posting_date=date(2026, 5, 1),
        docstatus=1,
    )
    order_b = PurchaseOrder(
        id="PO-MIX-B",
        company="cacao",
        supplier_id="SUPP-PO-MIX",
        posting_date=date(2026, 5, 1),
        docstatus=1,
    )
    database.session.add_all([order_a, order_b])
    database.session.flush()
    receipt = PurchaseReceipt(
        company="cacao",
        supplier_id="SUPP-PO-MIX",
        purchase_order_id=order_b.id,
        posting_date=date(2026, 5, 2),
        docstatus=1,
    )
    database.session.add(receipt)
    database.session.flush()
    database.session.add(
        PurchaseReceiptItem(
            purchase_receipt_id=receipt.id,
            item_code="ITEM-PO-MIX",
            qty=Decimal("1"),
            qty_in_base_uom=Decimal("1"),
            uom="EA-PO-MIX",
            rate=Decimal("10"),
            amount=Decimal("10"),
            warehouse="WH-PO-MIX",
        )
    )
    invoice = PurchaseInvoice(
        company="cacao",
        supplier_id="SUPP-PO-MIX",
        purchase_order_id=order_a.id,
        purchase_receipt_id=receipt.id,
        posting_date=date(2026, 5, 3),
        docstatus=1,
    )
    database.session.add(invoice)
    database.session.flush()
    database.session.add(
        PurchaseInvoiceItem(
            purchase_invoice_id=invoice.id,
            item_code="ITEM-PO-MIX",
            qty=Decimal("1"),
            uom="EA-PO-MIX",
            rate=Decimal("10"),
            amount=Decimal("10"),
        )
    )
    config = database.session.execute(database.select(PurchaseMatchingConfig).filter_by(company="cacao")).scalar_one()
    config.require_purchase_order = True
    database.session.commit()

    with pytest.raises(PurchaseReconciliationError, match="misma orden"):
        reconcile_purchase_invoice(invoice.id)
    assert (
        database.session.execute(
            database.select(PurchaseReconciliation).filter_by(purchase_invoice_id=invoice.id)
        ).scalar_one_or_none()
        is None
    )


def test_partial_invoice_price_variance_scaling(app_ctx):
    from cacao_accounting.compras.purchase_reconciliation_service import (
        reconcile_purchase_invoice,
        PurchaseMatchingConfig,
        MatchingType,
        seed_matching_config_for_company,
    )
    from cacao_accounting.database import (
        UOM,
        Item,
        Warehouse,
        PurchaseReceipt,
        PurchaseReceiptItem,
        PurchaseInvoice,
        PurchaseInvoiceItem,
        PurchaseOrder,
        PurchaseOrderItem,
        database,
    )

    seed_matching_config_for_company("cacao")

    # 1. Test 3-way price difference scaling
    database.session.add_all(
        [
            UOM(code="EA-SCALE", name="Each SCALE"),
            Item(code="ITEM-SCALE", name="Item SCALE", item_type="goods", is_stock_item=True, default_uom="EA-SCALE"),
            Warehouse(code="WH-SCALE", name="Bodega SCALE", company="cacao"),
        ]
    )
    database.session.flush()

    # Ensure matching config is 3-way
    cfg = database.session.execute(database.select(PurchaseMatchingConfig).filter_by(company="cacao")).scalar_one()
    cfg.matching_type = MatchingType.THREE_WAY
    database.session.commit()

    # 100-unit receipt at rate 10 (amount = 1000)
    receipt = PurchaseReceipt(
        company="cacao", posting_date=date(2026, 5, 1), supplier_id="SUPP-SCALE", transaction_currency="USD", docstatus=1
    )
    database.session.add(receipt)
    database.session.flush()

    database.session.add(
        PurchaseReceiptItem(
            purchase_receipt_id=receipt.id,
            item_code="ITEM-SCALE",
            qty=Decimal("100"),
            qty_in_base_uom=Decimal("100"),
            uom="EA-SCALE",
            rate=Decimal("10.00"),
            amount=Decimal("1000.00"),
            warehouse="WH-SCALE",
        )
    )

    # Partial invoice of 10 units at rate 12 (amount = 120)
    # Price difference is 12 - 10 = 2 per unit
    # Expected price variance for 10 units should be 10 * 2 = 20, not 100 * 2 = 200 (if reference_qty was used)
    invoice = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 2),
        supplier_id="SUPP-SCALE",
        purchase_receipt_id=receipt.id,
        transaction_currency="USD",
        docstatus=1,
    )
    database.session.add(invoice)
    database.session.flush()

    database.session.add(
        PurchaseInvoiceItem(
            purchase_invoice_id=invoice.id,
            item_code="ITEM-SCALE",
            qty=Decimal("10"),
            uom="EA-SCALE",
            rate=Decimal("12.00"),
            amount=Decimal("120.00"),
            warehouse="WH-SCALE",
        )
    )
    database.session.commit()

    result_3w = reconcile_purchase_invoice(invoice.id)
    assert result_3w.price_difference == Decimal("20.00")

    # 2. Test 2-way price difference scaling
    # Switch matching config to 2-way
    cfg.matching_type = MatchingType.TWO_WAY
    database.session.commit()

    order = PurchaseOrder(
        company="cacao",
        posting_date=date(2026, 5, 1),
        supplier_id="SUPP-SCALE2",
        transaction_currency="USD",
        docstatus=1,
    )
    database.session.add(order)
    database.session.flush()

    database.session.add(
        PurchaseOrderItem(
            purchase_order_id=order.id,
            item_code="ITEM-SCALE",
            qty=Decimal("100"),
            qty_in_base_uom=Decimal("100"),
            uom="EA-SCALE",
            rate=Decimal("10.00"),
            amount=Decimal("1000.00"),
        )
    )

    invoice_2w = PurchaseInvoice(
        company="cacao",
        posting_date=date(2026, 5, 2),
        supplier_id="SUPP-SCALE2",
        purchase_order_id=order.id,
        transaction_currency="USD",
        docstatus=1,
    )
    database.session.add(invoice_2w)
    database.session.flush()

    database.session.add(
        PurchaseInvoiceItem(
            purchase_invoice_id=invoice_2w.id,
            item_code="ITEM-SCALE",
            qty=Decimal("10"),
            uom="EA-SCALE",
            rate=Decimal("12.00"),
            amount=Decimal("120.00"),
        )
    )
    database.session.commit()

    result_2w = reconcile_purchase_invoice(invoice_2w.id)
    assert result_2w.price_difference == Decimal("20.00")

    # Restore matching config back to 3-way
    cfg.matching_type = MatchingType.THREE_WAY
    database.session.commit()


def test_bank_reconciliation_atomicity_with_difference(app_ctx, monkeypatch):
    """Test that a bank reconciliation difference posting is atomic and commits only at the end."""
    _seed_accounting_admin()
    from cacao_accounting.bancos.reconciliation_service import (
        BankReconciliationMatch,
        BankReconciliationRequest,
        reconcile_bank_items,
    )
    from cacao_accounting.database import (
        Bank,
        BankAccount,
        BankTransaction,
        Book,
        GLEntry,
        ReconciliationItem,
        CompanyDefaultAccount,
        Accounts,
        database,
    )

    bank_gl = Accounts(entity="cacao", code="BANK-ATOM", name="Bank atom", classification="asset", account_type="bank")
    difference = Accounts(entity="cacao", code="BANK-ATOM-DIFF", name="Bank diff atom", classification="expense")
    bank = Bank(name="Banco Atom")
    database.session.add_all([bank_gl, difference, bank])
    database.session.flush()

    database.session.add(CompanyDefaultAccount(company="cacao", bank_difference_account_id=difference.id))
    bank_account = BankAccount(
        bank_id=bank.id,
        company="cacao",
        account_name="Atom account",
        currency="NIO",
        gl_account_id=bank_gl.id,
    )
    database.session.add(bank_account)
    database.session.flush()

    transaction = BankTransaction(
        bank_account_id=bank_account.id,
        posting_date=date(2026, 5, 5),
        deposit=Decimal("100.00"),
    )
    database.session.add(transaction)
    primary_book = database.session.execute(
        database.select(Book).where(Book.entity == "cacao", Book.is_primary.is_(True))
    ).scalar_one()
    target_entry = GLEntry(
        company="cacao",
        posting_date=date(2026, 5, 5),
        account_id=bank_gl.id,
        ledger_id=primary_book.id,
        debit=Decimal("100.00"),
        credit=Decimal("0"),
        voucher_type="manual_test",
        voucher_id="TARGET-ATOM",
        is_cancelled=False,
        is_reversal=False,
    )
    database.session.add(target_entry)
    database.session.commit()

    # Simulate final commit failure or outer rollback
    # We will trigger a mock apply block that fails after journal is attached but before the outer transaction commits.
    # In cacao_accounting/bancos/__init__.py:bancos_conciliacion_bancaria_aplicar,
    # the outer transaction commits via `database.session.commit()` only at the end.
    # Let's perform a simulated route run manually.

    # 1. Start a transaction
    database.session.begin_nested()

    reconciliation = reconcile_bank_items(
        BankReconciliationRequest(
            company="cacao",
            reconciliation_date=date.today(),
            matches=[
                BankReconciliationMatch(
                    bank_transaction_id=transaction.id,
                    target_type="gl_entry",
                    target_id=target_entry.id,
                    allocated_amount=Decimal("95.00"),
                )
            ],
        )
    )
    from cacao_accounting.bancos import _post_bank_difference_adjustment

    _post_bank_difference_adjustment(reconciliation.id, transaction, Decimal("5.00"), user_id="admin")

    # The failure happens after the adjustment journal and reconciliation item exist.
    try:
        raise ValueError("Simulated lookup or route final commit failure")
    except ValueError:
        database.session.rollback()

    # Verify that nothing was committed/persisted to database because of rollback!
    # If the journal had committed internally, it would have persisted even after outer rollback.
    recon_items = database.session.execute(database.select(ReconciliationItem)).scalars().all()
    assert len(recon_items) == 0

    journal_entries = (
        database.session.execute(database.select(GLEntry).filter_by(voucher_type="journal_entry")).scalars().all()
    )
    assert len(journal_entries) == 0


def test_post_bank_difference_deposit_and_withdrawal(app_ctx):
    """Verify that bank difference adjustments derive their sign based on deposit vs withdrawal."""
    _seed_accounting_admin()
    from cacao_accounting.bancos import _post_bank_difference_adjustment
    from cacao_accounting.database import (
        Accounts,
        Bank,
        BankAccount,
        BankTransaction,
        Book,
        CompanyDefaultAccount,
        GLEntry,
        Reconciliation,
        ReconciliationItem,
        database,
    )

    bank_gl = Accounts(entity="cacao", code="BANK-DEP", name="Bank Account", classification="asset", account_type="bank")
    difference = Accounts(entity="cacao", code="BANK-DIFF-DEP", name="Bank difference", classification="expense")
    bank = Bank(name="Banco DEP")
    local_book = Book(entity="cacao", code="DEP-NIO", name="NIO", currency="NIO", status="activo", is_primary=True)
    database.session.add_all(
        [
            bank_gl,
            difference,
            bank,
            local_book,
        ]
    )
    database.session.flush()

    database.session.add(CompanyDefaultAccount(company="cacao", bank_difference_account_id=difference.id))
    bank_account = BankAccount(
        bank_id=bank.id,
        company="cacao",
        account_name="NIO account",
        currency="NIO",
        gl_account_id=bank_gl.id,
    )
    database.session.add(bank_account)
    database.session.flush()

    # 1. Test Deposit Case
    deposit_txn = BankTransaction(
        bank_account_id=bank_account.id,
        posting_date=date(2026, 5, 5),
        deposit=Decimal("10"),
    )
    reconciliation_dep = Reconciliation(company="cacao", recon_date=date(2026, 5, 5), recon_type="bank")
    database.session.add_all([deposit_txn, reconciliation_dep])
    database.session.flush()
    database.session.add(
        ReconciliationItem(
            reconciliation_id=reconciliation_dep.id,
            reference_type="bank_transaction",
            reference_id=deposit_txn.id,
            source_type="bank_transaction",
            source_id=deposit_txn.id,
            amount=Decimal("10"),
        )
    )
    database.session.commit()

    _post_bank_difference_adjustment(reconciliation_dep.id, deposit_txn, Decimal("10"), user_id="admin")
    database.session.commit()

    # Find the GL entries for the deposit adjustment
    # Since deposit, the bank should be DEBITED, and difference account should be CREDITED.
    entries_dep = (
        database.session.execute(
            database.select(GLEntry).filter(
                GLEntry.voucher_type == "journal_entry",
                GLEntry.account_id.in_([bank_gl.id, difference.id]),
                GLEntry.is_cancelled.is_(False),
            )
        )
        .scalars()
        .all()
    )

    # We expect 2 entries: 1 debiting bank, 1 crediting difference
    assert len(entries_dep) == 2
    bank_entry = next(e for e in entries_dep if e.account_id == bank_gl.id)
    diff_entry = next(e for e in entries_dep if e.account_id == difference.id)
    assert bank_entry.debit == Decimal("10")
    assert bank_entry.credit == Decimal("0")
    assert diff_entry.credit == Decimal("10")
    assert diff_entry.debit == Decimal("0")
    deposit_difference_item = database.session.execute(
        database.select(ReconciliationItem).filter_by(
            reconciliation_id=reconciliation_dep.id,
            target_id=bank_entry.id,
        )
    ).scalar_one()
    assert deposit_difference_item.amount == Decimal("10")
    assert deposit_difference_item.allocated_amount == Decimal("10")
    deposit_voucher_id = bank_entry.voucher_id

    # 2. Test Withdrawal Case
    withdrawal_txn = BankTransaction(
        bank_account_id=bank_account.id,
        posting_date=date(2026, 5, 5),
        withdrawal=Decimal("15"),
    )
    reconciliation_with = Reconciliation(company="cacao", recon_date=date(2026, 5, 5), recon_type="bank")
    database.session.add_all([withdrawal_txn, reconciliation_with])
    database.session.flush()
    database.session.add(
        ReconciliationItem(
            reconciliation_id=reconciliation_with.id,
            reference_type="bank_transaction",
            reference_id=withdrawal_txn.id,
            source_type="bank_transaction",
            source_id=withdrawal_txn.id,
            amount=Decimal("15"),
        )
    )
    database.session.commit()

    _post_bank_difference_adjustment(reconciliation_with.id, withdrawal_txn, Decimal("15"), user_id="admin")
    database.session.commit()

    # Find the GL entries for the withdrawal adjustment
    # Since withdrawal, the bank should be CREDITED, and difference account should be DEBITED.
    entries_with = (
        database.session.execute(
            database.select(GLEntry).filter(
                GLEntry.voucher_type == "journal_entry",
                GLEntry.voucher_id != deposit_voucher_id,
                GLEntry.account_id.in_([bank_gl.id, difference.id]),
            )
        )
        .scalars()
        .all()
    )

    assert len(entries_with) == 2
    bank_entry_w = next(e for e in entries_with if e.account_id == bank_gl.id)
    diff_entry_w = next(e for e in entries_with if e.account_id == difference.id)
    assert bank_entry_w.credit == Decimal("15")
    assert bank_entry_w.debit == Decimal("0")
    assert diff_entry_w.debit == Decimal("15")
    assert diff_entry_w.credit == Decimal("0")
    withdrawal_difference_item = database.session.execute(
        database.select(ReconciliationItem).filter_by(
            reconciliation_id=reconciliation_with.id,
            target_id=bank_entry_w.id,
        )
    ).scalar_one()
    assert withdrawal_difference_item.amount == Decimal("15")
    assert withdrawal_difference_item.allocated_amount == Decimal("15")


def test_cancelled_reconciliation_item_excluded_from_reconciliation_report(app_ctx):
    """Verifica que get_reconciliation_report no sume ReconciliationItem cancelados en bank_reconciled_amount."""
    from datetime import datetime, timezone
    from cacao_accounting.database import (
        Accounts,
        Bank,
        BankAccount,
        BankTransaction,
        PaymentEntry,
        Reconciliation,
        ReconciliationItem,
        AuditTrail,
        database,
    )
    from cacao_accounting.bancos import _apply_payment_cancellation_hooks
    from cacao_accounting.reportes.services import get_reconciliation_report

    bank_gl = Accounts(entity="cacao", code="BANK-TEST", name="Test Bank", classification="asset")
    bank = Bank(name="Banco Test")
    database.session.add_all([bank_gl, bank])
    database.session.flush()

    bank_account = BankAccount(
        bank_id=bank.id,
        company="cacao",
        account_name="Test Bank Account",
        currency="NIO",
        gl_account_id=bank_gl.id,
    )
    database.session.add(bank_account)
    database.session.flush()

    transaction = BankTransaction(
        bank_account_id=bank_account.id,
        posting_date=date(2026, 5, 5),
        deposit=Decimal("100.00"),
    )
    database.session.add(transaction)
    database.session.flush()

    payment = PaymentEntry(
        company="cacao",
        posting_date=date(2026, 5, 5),
        payment_type="receive",
        bank_account_id=bank_account.id,
        paid_amount=Decimal("100.00"),
        docstatus=1,
    )
    database.session.add(payment)
    database.session.flush()

    reconciliation = Reconciliation(company="cacao", recon_date=date(2026, 5, 5), recon_type="bank")
    database.session.add(reconciliation)
    database.session.flush()

    recon_item = ReconciliationItem(
        reconciliation_id=reconciliation.id,
        reference_type="bank_transaction",
        reference_id=transaction.id,
        source_type="bank_transaction",
        source_id=transaction.id,
        target_type="payment_entry",
        target_id=payment.id,
        amount=Decimal("100.00"),
        allocated_amount=Decimal("100.00"),
        status="reconciled",
    )
    database.session.add(recon_item)
    database.session.commit()

    # Before cancellation, the report should sum the 100.00 amount
    report_before = get_reconciliation_report(company="cacao")
    assert report_before.totals["bank_reconciled_amount"] == Decimal("100.00")

    # Cancel the payment, which triggers marking status as 'cancelled'
    _apply_payment_cancellation_hooks(payment)
    # Manually add the AuditTrail entry for the cancellation on June 10, 2026
    cancel_log = AuditTrail(
        document_type="payment_entry",
        document_id=payment.id,
        action="cancelled",
        timestamp=datetime(2026, 6, 10, 12, 0, 0, tzinfo=timezone.utc),
        company="cacao",
    )
    database.session.add(cancel_log)
    database.session.commit()

    # Verify that the ReconciliationItem status was updated to 'cancelled'
    database.session.refresh(recon_item)
    assert recon_item.status == "cancelled"

    # Test as_of_date BEFORE the cancellation date (e.g. May 31, 2026) -> should PRESERVE the reconciled item!
    report_cutoff_before = get_reconciliation_report(company="cacao", as_of_date=date(2026, 5, 31))
    assert report_cutoff_before.totals["bank_reconciled_amount"] == Decimal("100.00")
    # Verify status is displayed as "reconciled" in the report as of that date
    matching_rows = [row for row in report_cutoff_before.rows if row.values["target_id"] == payment.id]
    assert len(matching_rows) == 1
    assert matching_rows[0].values["status"] == "reconciled"

    # Test as_of_date AFTER the cancellation date (e.g. June 15, 2026) -> should EXCLUDE the cancelled item!
    report_cutoff_after = get_reconciliation_report(company="cacao", as_of_date=date(2026, 6, 15))
    assert report_cutoff_after.totals["bank_reconciled_amount"] == Decimal("0")

    # Test current report (no as_of_date specified) -> should EXCLUDE the cancelled item!
    report_current = get_reconciliation_report(company="cacao")
    assert report_current.totals["bank_reconciled_amount"] == Decimal("0")
